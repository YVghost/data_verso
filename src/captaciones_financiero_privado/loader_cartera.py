"""
ETL completo: Colocaciones / Cartera — Bancos Privados
Flujo: Excel → stg_cartera → cartera

Columnas extra vs depósitos:
  tipo_colocacion    → del nombre del archivo  (banca_privada_consumo, etc.)
  subtipo_colocacion → del nombre de la hoja   (consumo_prioritario, inmobiliario, etc.)

Métricas de cartera (distintas a depósitos):
  por_vencer | no_devenga_intereses | vencida | total_cartera

Variantes de archivo/hoja:
  2017–2020 reporte : ENTIDAD en col 0, fechas como datetime, solo por_vencer
  2021+    reporte  : col 0 = 'N' (subtipo), ENTIDAD en col 1, fechas como string dd/mm/aaaa
  2021+    tabular  : BASE B PRIVADA … → 4 métricas completas
"""

import sys
import re
import hashlib
import logging
from collections import defaultdict
from pathlib import Path
from typing import Optional
import datetime as dt_mod

import pandas as pd
import openpyxl
from sqlalchemy import text, inspect as sa_inspect

sys.path.append(str(Path(__file__).resolve().parents[2]))
from utils.base_engine import get_master_engine

STG_TABLE   = "stg_cartera"
FINAL_TABLE = "cartera"

logging.basicConfig(level=logging.INFO, format="[%(levelname)s] %(message)s")
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Mapeos de normalización
# ---------------------------------------------------------------------------

_TIPO_COLOCACION_MAP = {
    "comercial":   "banca_privada_comercial",
    "consumo":     "banca_privada_consumo",
    "educativo":   "banca_privada_educativo",
    "microcredito":"banca_privada_microcredito",
    "microcrédito":"banca_privada_microcredito",
    "vivienda":    "banca_privada_vivienda",
    "productivo":  "banca_privada_productivo",
}

# Palabras clave en nombre de hoja → subtipo normalizado
# Evaluados en orden; el primero que encaje gana
_SUBTIPO_RULES = [
    (("comercial", "prioritario"),  "comercial_prioritario"),
    (("comercial", "ordinario"),    "comercial_ordinario"),
    (("consumo",   "prioritario"),  "consumo_prioritario"),
    (("consumo",   "ordinario"),    "consumo_ordinario"),
    (("vivienda",  "interes"),      "vivienda_interes_publico"),
    (("inmobiliario",),             "inmobiliario"),
    (("consumo",),                  "consumo"),
    (("educativo",),                "educativo"),
    (("microcredito",),             "microcredito"),
    (("microcrédito",),             "microcredito"),
    (("productivo",),               "productivo"),
]


def _tipo_from_filename(filename: str) -> str:
    fn = filename.lower()
    for kw, tipo in _TIPO_COLOCACION_MAP.items():
        if kw in fn:
            return tipo
    return "banca_privada_desconocido"


def _subtipo_from_sheetname(sheet: str) -> str:
    sh = sheet.lower()
    for keywords, subtipo in _SUBTIPO_RULES:
        if all(kw in sh for kw in keywords):
            return subtipo
    return re.sub(r"\s+", "_", sheet.strip().lower())


# ---------------------------------------------------------------------------
# Punto de entrada principal
# ---------------------------------------------------------------------------

def load(excel_paths: list) -> None:
    """
    Procesa todos los Excels de cartera y los carga en stg_cartera → cartera.
    Acepta lista de Path (salida de bot.download_and_extract["cartera"]).
    """
    if not excel_paths:
        log.info("[cartera] Sin archivos para procesar.")
        return

    engine = get_master_engine()
    _ensure_tables(engine)
    _ensure_columns(engine)

    total_stg = 0
    for path in excel_paths:
        path = Path(path)
        if not path.exists() or path.name.startswith("~$"):
            continue
        log.info(f"[cartera] Procesando: {path.name}")
        try:
            rows_loaded = _process_file(path, engine)
            total_stg += rows_loaded
        except Exception as ex:
            log.error(f"[cartera] Error en {path.name}: {ex}")

    log.info(f"[cartera] Total filas cargadas a staging: {total_stg}")
    _consolidate(engine)


# ---------------------------------------------------------------------------
# Procesamiento de un archivo Excel
# ---------------------------------------------------------------------------

def _process_file(path: Path, engine) -> int:
    """
    Un archivo puede tener varias hojas reporte y/o varias hojas tabular.
    Cada hoja genera su propio DataFrame etiquetado con subtipo_colocacion.
    Prioridad por hoja: si existe tabular para un subtipo, descarta el reporte
    del mismo subtipo.
    """
    tipo_col = _tipo_from_filename(path.name)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    tabular_subtipos: dict[str, pd.DataFrame] = {}
    reporte_subtipos: dict[str, pd.DataFrame] = {}

    for sh in wb.sheetnames:
        ws   = wb[sh]
        rows = list(ws.iter_rows(values_only=True))
        subtipo  = _subtipo_from_sheetname(sh)
        sh_type  = _detect_sheet_type(rows)
        log.info(f"  Hoja '{sh}': tipo={sh_type}  subtipo={subtipo}")

        if sh_type == "tabular":
            df = _parse_tabular(rows, path.name, sh, tipo_col, subtipo)
            if df is not None and not df.empty:
                tabular_subtipos[subtipo] = df

        elif sh_type in ("reporte_viejo", "reporte_nuevo"):
            df = _parse_reporte(rows, path.name, sh, tipo_col, subtipo, sh_type)
            if df is not None and not df.empty:
                reporte_subtipos[subtipo] = df

    wb.close()

    # Combinar: tabular tiene prioridad por subtipo.
    # Si también existe reporte para el mismo subtipo, inyectar morosidad del reporte
    # en las filas tabular (tabular no incluye morosidad en sus columnas).
    _MERGE_KEYS = ["fecha", "entidad", "provincia", "canton"]
    frames = []
    for subtipo, df_tab in tabular_subtipos.items():
        if subtipo in reporte_subtipos:
            df_rep = reporte_subtipos[subtipo]
            if "morosidad" in df_rep.columns:
                df_mor = (
                    df_rep[_MERGE_KEYS + ["morosidad"]]
                    .dropna(subset=["morosidad"])
                    .drop_duplicates(subset=_MERGE_KEYS)
                )
                df_tab = df_tab.merge(df_mor, on=_MERGE_KEYS, how="left",
                                      suffixes=("", "_rep"))
                if "morosidad_rep" in df_tab.columns:
                    df_tab["morosidad"] = df_tab["morosidad_rep"]
                    df_tab.drop(columns=["morosidad_rep"], inplace=True)
        frames.append(df_tab)
    for subtipo, df in reporte_subtipos.items():
        if subtipo not in tabular_subtipos:
            frames.append(df)

    if not frames:
        log.warning(f"  Sin datos utilizables en {path.name}")
        return 0

    combined = pd.concat(frames, ignore_index=True)
    combined = _validate(combined)
    log.info(f"  Filas válidas: {len(combined)}")
    return _load_staging(combined, engine)


# ---------------------------------------------------------------------------
# Detección de tipo de hoja
# ---------------------------------------------------------------------------

def _detect_sheet_type(rows: list) -> Optional[str]:
    """
    tabular       → fila con 'POR VENCER' y 'NO DEVENGA INTERESES'
    reporte_nuevo → fila con 'N','ENTIDAD','PROVINCIA','CANTON' + fechas como string
    reporte_viejo → fila con 'ENTIDAD' en col 0 + 'CANTON' + datetime en cols siguientes
    """
    for row in rows[:15]:
        vals_upper = [str(v).strip().upper() for v in row if v is not None and str(v).strip()]

        # Tabular (2021+)
        if "POR VENCER" in vals_upper and "NO DEVENGA INTERESES" in vals_upper:
            return "tabular"

        # Reporte nuevo (2021+): col 0 o 1 tiene 'N', fechas como string dd/mm/aaaa
        if "ENTIDAD" in vals_upper and "CANTON" in vals_upper:
            has_str_dates = any(
                isinstance(v, str) and re.match(r"\d{1,2}/\d{1,2}/\d{4}", str(v).strip())
                for v in row if v is not None
            )
            has_dt_dates = any(
                isinstance(v, (dt_mod.datetime, dt_mod.date))
                for v in row
            )
            if has_str_dates:
                return "reporte_nuevo"
            if has_dt_dates:
                return "reporte_viejo"

    return None


# ---------------------------------------------------------------------------
# Parser hoja TABULAR (BASE B PRIVADA …, 2021+)
# ---------------------------------------------------------------------------

def _parse_tabular(rows, archivo, hoja, tipo_col, subtipo) -> Optional[pd.DataFrame]:
    """
    Estructura: [None, subtipo_str?, FECHA, ENTIDAD, PROVINCIA, CANTON,
                 POR VENCER, NO DEVENGA INTERESES, VENCIDA, TOTAL CARTERA]
    """
    header_idx = None
    for i, row in enumerate(rows[:15]):
        vals = [str(v).strip().upper() for v in row if v is not None and str(v).strip()]
        if "FECHA" in vals and "POR VENCER" in vals:
            header_idx = i
            break

    if header_idx is None:
        log.warning(f"  [{hoja}] No se encontró encabezado tabular.")
        return None

    raw_headers = rows[header_idx]
    data_rows   = rows[header_idx + 1:]

    col_map = {
        "FECHA":                 "fecha",
        "ENTIDAD":               "entidad",
        "PROVINCIA":             "provincia",
        "CANTON":                "canton",
        "POR VENCER":            "por_vencer",
        "NO DEVENGA INTERESES":  "no_devenga_intereses",
        "VENCIDA":               "vencida",
        "TOTAL CARTERA":         "total_cartera",
        "TOTAL SALDO":           "total_cartera",
        "TOTAL":                 "total_cartera",
        "MOROSIDAD":             "morosidad",
    }
    col_pos = {}
    for j, h in enumerate(raw_headers):
        if h is None:
            continue
        hn = str(h).strip().upper()
        if hn in col_map:
            col_pos[col_map[hn]] = j

    required = {"fecha", "entidad", "provincia", "canton"}
    if not required.issubset(col_pos):
        log.warning(f"  [{hoja}] Faltan columnas: {required - set(col_pos)}")
        return None

    records = []
    current_subtipo = subtipo
    for row in data_rows:
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        rec = {col: (row[pos] if pos < len(row) else None) for col, pos in col_pos.items()}

        # Col 1 puede contener el subtipo explícito (ej. "INMOBILIARIO")
        if len(row) > 1 and row[1] is not None and str(row[1]).strip():
            current_subtipo = _subtipo_from_sheetname(str(row[1]).strip())
        rec["subtipo_colocacion"] = current_subtipo
        records.append(rec)

    df = pd.DataFrame(records)
    df["tipo_colocacion"] = tipo_col
    df["tipo_hoja"]       = "tabular"
    df["archivo_origen"]  = archivo
    df["hoja_origen"]     = hoja
    df["fecha_carga"]     = pd.Timestamp.now()
    return _normalize_columns(df)


# ---------------------------------------------------------------------------
# Parser hoja REPORTE (ambas variantes)
# ---------------------------------------------------------------------------

def _parse_reporte(rows, archivo, hoja, tipo_col, subtipo, sh_type) -> Optional[pd.DataFrame]:
    """
    reporte_viejo (2017-2020): ENTIDAD col 0, fechas datetime, solo por_vencer
    reporte_nuevo (2021+)    : fila N-1 = grupos de métrica, fila N = N/ENTIDAD/CANTON/fechas repetidas
    Cada grupo de métrica (CARTERA POR VENCER, NO DEVENGA INTERESES, VENCIDA, SALDO TOTAL,
    MOROSIDAD) repite el mismo set de fechas. El parser genera una fila por fecha con las
    5 métricas ya resueltas.
    """
    # 1. Localizar fila de encabezado (tiene ENTIDAD y CANTON)
    header_idx = None
    for i, row in enumerate(rows[:15]):
        vals_upper = [str(v).strip().upper() for v in row if v is not None and str(v).strip()]
        if "ENTIDAD" in vals_upper and "CANTON" in vals_upper:
            header_idx = i
            break

    if header_idx is None:
        log.warning(f"  [{hoja}] No se encontró encabezado de reporte.")
        return None

    # 2. Detectar fila de grupos de métrica (inmediatamente anterior al encabezado)
    _METRIC_NAMES = {
        "CARTERA POR VENCER":               "por_vencer",
        "CARTERA QUE NO DEVENGA INTERESES": "no_devenga_intereses",
        "CARTERA VENCIDA":                  "vencida",
        "SALDO TOTAL":                      "total_cartera",
        "MOROSIDAD":                        "morosidad",
    }
    col_metric: dict[int, str] = {}   # col_idx → métrica

    if header_idx > 0:
        metric_row = rows[header_idx - 1]
        current_metric = None
        for j, cell in enumerate(metric_row):
            if cell is not None:
                cs = str(cell).strip().upper()
                for mk, mv in _METRIC_NAMES.items():
                    if mk in cs:
                        current_metric = mv
                        break
            if current_metric:
                col_metric[j] = current_metric

    # 3. Parsear encabezado: cols descriptivas + cols de fecha→métrica
    header_row = rows[header_idx]
    data_rows  = rows[header_idx + 1:]

    DESC_MAP = {"ENTIDAD": "entidad", "PROVINCIA": "provincia", "CANTON": "canton"}
    desc_cols: dict[str, int] = {}
    subtipo_col_idx: Optional[int] = None
    date_col_list: list[tuple[int, dt_mod.date, str]] = []  # (col_idx, fecha, métrica)

    for j, cell in enumerate(header_row):
        if cell is None:
            continue
        cs = str(cell).strip().upper()
        if cs in DESC_MAP:
            desc_cols[DESC_MAP[cs]] = j
        elif cs == "N":
            subtipo_col_idx = j
        elif isinstance(cell, (dt_mod.datetime, dt_mod.date)):
            d = cell.date() if isinstance(cell, dt_mod.datetime) else cell
            date_col_list.append((j, d, col_metric.get(j, "por_vencer")))
        elif isinstance(cell, str):
            m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", cell.strip())
            if m:
                part1, part2, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
                if part1 > 12:
                    d = dt_mod.date(year, part2, part1)
                elif part2 > 12:
                    d = dt_mod.date(year, part1, part2)
                else:
                    d = dt_mod.date(year, part2, part1)
                date_col_list.append((j, d, col_metric.get(j, "por_vencer")))

    if not date_col_list:
        log.warning(f"  [{hoja}] No se encontraron columnas de fecha.")
        return None

    # Agrupar cols por fecha → {métrica: col_idx}
    date_groups: dict[dt_mod.date, dict[str, int]] = defaultdict(dict)
    for col_idx, d, metric in date_col_list:
        date_groups[d][metric] = col_idx

    all_col_idxs = [c for c, _, _ in date_col_list] + list(desc_cols.values())
    if subtipo_col_idx is not None:
        all_col_idxs.append(subtipo_col_idx)
    max_col = max(all_col_idxs) + 1

    # 4. Leer filas de datos → registros anchos (uno por fila fuente)
    wide_records = []
    for row in data_rows:
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        padded = list(row) + [None] * max_col
        rec = {col: padded[idx] for col, idx in desc_cols.items()}

        if subtipo_col_idx is not None:
            val = padded[subtipo_col_idx]
            rec["_subtipo_raw"] = str(val).strip() if val is not None and str(val).strip() else None
        else:
            rec["_subtipo_raw"] = None

        date_vals: dict[dt_mod.date, dict[str, object]] = {}
        for d, metric_map in date_groups.items():
            date_vals[d] = {mf: padded[ci] for mf, ci in metric_map.items()}
        rec["_date_vals"] = date_vals
        wide_records.append(rec)

    if not wide_records:
        return None

    df_wide = pd.DataFrame(wide_records)

    # 5. Forward-fill jerarquía visual
    for col in ["entidad", "provincia", "canton", "_subtipo_raw"]:
        if col in df_wide.columns:
            df_wide[col] = df_wide[col].ffill()

    # 6. Excluir filas de totales
    total_mask = df_wide.apply(_is_total_row, axis=1)
    df_wide = df_wide[~total_mask].copy()
    if total_mask.sum():
        log.info(f"  [{hoja}] Descartadas {total_mask.sum()} filas de totales.")

    # 7. Melt: una fila por fecha con todas las métricas
    ALL_METRICS = ["por_vencer", "no_devenga_intereses", "vencida", "total_cartera", "morosidad"]
    rows_long = []
    for _, row in df_wide.iterrows():
        raw_subtipo = row.get("_subtipo_raw")
        effective_subtipo = _subtipo_from_sheetname(str(raw_subtipo)) if raw_subtipo else subtipo
        for d, metric_vals in row["_date_vals"].items():
            long_rec = {
                "fecha":              d,
                "entidad":            row.get("entidad"),
                "provincia":          row.get("provincia"),
                "canton":             row.get("canton"),
                "subtipo_colocacion": effective_subtipo,
            }
            for mf in ALL_METRICS:
                long_rec[mf] = metric_vals.get(mf)
            rows_long.append(long_rec)

    if not rows_long:
        return None

    df_long = pd.DataFrame(rows_long)
    df_long["tipo_colocacion"] = tipo_col
    df_long["tipo_hoja"]       = "reporte"
    df_long["archivo_origen"]  = archivo
    df_long["hoja_origen"]     = hoja
    df_long["fecha_carga"]     = pd.Timestamp.now()
    return _normalize_columns(df_long)


# ---------------------------------------------------------------------------
# Normalización compartida
# ---------------------------------------------------------------------------

def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    # fecha → date
    if "fecha" in df.columns:
        def _to_date(v):
            if isinstance(v, dt_mod.datetime):
                return v.date()
            if isinstance(v, dt_mod.date):
                return v
            try:
                return pd.to_datetime(v, dayfirst=True).date()
            except Exception:
                return None
        df["fecha"] = df["fecha"].apply(_to_date)

    # texto → strip + upper
    for col in ["entidad", "provincia", "canton", "tipo_colocacion", "subtipo_colocacion"]:
        if col in df.columns:
            df[col] = df[col].apply(
                lambda v: str(v).strip().upper() if v is not None and str(v).strip() else None
            )

    # numéricos
    for col in ["por_vencer", "no_devenga_intereses", "vencida", "total_cartera", "morosidad"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    df["hash_registro"] = df.apply(_compute_hash, axis=1)
    return df


def _is_total_row(row: pd.Series) -> bool:
    for col in ["entidad", "provincia", "canton"]:
        if col in row and row[col] is not None:
            if str(row[col]).strip().upper().startswith("TOTAL"):
                return True
    return False


def _compute_hash(row: pd.Series) -> str:
    parts = [
        str(row.get("fecha",              "") or ""),
        str(row.get("entidad",            "") or ""),
        str(row.get("provincia",          "") or ""),
        str(row.get("canton",             "") or ""),
        str(row.get("tipo_colocacion",    "") or ""),
        str(row.get("subtipo_colocacion", "") or ""),
    ]
    return hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()


# ---------------------------------------------------------------------------
# Validación
# ---------------------------------------------------------------------------

def _validate(df: pd.DataFrame) -> pd.DataFrame:
    initial = len(df)
    mask = (
        df["fecha"].notna() &
        df["entidad"].notna() &
        df["tipo_colocacion"].notna() &
        df["subtipo_colocacion"].notna()
    )
    df = df[mask].copy()
    dropped = initial - len(df)
    if dropped:
        log.info(f"  Descartadas por calidad: {dropped}")
    return df


# ---------------------------------------------------------------------------
# Carga a staging
# ---------------------------------------------------------------------------

_STG_COLS = [
    "archivo_origen", "hoja_origen", "tipo_hoja",
    "tipo_colocacion", "subtipo_colocacion",
    "fecha", "entidad", "provincia", "canton",
    "por_vencer", "no_devenga_intereses", "vencida", "total_cartera", "morosidad",
    "hash_registro", "fecha_carga",
]

def _load_staging(df: pd.DataFrame, engine) -> int:
    for col in _STG_COLS:
        if col not in df.columns:
            df[col] = None
    df = df[_STG_COLS].drop_duplicates(subset=["hash_registro"])

    with engine.connect() as conn:
        existing = set(
            pd.read_sql(f"SELECT hash_registro FROM {STG_TABLE}", conn)["hash_registro"].tolist()
        )
    new_df = df[~df["hash_registro"].isin(existing)]

    if new_df.empty:
        log.info(f"  [staging] Sin registros nuevos.")
        return 0

    new_df.to_sql(STG_TABLE, engine, if_exists="append", index=False)
    log.info(f"  [staging] {len(new_df)} filas insertadas.")
    return len(new_df)


# ---------------------------------------------------------------------------
# Consolidación staging → final
# ---------------------------------------------------------------------------

def _consolidate(engine) -> None:
    """
    Prioridad: tabular > reporte para el mismo subtipo/fecha/entidad/provincia/canton.
    """
    sql = text(f"""
        INSERT INTO {FINAL_TABLE} (
            tipo_colocacion, subtipo_colocacion,
            fecha, entidad, provincia, canton,
            por_vencer, no_devenga_intereses, vencida, total_cartera, morosidad,
            archivo_origen, hoja_origen, tipo_hoja,
            hash_registro, fecha_carga
        )
        SELECT
            s.tipo_colocacion, s.subtipo_colocacion,
            s.fecha, s.entidad, s.provincia, s.canton,
            s.por_vencer, s.no_devenga_intereses, s.vencida, s.total_cartera, s.morosidad,
            s.archivo_origen, s.hoja_origen, s.tipo_hoja,
            s.hash_registro, s.fecha_carga
        FROM {STG_TABLE} s
        WHERE s.hash_registro NOT IN (SELECT hash_registro FROM {FINAL_TABLE})
          AND s.tipo_hoja = 'tabular'

        UNION ALL

        SELECT
            s.tipo_colocacion, s.subtipo_colocacion,
            s.fecha, s.entidad, s.provincia, s.canton,
            s.por_vencer, s.no_devenga_intereses, s.vencida, s.total_cartera, s.morosidad,
            s.archivo_origen, s.hoja_origen, s.tipo_hoja,
            s.hash_registro, s.fecha_carga
        FROM {STG_TABLE} s
        WHERE s.hash_registro NOT IN (SELECT hash_registro FROM {FINAL_TABLE})
          AND s.tipo_hoja = 'reporte'
          AND NOT EXISTS (
              SELECT 1 FROM {STG_TABLE} s2
              WHERE s2.tipo_hoja          = 'tabular'
                AND s2.fecha              = s.fecha
                AND s2.entidad            = s.entidad
                AND COALESCE(s2.provincia,'') = COALESCE(s.provincia,'')
                AND COALESCE(s2.canton,'')    = COALESCE(s.canton,'')
                AND s2.tipo_colocacion    = s.tipo_colocacion
                AND s2.subtipo_colocacion = s.subtipo_colocacion
          )
    """)
    with engine.begin() as conn:
        result = conn.execute(sql)
        inserted = result.rowcount if result.rowcount >= 0 else "?"
    log.info(f"[cartera] [{FINAL_TABLE}] {inserted} filas consolidadas.")


# ---------------------------------------------------------------------------
# Creación / migración de tablas
# ---------------------------------------------------------------------------

def _ensure_columns(engine) -> None:
    """Agrega columnas que faltan en tablas ya existentes (migración no destructiva)."""
    inspector = sa_inspect(engine)
    new_cols = {"morosidad": "FLOAT"}
    for table in [STG_TABLE, FINAL_TABLE]:
        if not inspector.has_table(table):
            continue
        existing = {c["name"].lower() for c in inspector.get_columns(table)}
        for col_name, col_type in new_cols.items():
            if col_name not in existing:
                with engine.begin() as conn:
                    conn.execute(text(f"ALTER TABLE {table} ADD {col_name} {col_type}"))
                log.info(f"[cartera] Columna '{col_name}' añadida a {table}.")


def _ensure_tables(engine) -> None:
    inspector = sa_inspect(engine)

    if not inspector.has_table(STG_TABLE):
        with engine.begin() as conn:
            conn.execute(text(f"""
                CREATE TABLE {STG_TABLE} (
                    id_stg                INT IDENTITY(1,1) PRIMARY KEY,
                    archivo_origen        NVARCHAR(255),
                    hoja_origen           NVARCHAR(100),
                    tipo_hoja             NVARCHAR(20),
                    tipo_colocacion       NVARCHAR(50),
                    subtipo_colocacion    NVARCHAR(50),
                    fecha                 DATE,
                    entidad               NVARCHAR(200),
                    provincia             NVARCHAR(100),
                    canton                NVARCHAR(100),
                    por_vencer            FLOAT,
                    no_devenga_intereses  FLOAT,
                    vencida               FLOAT,
                    total_cartera         FLOAT,
                    morosidad             FLOAT,
                    hash_registro         NVARCHAR(64),
                    fecha_carga           DATETIME
                )
            """))
        log.info(f"[cartera] Tabla {STG_TABLE} creada.")

    if not inspector.has_table(FINAL_TABLE):
        with engine.begin() as conn:
            conn.execute(text(f"""
                CREATE TABLE {FINAL_TABLE} (
                    id                    INT IDENTITY(1,1) NOT NULL,
                    tipo_colocacion       NVARCHAR(50)  NOT NULL,
                    subtipo_colocacion    NVARCHAR(50)  NOT NULL,
                    fecha                 DATE          NOT NULL,
                    entidad               NVARCHAR(200) NOT NULL,
                    provincia             NVARCHAR(100),
                    canton                NVARCHAR(100),
                    por_vencer            FLOAT,
                    no_devenga_intereses  FLOAT,
                    vencida               FLOAT,
                    total_cartera         FLOAT,
                    morosidad             FLOAT,
                    archivo_origen        NVARCHAR(255),
                    hoja_origen           NVARCHAR(100),
                    tipo_hoja             NVARCHAR(20),
                    hash_registro         NVARCHAR(64),
                    fecha_carga           DATETIME,
                    CONSTRAINT PK_{FINAL_TABLE} PRIMARY KEY NONCLUSTERED (id),
                    CONSTRAINT UQ_{FINAL_TABLE}_hash UNIQUE NONCLUSTERED (hash_registro)
                )
            """))
            conn.execute(text(f"""
                CREATE CLUSTERED INDEX CIX_{FINAL_TABLE}_fecha_prov_canton
                    ON {FINAL_TABLE} (fecha, provincia, canton)
            """))
        log.info(f"[cartera] Tabla {FINAL_TABLE} creada (orden: fecha → provincia → canton).")
