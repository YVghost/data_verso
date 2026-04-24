"""
ETL completo: Captaciones / Depósitos — Bancos Privados
Flujo: Excel → stg_captaciones → captaciones

Variantes de archivo por año:
  2017–2020 : 1 hoja reporte (BANCOS PRIVADOS)  → solo NUMERO DE CLIENTES, formato ancho
  2021–2025 : 2 hojas → reporte (BANCA PRIVADA) + tabular (BASE BANCA PRIVADA)
              Prioridad: tabular (tiene 3 métricas + cuenta)

Detección automática de tipo de hoja por contenido, no por nombre.
"""

import sys
import hashlib
import logging
from datetime import date as date_type
from pathlib import Path
from typing import Optional

import pandas as pd
import openpyxl
from sqlalchemy import text, inspect as sa_inspect

sys.path.append(str(Path(__file__).resolve().parents[2]))
from utils.base_engine import get_master_engine

STG_TABLE   = "stg_captaciones"
FINAL_TABLE = "captaciones"

logging.basicConfig(level=logging.INFO, format="[%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

# Palabras que indican fila de total/subtotal (case-insensitive, al inicio del texto)
_TOTAL_PREFIXES = ("total ", "total\t", "subtotal")


# ---------------------------------------------------------------------------
# Punto de entrada principal
# ---------------------------------------------------------------------------

def load(excel_paths: list) -> None:
    """
    Procesa todos los Excels de depósitos y los carga en stg_captaciones → captaciones.
    Acepta lista de Path (salida de bot.download_and_extract["depositos"]).
    """
    if not excel_paths:
        log.info("[depositos] Sin archivos para procesar.")
        return

    engine = get_master_engine()
    _ensure_tables(engine)

    total_stg = 0
    for path in excel_paths:
        path = Path(path)
        if not path.exists() or path.name.startswith("~$"):
            continue
        log.info(f"[depositos] Procesando: {path.name}")
        try:
            rows_loaded = _process_file(path, engine)
            total_stg += rows_loaded
        except Exception as ex:
            log.error(f"[depositos] Error en {path.name}: {ex}")

    log.info(f"[depositos] Total filas cargadas a staging: {total_stg}")
    _consolidate(engine)


# ---------------------------------------------------------------------------
# Procesamiento de un archivo Excel
# ---------------------------------------------------------------------------

def _process_file(path: Path, engine) -> int:
    """
    Detecta las hojas del Excel, las transforma y carga a staging.
    Retorna el número de filas nuevas insertadas en staging.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = wb.sheetnames

    tabular_df  = None
    reporte_df  = None

    for sh in sheets:
        ws = wb[sh]
        rows = list(ws.iter_rows(values_only=True))
        wb_type = _detect_sheet_type(rows)
        log.info(f"  Hoja '{sh}': tipo={wb_type}")

        if wb_type == "tabular":
            tabular_df = _parse_tabular(rows, path.name, sh)
        elif wb_type == "reporte":
            reporte_df = _parse_reporte(rows, path.name, sh)

    wb.close()

    # Prioridad: tabular > reporte
    if tabular_df is not None and not tabular_df.empty:
        df = tabular_df
    elif reporte_df is not None and not reporte_df.empty:
        df = reporte_df
    else:
        log.warning(f"  Sin datos utilizables en {path.name}")
        return 0

    df = _validate(df)
    log.info(f"  Filas válidas: {len(df)}")
    return _load_staging(df, engine)


# ---------------------------------------------------------------------------
# Detección de tipo de hoja
# ---------------------------------------------------------------------------

def _detect_sheet_type(rows: list) -> Optional[str]:
    """
    Analiza las primeras filas y determina 'tabular', 'reporte' o None.

    Tabular:  hay una fila encabezado con 'FECHA' y 'CUENTA'.
    Reporte:  hay una fila encabezado con 'ENTIDAD' y 'TIPO DE DEPOSITO',
              más columnas de tipo datetime (fechas mensuales).
    """
    for row in rows[:15]:
        vals = [str(v).strip().upper() for v in row if v is not None]
        joined = " | ".join(vals)
        if "FECHA" in vals and "CUENTA" in vals:
            return "tabular"
        if "ENTIDAD" in vals and "TIPO DE DEPOSITO" in vals:
            # Verificar que haya fechas en las columnas siguientes
            has_dates = any(isinstance(v, __import__("datetime").datetime) for v in row)
            if has_dates:
                return "reporte"
    return None


# ---------------------------------------------------------------------------
# Parser hoja TABULAR (BASE BANCA PRIVADA, 2021+)
# ---------------------------------------------------------------------------

def _parse_tabular(rows: list, archivo: str, hoja: str) -> pd.DataFrame:
    """
    Hoja tabular: encabezados en la primera fila no vacía que contenga 'FECHA'.
    Estructura: [vacío?, FECHA, ENTIDAD, REGION, PROVINCIA, CANTON, CUENTA,
                 TIPO DE DEPOSITO, NUMERO DE CUENTAS, NUMERO DE CLIENTES, SALDO]
    """
    header_idx = None
    for i, row in enumerate(rows):
        vals = [str(v).strip().upper() for v in row if v is not None]
        if "FECHA" in vals and "ENTIDAD" in vals:
            header_idx = i
            break

    if header_idx is None:
        log.warning(f"  [{hoja}] No se encontró encabezado tabular.")
        return pd.DataFrame()

    raw_headers = rows[header_idx]
    data_rows   = rows[header_idx + 1:]

    # Normalizar nombres de columna
    col_map = {
        "FECHA": "fecha",
        "ENTIDAD": "entidad",
        "REGION": "region",
        "PROVINCIA": "provincia",
        "CANTON": "canton",
        "CUENTA": "cuenta",
        "TIPO DE DEPOSITO": "tipo_deposito",
        "TIPO DE DEPÓSITO": "tipo_deposito",
        "NUMERO DE CUENTAS": "numero_cuentas",
        "NUMERO DE CLIENTES": "numero_clientes",
        "SALDO": "saldo",
    }

    headers = []
    col_positions = {}
    for j, h in enumerate(raw_headers):
        if h is None:
            continue
        hn = str(h).strip().upper()
        if hn in col_map:
            col_positions[col_map[hn]] = j
            headers.append((j, col_map[hn]))

    required = {"fecha", "entidad", "tipo_deposito"}
    if not required.issubset(col_positions.keys()):
        log.warning(f"  [{hoja}] Faltan columnas requeridas: {required - set(col_positions.keys())}")
        return pd.DataFrame()

    records = []
    for row in data_rows:
        if all(v is None for v in row):
            continue
        rec = {}
        for col_name, pos in col_positions.items():
            rec[col_name] = row[pos] if pos < len(row) else None
        records.append(rec)

    df = pd.DataFrame(records)
    df["tipo_hoja"]      = "tabular"
    df["archivo_origen"] = archivo
    df["hoja_origen"]    = hoja
    df["fecha_carga"]    = pd.Timestamp.now()

    df = _normalize_columns(df)
    return df


# ---------------------------------------------------------------------------
# Parser hoja REPORTE (BANCOS PRIVADOS / BANCA PRIVADA, todos los años)
# ---------------------------------------------------------------------------

def _parse_reporte(rows: list, archivo: str, hoja: str) -> pd.DataFrame:
    """
    Hoja reporte: formato ancho con fechas como columnas.
    Solo contiene NUMERO DE CLIENTES.

    Detección dinámica del encabezado: busca la fila con 'ENTIDAD'.
    Columnas descriptivas: ENTIDAD, REGION, PROVINCIA, CANTON, TIPO DE DEPOSITO.
    Columnas métricas: las columnas datetime → se transforman a filas (melt).
    """
    header_idx   = None
    metric_label = "NUMERO DE CLIENTES"

    for i, row in enumerate(rows[:20]):
        vals_upper = [str(v).strip().upper() for v in row if v is not None]
        if "ENTIDAD" in vals_upper and "TIPO DE DEPOSITO" in vals_upper:
            header_idx = i
            # Buscar la etiqueta de métrica en las filas anteriores
            for prev_row in rows[max(0, i - 3):i]:
                for cell in prev_row:
                    if cell is not None and "NUMERO DE CLIENTES" in str(cell).upper():
                        metric_label = "numero_clientes"
                        break
            break

    if header_idx is None:
        log.warning(f"  [{hoja}] No se encontró encabezado de reporte.")
        return pd.DataFrame()

    header_row = rows[header_idx]
    data_rows  = rows[header_idx + 1:]

    # Identificar columnas descriptivas y de fechas
    desc_cols  = {}   # nombre_estándar -> índice
    date_cols  = []   # (índice, datetime)

    DESC_MAP = {
        "ENTIDAD": "entidad",
        "REGION": "region",
        "PROVINCIA": "provincia",
        "CANTON": "canton",
        "TIPO DE DEPOSITO": "tipo_deposito",
        "TIPO DE DEPÓSITO": "tipo_deposito",
    }

    import datetime as dt_mod
    for j, cell in enumerate(header_row):
        if cell is None:
            continue
        if isinstance(cell, (dt_mod.datetime, dt_mod.date)):
            date_cols.append((j, cell.date() if isinstance(cell, dt_mod.datetime) else cell))
        elif str(cell).strip().upper() in DESC_MAP:
            desc_cols[DESC_MAP[str(cell).strip().upper()]] = j

    if not date_cols:
        log.warning(f"  [{hoja}] No se encontraron columnas de fecha en reporte.")
        return pd.DataFrame()

    # Construir DataFrame ancho
    col_names = {}
    for name, idx in desc_cols.items():
        col_names[idx] = name
    for idx, d in date_cols:
        col_names[idx] = d  # La fecha como nombre de columna

    max_col = max(col_names.keys()) + 1
    records = []
    for row in data_rows:
        if all(v is None for v in row):
            continue
        padded = list(row) + [None] * max_col
        rec = {name: padded[idx] for idx, name in col_names.items()}
        records.append(rec)

    df_wide = pd.DataFrame(records)

    # Forward fill columnas descriptivas (jerarquía visual del reporte)
    for col in ["entidad", "region", "provincia", "canton"]:
        if col in df_wide.columns:
            df_wide[col] = df_wide[col].ffill()

    # Excluir filas de totales
    total_mask = df_wide.apply(_is_total_row, axis=1)
    n_totals = total_mask.sum()
    if n_totals:
        log.info(f"  [{hoja}] Descartadas {n_totals} filas de totales.")
    df_wide = df_wide[~total_mask].copy()

    # Melt: una fila por (descriptoras + fecha)
    date_col_names = [d for _, d in date_cols]
    id_vars = [c for c in ["entidad", "region", "provincia", "canton", "tipo_deposito"]
               if c in df_wide.columns]

    df_long = df_wide.melt(
        id_vars=id_vars,
        value_vars=date_col_names,
        var_name="fecha",
        value_name="numero_clientes",
    )

    # Columnas ausentes en reporte
    df_long["numero_cuentas"] = None
    df_long["saldo"]          = None
    df_long["cuenta"]         = None
    df_long["tipo_hoja"]      = "reporte"
    df_long["archivo_origen"] = archivo
    df_long["hoja_origen"]    = hoja
    df_long["fecha_carga"]    = pd.Timestamp.now()

    df_long = _normalize_columns(df_long)
    return df_long


# ---------------------------------------------------------------------------
# Normalización compartida
# ---------------------------------------------------------------------------

def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Limpia tipos y textos en el DataFrame ya transformado a formato largo."""
    import datetime as dt_mod

    # fecha → date
    if "fecha" in df.columns:
        def _to_date(v):
            if isinstance(v, (dt_mod.datetime,)):
                return v.date()
            if isinstance(v, dt_mod.date):
                return v
            try:
                return pd.to_datetime(v).date()
            except Exception:
                return None
        df["fecha"] = df["fecha"].apply(_to_date)

    # texto → strip + upper
    for col in ["entidad", "region", "provincia", "canton", "tipo_deposito", "cuenta"]:
        if col in df.columns:
            df[col] = df[col].apply(
                lambda v: str(v).strip().upper() if v is not None and str(v).strip() else None
            )

    # normalizar tipo_deposito: variantes de tildes/mayúsculas
    if "tipo_deposito" in df.columns:
        df["tipo_deposito"] = df["tipo_deposito"].apply(_normalize_tipo_deposito)

    # numéricos
    for col in ["numero_cuentas", "numero_clientes", "saldo"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # hash_registro
    df["hash_registro"] = df.apply(_compute_hash, axis=1)

    return df


_TIPO_NORMALIZATION = {
    "DE 181 A 360 DIAS":     "DE 181 A 360 DÍAS",
    "DE MAS DE 361 DIAS":    "DE MÁS DE 361 DÍAS",
    "DE MAS DE 361 DÍAS":    "DE MÁS DE 361 DÍAS",
    "DE MÁS DE 361 DIAS":    "DE MÁS DE 361 DÍAS",
    "DE MAS DE 361 DIAS":    "DE MÁS DE 361 DÍAS",
    "DEPOSITOS MONETARIOS QUE GENERAN INTERESES":      "DEPÓSITOS MONETARIOS QUE GENERAN INTERESES",
    "DEPOSITOS MONETARIOS QUE NO GENERAN INTERESES":   "DEPÓSITOS MONETARIOS QUE NO GENERAN INTERESES",
    "DEPOSITOS MONETARIOS DE INSTITUCIONES FINANCIERAS":"DEPÓSITOS MONETARIOS DE INSTITUCIONES FINANCIERAS",
    "DEPOSITOS DE AHORRO":   "DEPÓSITOS DE AHORRO",
    "DEPOSITOS A PLAZO":     "DEPÓSITOS A PLAZO",
}

def _normalize_tipo_deposito(v: Optional[str]) -> Optional[str]:
    if v is None:
        return None
    norm = v.strip().upper()
    # Quitar dobles espacios
    norm = " ".join(norm.split())
    return _TIPO_NORMALIZATION.get(norm, norm)


def _is_total_row(row: pd.Series) -> bool:
    """True si alguna celda descriptiva comienza con 'Total' (case-insensitive)."""
    for col in ["entidad", "region", "provincia", "canton", "tipo_deposito"]:
        if col in row and row[col] is not None:
            v = str(row[col]).strip().upper()
            if v.startswith("TOTAL"):
                return True
    return False


def _compute_hash(row: pd.Series) -> str:
    """SHA-256 de la clave natural del registro."""
    parts = [
        str(row.get("fecha", "") or ""),
        str(row.get("entidad", "") or ""),
        str(row.get("region", "") or ""),
        str(row.get("provincia", "") or ""),
        str(row.get("canton", "") or ""),
        str(row.get("cuenta", "") or ""),
        str(row.get("tipo_deposito", "") or ""),
    ]
    return hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()


# ---------------------------------------------------------------------------
# Validación
# ---------------------------------------------------------------------------

def _validate(df: pd.DataFrame) -> pd.DataFrame:
    """Descarta filas que no cumplan reglas mínimas de calidad."""
    initial = len(df)

    mask_ok = (
        df["fecha"].notna() &
        df["entidad"].notna() &
        df["tipo_deposito"].notna()
    )
    df = df[mask_ok].copy()

    dropped = initial - len(df)
    if dropped:
        log.info(f"  Descartadas por calidad (fecha/entidad/tipo nulos): {dropped}")
    return df


# ---------------------------------------------------------------------------
# Carga a staging
# ---------------------------------------------------------------------------

_STG_COLS = [
    "archivo_origen", "hoja_origen", "tipo_hoja",
    "fecha", "entidad", "region", "provincia", "canton",
    "cuenta", "tipo_deposito",
    "numero_cuentas", "numero_clientes", "saldo",
    "hash_registro", "fecha_carga",
]

def _load_staging(df: pd.DataFrame, engine) -> int:
    """
    Inserta en stg_captaciones solo los registros cuyo hash_registro
    no existe aún en staging (idempotente).
    """
    # Alinear columnas
    for col in _STG_COLS:
        if col not in df.columns:
            df[col] = None
    df = df[_STG_COLS]

    # Deduplicar dentro del propio DataFrame
    df = df.drop_duplicates(subset=["hash_registro"])

    # Excluir hashes ya presentes en staging
    with engine.connect() as conn:
        existing_hashes = set(
            pd.read_sql(f"SELECT hash_registro FROM {STG_TABLE}", conn)["hash_registro"].tolist()
        )
    new_df = df[~df["hash_registro"].isin(existing_hashes)]

    if new_df.empty:
        log.info(f"  [staging] Sin registros nuevos.")
        return 0

    new_df.to_sql(STG_TABLE, engine, if_exists="append", index=False)
    log.info(f"  [staging] {len(new_df)} filas insertadas.")
    return len(new_df)


# ---------------------------------------------------------------------------
# Consolidación staging → final
# ---------------------------------------------------------------------------

def _consolidate(engine) -> None:
    """
    Mueve registros de stg_captaciones a captaciones sin duplicar.
    Prioridad: tabular > reporte (si hay solapamiento de clave natural).
    Clave natural: fecha + entidad + region + provincia + canton + cuenta + tipo_deposito
    """
    sql = text(f"""
        INSERT INTO {FINAL_TABLE} (
            fecha, entidad, region, provincia, canton,
            cuenta, tipo_deposito,
            numero_cuentas, numero_clientes, saldo,
            archivo_origen, hoja_origen, tipo_hoja,
            hash_registro, fecha_carga
        )
        SELECT
            s.fecha, s.entidad, s.region, s.provincia, s.canton,
            s.cuenta, s.tipo_deposito,
            s.numero_cuentas, s.numero_clientes, s.saldo,
            s.archivo_origen, s.hoja_origen, s.tipo_hoja,
            s.hash_registro, s.fecha_carga
        FROM {STG_TABLE} s
        WHERE s.hash_registro NOT IN (SELECT hash_registro FROM {FINAL_TABLE})
          AND s.tipo_hoja = 'tabular'

        UNION ALL

        SELECT
            s.fecha, s.entidad, s.region, s.provincia, s.canton,
            s.cuenta, s.tipo_deposito,
            s.numero_cuentas, s.numero_clientes, s.saldo,
            s.archivo_origen, s.hoja_origen, s.tipo_hoja,
            s.hash_registro, s.fecha_carga
        FROM {STG_TABLE} s
        WHERE s.hash_registro NOT IN (SELECT hash_registro FROM {FINAL_TABLE})
          AND s.tipo_hoja = 'reporte'
          AND NOT EXISTS (
              SELECT 1 FROM {STG_TABLE} s2
              WHERE s2.tipo_hoja = 'tabular'
                AND s2.fecha          = s.fecha
                AND s2.entidad        = s.entidad
                AND COALESCE(s2.region,'')   = COALESCE(s.region,'')
                AND COALESCE(s2.provincia,'')= COALESCE(s.provincia,'')
                AND COALESCE(s2.canton,'')   = COALESCE(s.canton,'')
                AND COALESCE(s2.tipo_deposito,'') = COALESCE(s.tipo_deposito,'')
          )
    """)

    with engine.begin() as conn:
        result = conn.execute(sql)
        inserted = result.rowcount if result.rowcount >= 0 else "?"
    log.info(f"[depositos] [{FINAL_TABLE}] {inserted} filas consolidadas.")


# ---------------------------------------------------------------------------
# Creación de tablas
# ---------------------------------------------------------------------------

def _ensure_tables(engine) -> None:
    inspector = sa_inspect(engine)

    if not inspector.has_table(STG_TABLE):
        with engine.begin() as conn:
            conn.execute(text(f"""
                CREATE TABLE {STG_TABLE} (
                    id_stg           INT IDENTITY(1,1) PRIMARY KEY,
                    archivo_origen   NVARCHAR(255),
                    hoja_origen      NVARCHAR(100),
                    tipo_hoja        NVARCHAR(20),
                    fecha            DATE,
                    entidad          NVARCHAR(200),
                    region           NVARCHAR(100),
                    provincia        NVARCHAR(100),
                    canton           NVARCHAR(100),
                    cuenta           NVARCHAR(20),
                    tipo_deposito    NVARCHAR(200),
                    numero_cuentas   FLOAT,
                    numero_clientes  FLOAT,
                    saldo            FLOAT,
                    hash_registro    NVARCHAR(64),
                    fecha_carga      DATETIME
                )
            """))
        log.info(f"[depositos] Tabla {STG_TABLE} creada.")

    if not inspector.has_table(FINAL_TABLE):
        with engine.begin() as conn:
            conn.execute(text(f"""
                CREATE TABLE {FINAL_TABLE} (
                    id               INT IDENTITY(1,1) NOT NULL,
                    fecha            DATE          NOT NULL,
                    entidad          NVARCHAR(200) NOT NULL,
                    region           NVARCHAR(100),
                    provincia        NVARCHAR(100),
                    canton           NVARCHAR(100),
                    cuenta           NVARCHAR(20),
                    tipo_deposito    NVARCHAR(200) NOT NULL,
                    numero_cuentas   FLOAT,
                    numero_clientes  FLOAT,
                    saldo            FLOAT,
                    archivo_origen   NVARCHAR(255),
                    hoja_origen      NVARCHAR(100),
                    tipo_hoja        NVARCHAR(20),
                    hash_registro    NVARCHAR(64),
                    fecha_carga      DATETIME,
                    -- PK no clusterizada para liberar el índice clusterizado
                    CONSTRAINT PK_{FINAL_TABLE} PRIMARY KEY NONCLUSTERED (id),
                    CONSTRAINT UQ_{FINAL_TABLE}_hash UNIQUE NONCLUSTERED (hash_registro)
                )
            """))
            # Índice clusterizado en (fecha, provincia, canton) →
            # los datos se almacenan físicamente en ese orden
            conn.execute(text(f"""
                CREATE CLUSTERED INDEX CIX_{FINAL_TABLE}_fecha_prov_canton
                    ON {FINAL_TABLE} (fecha, provincia, canton)
            """))
        log.info(f"[depositos] Tabla {FINAL_TABLE} creada (orden: fecha → provincia → canton).")
        return

    # Tabla ya existe: verificar que el índice clusterizado es el correcto
    _ensure_clustered_index(engine)


def _ensure_clustered_index(engine) -> None:
    """
    Si la tabla captaciones existe pero su índice clusterizado no es
    (fecha, provincia, canton), lo recrea:
      1. Elimina el índice clusterizado actual (puede ser el PK).
      2. Convierte el PK a NONCLUSTERED si hace falta.
      3. Crea el nuevo índice clusterizado en (fecha, provincia, canton).
    """
    check_sql = text("""
        SELECT i.name, i.is_primary_key,
               STRING_AGG(c.name, ',') WITHIN GROUP (ORDER BY ic.key_ordinal) AS cols
        FROM sys.indexes i
        JOIN sys.index_columns ic ON ic.object_id = i.object_id AND ic.index_id = i.index_id
        JOIN sys.columns c        ON c.object_id  = i.object_id AND c.column_id = ic.column_id
        WHERE i.object_id = OBJECT_ID(:tbl)
          AND i.type = 1  -- 1 = CLUSTERED
        GROUP BY i.name, i.is_primary_key
    """)

    with engine.connect() as conn:
        row = conn.execute(check_sql, {"tbl": FINAL_TABLE}).fetchone()

    if row is None:
        # No hay índice clusterizado aún — crearlo directamente
        with engine.begin() as conn:
            conn.execute(text(f"""
                CREATE CLUSTERED INDEX CIX_{FINAL_TABLE}_fecha_prov_canton
                    ON {FINAL_TABLE} (fecha, provincia, canton)
            """))
        log.info(f"[depositos] Índice clusterizado creado en {FINAL_TABLE}.")
        return

    idx_name, is_pk, cols = row
    if cols == "fecha,provincia,canton":
        log.info(f"[depositos] Índice clusterizado ya es correcto: {cols}")
        return

    log.info(f"[depositos] Reordenando índice clusterizado de '{cols}' → 'fecha,provincia,canton'")
    with engine.begin() as conn:
        if is_pk:
            # El PK es clusterizado: hay que eliminarlo y recrearlo como NONCLUSTERED
            conn.execute(text(f"ALTER TABLE {FINAL_TABLE} DROP CONSTRAINT {idx_name}"))
            conn.execute(text(f"""
                ALTER TABLE {FINAL_TABLE}
                    ADD CONSTRAINT PK_{FINAL_TABLE} PRIMARY KEY NONCLUSTERED (id)
            """))
        else:
            conn.execute(text(f"DROP INDEX {idx_name} ON {FINAL_TABLE}"))

        conn.execute(text(f"""
            CREATE CLUSTERED INDEX CIX_{FINAL_TABLE}_fecha_prov_canton
                ON {FINAL_TABLE} (fecha, provincia, canton)
        """))
    log.info(f"[depositos] Índice clusterizado actualizado.")
