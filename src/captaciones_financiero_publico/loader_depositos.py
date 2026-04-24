"""
ETL completo: Captaciones / Depósitos — Instituciones Financieras Públicas
Flujo: Excel → stg_captaciones_publicas → captaciones_publicas

Misma lógica de parseo que captaciones_financiero_privado/loader_depositos.py.
Solo cambian los nombres de tabla.
"""

import sys
import hashlib
import logging
from pathlib import Path
from typing import Optional

import pandas as pd
import openpyxl
from sqlalchemy import text, inspect as sa_inspect

sys.path.append(str(Path(__file__).resolve().parents[2]))
from utils.base_engine import get_master_engine

STG_TABLE   = "stg_captaciones_publicas"
FINAL_TABLE = "captaciones_publicas"

logging.basicConfig(level=logging.INFO, format="[%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

_TOTAL_PREFIXES = ("total ", "total\t", "subtotal")


# ---------------------------------------------------------------------------
# Punto de entrada principal
# ---------------------------------------------------------------------------

def load(excel_paths: list) -> None:
    if not excel_paths:
        log.info("[depositos_pub] Sin archivos para procesar.")
        return

    engine = get_master_engine()
    _ensure_tables(engine)

    total_stg = 0
    for path in excel_paths:
        path = Path(path)
        if not path.exists() or path.name.startswith("~$"):
            continue
        log.info(f"[depositos_pub] Procesando: {path.name}")
        try:
            rows_loaded = _process_file(path, engine)
            total_stg += rows_loaded
        except Exception as ex:
            log.error(f"[depositos_pub] Error en {path.name}: {ex}")

    log.info(f"[depositos_pub] Total filas cargadas a staging: {total_stg}")
    _consolidate(engine)


# ---------------------------------------------------------------------------
# Procesamiento de un archivo Excel
# ---------------------------------------------------------------------------

def _process_file(path: Path, engine) -> int:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = wb.sheetnames

    tabular_df = None
    reporte_df = None

    for sh in sheets:
        ws = wb[sh]
        rows = list(ws.iter_rows(values_only=True))
        wb_type = _detect_sheet_type(rows)
        log.info(f"  Hoja '{sh}': tipo={wb_type}")

        if wb_type == "tabular":
            tabular_df = _parse_tabular(rows, path.name, sh)
        elif wb_type == "reporte":
            reporte_df = _parse_reporte(rows, path.name, sh)

    wb.close()

    if tabular_df is not None and not tabular_df.empty:
        df = tabular_df
    elif reporte_df is not None and not reporte_df.empty:
        df = reporte_df
    else:
        log.warning(f"  Sin datos utilizables en {path.name}")
        return 0

    df = _validate(df)
    log.info(f"  Filas válidas: {len(df)}")
    return _load_staging(df, engine)


# ---------------------------------------------------------------------------
# Detección de tipo de hoja
# ---------------------------------------------------------------------------

def _detect_sheet_type(rows: list) -> Optional[str]:
    for row in rows[:15]:
        vals = [str(v).strip().upper() for v in row if v is not None]
        if "FECHA" in vals and "CUENTA" in vals:
            return "tabular"
        if "ENTIDAD" in vals and "TIPO DE DEPOSITO" in vals:
            has_dates = any(isinstance(v, __import__("datetime").datetime) for v in row)
            if has_dates:
                return "reporte"
    return None


# ---------------------------------------------------------------------------
# Parser hoja TABULAR
# ---------------------------------------------------------------------------

def _parse_tabular(rows: list, archivo: str, hoja: str) -> pd.DataFrame:
    header_idx = None
    for i, row in enumerate(rows):
        vals = [str(v).strip().upper() for v in row if v is not None]
        if "FECHA" in vals and "ENTIDAD" in vals:
            header_idx = i
            break

    if header_idx is None:
        log.warning(f"  [{hoja}] No se encontró encabezado tabular.")
        return pd.DataFrame()

    raw_headers = rows[header_idx]
    data_rows   = rows[header_idx + 1:]

    col_map = {
        "FECHA":              "fecha",
        "ENTIDAD":            "entidad",
        "REGION":             "region",
        "PROVINCIA":          "provincia",
        "CANTON":             "canton",
        "CUENTA":             "cuenta",
        "TIPO DE DEPOSITO":   "tipo_deposito",
        "TIPO DE DEPÓSITO":   "tipo_deposito",
        "NUMERO DE CUENTAS":  "numero_cuentas",
        "NUMERO DE CLIENTES": "numero_clientes",
        "SALDO":              "saldo",
    }

    col_positions = {}
    for j, h in enumerate(raw_headers):
        if h is None:
            continue
        hn = str(h).strip().upper()
        if hn in col_map:
            col_positions[col_map[hn]] = j

    required = {"fecha", "entidad", "tipo_deposito"}
    if not required.issubset(col_positions.keys()):
        log.warning(f"  [{hoja}] Faltan columnas: {required - set(col_positions.keys())}")
        return pd.DataFrame()

    records = []
    for row in data_rows:
        if all(v is None for v in row):
            continue
        rec = {col: (row[pos] if pos < len(row) else None) for col, pos in col_positions.items()}
        records.append(rec)

    df = pd.DataFrame(records)
    df["tipo_hoja"]      = "tabular"
    df["archivo_origen"] = archivo
    df["hoja_origen"]    = hoja
    df["fecha_carga"]    = pd.Timestamp.now()
    return _normalize_columns(df)


# ---------------------------------------------------------------------------
# Parser hoja REPORTE
# ---------------------------------------------------------------------------

def _parse_reporte(rows: list, archivo: str, hoja: str) -> pd.DataFrame:
    header_idx = None
    for i, row in enumerate(rows[:20]):
        vals_upper = [str(v).strip().upper() for v in row if v is not None]
        if "ENTIDAD" in vals_upper and "TIPO DE DEPOSITO" in vals_upper:
            header_idx = i
            break

    if header_idx is None:
        log.warning(f"  [{hoja}] No se encontró encabezado de reporte.")
        return pd.DataFrame()

    header_row = rows[header_idx]
    data_rows  = rows[header_idx + 1:]

    desc_cols = {}
    date_cols = []

    DESC_MAP = {
        "ENTIDAD":          "entidad",
        "REGION":           "region",
        "PROVINCIA":        "provincia",
        "CANTON":           "canton",
        "TIPO DE DEPOSITO": "tipo_deposito",
        "TIPO DE DEPÓSITO": "tipo_deposito",
    }

    import datetime as dt_mod
    for j, cell in enumerate(header_row):
        if cell is None:
            continue
        if isinstance(cell, (dt_mod.datetime, dt_mod.date)):
            date_cols.append((j, cell.date() if isinstance(cell, dt_mod.datetime) else cell))
        elif str(cell).strip().upper() in DESC_MAP:
            desc_cols[DESC_MAP[str(cell).strip().upper()]] = j

    if not date_cols:
        log.warning(f"  [{hoja}] No se encontraron columnas de fecha en reporte.")
        return pd.DataFrame()

    col_names = {}
    for name, idx in desc_cols.items():
        col_names[idx] = name
    for idx, d in date_cols:
        col_names[idx] = d

    max_col = max(col_names.keys()) + 1
    records = []
    for row in data_rows:
        if all(v is None for v in row):
            continue
        padded = list(row) + [None] * max_col
        rec = {name: padded[idx] for idx, name in col_names.items()}
        records.append(rec)

    df_wide = pd.DataFrame(records)

    for col in ["entidad", "region", "provincia", "canton"]:
        if col in df_wide.columns:
            df_wide[col] = df_wide[col].ffill()

    total_mask = df_wide.apply(_is_total_row, axis=1)
    if total_mask.sum():
        log.info(f"  [{hoja}] Descartadas {total_mask.sum()} filas de totales.")
    df_wide = df_wide[~total_mask].copy()

    date_col_names = [d for _, d in date_cols]
    id_vars = [c for c in ["entidad", "region", "provincia", "canton", "tipo_deposito"]
               if c in df_wide.columns]

    df_long = df_wide.melt(
        id_vars=id_vars,
        value_vars=date_col_names,
        var_name="fecha",
        value_name="numero_clientes",
    )

    df_long["numero_cuentas"] = None
    df_long["saldo"]          = None
    df_long["cuenta"]         = None
    df_long["tipo_hoja"]      = "reporte"
    df_long["archivo_origen"] = archivo
    df_long["hoja_origen"]    = hoja
    df_long["fecha_carga"]    = pd.Timestamp.now()
    return _normalize_columns(df_long)


# ---------------------------------------------------------------------------
# Normalización compartida
# ---------------------------------------------------------------------------

def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    import datetime as dt_mod

    if "fecha" in df.columns:
        def _to_date(v):
            if isinstance(v, dt_mod.datetime):
                return v.date()
            if isinstance(v, dt_mod.date):
                return v
            try:
                return pd.to_datetime(v).date()
            except Exception:
                return None
        df["fecha"] = df["fecha"].apply(_to_date)

    for col in ["entidad", "region", "provincia", "canton", "tipo_deposito", "cuenta"]:
        if col in df.columns:
            df[col] = df[col].apply(
                lambda v: str(v).strip().upper() if v is not None and str(v).strip() else None
            )

    if "tipo_deposito" in df.columns:
        df["tipo_deposito"] = df["tipo_deposito"].apply(_normalize_tipo_deposito)

    for col in ["numero_cuentas", "numero_clientes", "saldo"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    df["hash_registro"] = df.apply(_compute_hash, axis=1)
    return df


_TIPO_NORMALIZATION = {
    "DE 181 A 360 DIAS":     "DE 181 A 360 DÍAS",
    "DE MAS DE 361 DIAS":    "DE MÁS DE 361 DÍAS",
    "DE MAS DE 361 DÍAS":    "DE MÁS DE 361 DÍAS",
    "DE MÁS DE 361 DIAS":    "DE MÁS DE 361 DÍAS",
    "DEPOSITOS MONETARIOS QUE GENERAN INTERESES":      "DEPÓSITOS MONETARIOS QUE GENERAN INTERESES",
    "DEPOSITOS MONETARIOS QUE NO GENERAN INTERESES":   "DEPÓSITOS MONETARIOS QUE NO GENERAN INTERESES",
    "DEPOSITOS MONETARIOS DE INSTITUCIONES FINANCIERAS":"DEPÓSITOS MONETARIOS DE INSTITUCIONES FINANCIERAS",
    "DEPOSITOS DE AHORRO":   "DEPÓSITOS DE AHORRO",
    "DEPOSITOS A PLAZO":     "DEPÓSITOS A PLAZO",
}

def _normalize_tipo_deposito(v: Optional[str]) -> Optional[str]:
    if v is None:
        return None
    norm = " ".join(v.strip().upper().split())
    return _TIPO_NORMALIZATION.get(norm, norm)


def _is_total_row(row: pd.Series) -> bool:
    for col in ["entidad", "region", "provincia", "canton", "tipo_deposito"]:
        if col in row and row[col] is not None:
            if str(row[col]).strip().upper().startswith("TOTAL"):
                return True
    return False


def _compute_hash(row: pd.Series) -> str:
    parts = [
        str(row.get("fecha",         "") or ""),
        str(row.get("entidad",       "") or ""),
        str(row.get("region",        "") or ""),
        str(row.get("provincia",     "") or ""),
        str(row.get("canton",        "") or ""),
        str(row.get("cuenta",        "") or ""),
        str(row.get("tipo_deposito", "") or ""),
    ]
    return hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()


# ---------------------------------------------------------------------------
# Validación
# ---------------------------------------------------------------------------

def _validate(df: pd.DataFrame) -> pd.DataFrame:
    initial = len(df)
    mask = df["fecha"].notna() & df["entidad"].notna() & df["tipo_deposito"].notna()
    df = df[mask].copy()
    dropped = initial - len(df)
    if dropped:
        log.info(f"  Descartadas por calidad: {dropped}")
    return df


# ---------------------------------------------------------------------------
# Carga a staging
# ---------------------------------------------------------------------------

_STG_COLS = [
    "archivo_origen", "hoja_origen", "tipo_hoja",
    "fecha", "entidad", "region", "provincia", "canton",
    "cuenta", "tipo_deposito",
    "numero_cuentas", "numero_clientes", "saldo",
    "hash_registro", "fecha_carga",
]

def _load_staging(df: pd.DataFrame, engine) -> int:
    for col in _STG_COLS:
        if col not in df.columns:
            df[col] = None
    df = df[_STG_COLS].drop_duplicates(subset=["hash_registro"])

    with engine.connect() as conn:
        existing = set(
            pd.read_sql(f"SELECT hash_registro FROM {STG_TABLE}", conn)["hash_registro"].tolist()
        )
    new_df = df[~df["hash_registro"].isin(existing)]

    if new_df.empty:
        log.info(f"  [staging] Sin registros nuevos.")
        return 0

    new_df.to_sql(STG_TABLE, engine, if_exists="append", index=False)
    log.info(f"  [staging] {len(new_df)} filas insertadas.")
    return len(new_df)


# ---------------------------------------------------------------------------
# Consolidación staging → final
# ---------------------------------------------------------------------------

def _consolidate(engine) -> None:
    sql = text(f"""
        INSERT INTO {FINAL_TABLE} (
            fecha, entidad, region, provincia, canton,
            cuenta, tipo_deposito,
            numero_cuentas, numero_clientes, saldo,
            archivo_origen, hoja_origen, tipo_hoja,
            hash_registro, fecha_carga
        )
        SELECT
            s.fecha, s.entidad, s.region, s.provincia, s.canton,
            s.cuenta, s.tipo_deposito,
            s.numero_cuentas, s.numero_clientes, s.saldo,
            s.archivo_origen, s.hoja_origen, s.tipo_hoja,
            s.hash_registro, s.fecha_carga
        FROM {STG_TABLE} s
        WHERE s.hash_registro NOT IN (SELECT hash_registro FROM {FINAL_TABLE})
          AND s.tipo_hoja = 'tabular'

        UNION ALL

        SELECT
            s.fecha, s.entidad, s.region, s.provincia, s.canton,
            s.cuenta, s.tipo_deposito,
            s.numero_cuentas, s.numero_clientes, s.saldo,
            s.archivo_origen, s.hoja_origen, s.tipo_hoja,
            s.hash_registro, s.fecha_carga
        FROM {STG_TABLE} s
        WHERE s.hash_registro NOT IN (SELECT hash_registro FROM {FINAL_TABLE})
          AND s.tipo_hoja = 'reporte'
          AND NOT EXISTS (
              SELECT 1 FROM {STG_TABLE} s2
              WHERE s2.tipo_hoja = 'tabular'
                AND s2.fecha              = s.fecha
                AND s2.entidad            = s.entidad
                AND COALESCE(s2.region,'')         = COALESCE(s.region,'')
                AND COALESCE(s2.provincia,'')      = COALESCE(s.provincia,'')
                AND COALESCE(s2.canton,'')         = COALESCE(s.canton,'')
                AND COALESCE(s2.tipo_deposito,'')  = COALESCE(s.tipo_deposito,'')
          )
    """)
    with engine.begin() as conn:
        result = conn.execute(sql)
        inserted = result.rowcount if result.rowcount >= 0 else "?"
    log.info(f"[depositos_pub] [{FINAL_TABLE}] {inserted} filas consolidadas.")


# ---------------------------------------------------------------------------
# Creación de tablas
# ---------------------------------------------------------------------------

def _ensure_tables(engine) -> None:
    inspector = sa_inspect(engine)

    if not inspector.has_table(STG_TABLE):
        with engine.begin() as conn:
            conn.execute(text(f"""
                CREATE TABLE {STG_TABLE} (
                    id_stg           INT IDENTITY(1,1) PRIMARY KEY,
                    archivo_origen   NVARCHAR(255),
                    hoja_origen      NVARCHAR(100),
                    tipo_hoja        NVARCHAR(20),
                    fecha            DATE,
                    entidad          NVARCHAR(200),
                    region           NVARCHAR(100),
                    provincia        NVARCHAR(100),
                    canton           NVARCHAR(100),
                    cuenta           NVARCHAR(20),
                    tipo_deposito    NVARCHAR(200),
                    numero_cuentas   FLOAT,
                    numero_clientes  FLOAT,
                    saldo            FLOAT,
                    hash_registro    NVARCHAR(64),
                    fecha_carga      DATETIME
                )
            """))
        log.info(f"[depositos_pub] Tabla {STG_TABLE} creada.")

    if not inspector.has_table(FINAL_TABLE):
        with engine.begin() as conn:
            conn.execute(text(f"""
                CREATE TABLE {FINAL_TABLE} (
                    id               INT IDENTITY(1,1) NOT NULL,
                    fecha            DATE          NOT NULL,
                    entidad          NVARCHAR(200) NOT NULL,
                    region           NVARCHAR(100),
                    provincia        NVARCHAR(100),
                    canton           NVARCHAR(100),
                    cuenta           NVARCHAR(20),
                    tipo_deposito    NVARCHAR(200) NOT NULL,
                    numero_cuentas   FLOAT,
                    numero_clientes  FLOAT,
                    saldo            FLOAT,
                    archivo_origen   NVARCHAR(255),
                    hoja_origen      NVARCHAR(100),
                    tipo_hoja        NVARCHAR(20),
                    hash_registro    NVARCHAR(64),
                    fecha_carga      DATETIME,
                    CONSTRAINT PK_{FINAL_TABLE} PRIMARY KEY NONCLUSTERED (id),
                    CONSTRAINT UQ_{FINAL_TABLE}_hash UNIQUE NONCLUSTERED (hash_registro)
                )
            """))
            conn.execute(text(f"""
                CREATE CLUSTERED INDEX CIX_{FINAL_TABLE}_fecha_prov_canton
                    ON {FINAL_TABLE} (fecha, provincia, canton)
            """))
        log.info(f"[depositos_pub] Tabla {FINAL_TABLE} creada.")
