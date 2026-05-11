"""
ETL: Información Monetaria Semanal — BCE Ecuador

IMS1 → tabla ancha depositos_gobierno_ims1 (una fila por año-mes, desde 2006)
Resto IMS* → una tabla por hoja, formato largo (indicador × período).

Nomenclatura de tablas:
  depositos_gobierno_ims1     ← hoja IMS1  (esquema ancho, ver abajo)
  depositos_gobierno_ims1_1   ← hoja IMS1.1
  depositos_gobierno_ims2     ← hoja IMS2
  depositos_gobierno_ims2_2   ← hoja IMS2 (2)
  ...etc.

Esquema IMS1 (ancho):
  id, anio, mes, dias_mes,
  rild, pasivos_monetarios_pm, emision_monetaria_em, reservas_bancarias_rb,
  depositos_vista, cuasidinero_total/ahorro/plazo/restringido/operaciones_reporto/otros_depositos,
  credito_sector_privado_total/cartera/otros,
  tasa_basica/pasiva/activa,
  inflacion_mensual/anual/acumulada,
  hash_registro, fecha_carga

Esquema resto de tablas IMS (largo):
  id, fecha_semana, anio, indicador, valor, hash_registro, fecha_carga

Hojas excluidas: Caratula/Carátula, Indice/Índice, Presentacion/Presentación, Hoja*.
Si una hoja no existe en el archivo de un año → se omite silenciosamente.
"""

import hashlib
import re
import sys
from datetime import date
from pathlib import Path

import pandas as pd
from sqlalchemy import inspect as sa_inspect, text

sys.path.append(str(Path(__file__).resolve().parents[2]))
from utils.base_engine import get_master_engine

# ---------------------------------------------------------------------------
# Constantes generales
# ---------------------------------------------------------------------------

TABLE_PREFIX = "depositos_gobierno"
_IMS1_TABLE  = "depositos_gobierno_ims1"

_MESES = {
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
}

_SKIP_SHEET_PATTERNS = {
    "caratula", "carátula",
    "indice", "índice",
    "presentacion", "presentación",
}

_FILE_DATE_RE = re.compile(
    r"(?:InfMonetariaSemanal|BMS)_(\d{2})(\d{2})(\d{4})\.(xls|xlsx)$",
    re.IGNORECASE,
)

# DDL para hojas largas (IMS2, IMS3, …)
_DDL_TEMPLATE = """
CREATE TABLE {table} (
    id             BIGINT IDENTITY(1,1) NOT NULL,
    fecha_semana   DATE           NOT NULL,
    anio           INT            NOT NULL,
    indicador      NVARCHAR(300)  NOT NULL,
    valor_millones FLOAT          NULL,
    hash_registro  NVARCHAR(64)   NOT NULL,
    fecha_carga    DATETIME2      DEFAULT GETDATE(),
    CONSTRAINT PK_{table} PRIMARY KEY NONCLUSTERED (id)
)"""

_IDX_TEMPLATE = """
CREATE CLUSTERED INDEX CIX_{table}
ON {table} (fecha_semana, indicador)"""

# DDL específico para IMS1 (formato ancho)
_IMS1_DDL = """
CREATE TABLE depositos_gobierno_ims1 (
    id                              BIGINT IDENTITY(1,1) NOT NULL,
    anio                            INT            NULL,
    mes                             NVARCHAR(20)   NULL,
    dias_mes                        INT            NULL,
    rild                            FLOAT          NULL,
    pasivos_monetarios_pm           FLOAT          NULL,
    emision_monetaria_em            FLOAT          NULL,
    reservas_bancarias_rb           FLOAT          NULL,
    depositos_vista                 FLOAT          NULL,
    cuasidinero_total               FLOAT          NULL,
    cuasidinero_ahorro              FLOAT          NULL,
    cuasidinero_plazo               FLOAT          NULL,
    cuasidinero_restringido         FLOAT          NULL,
    cuasidinero_operaciones_reporto FLOAT          NULL,
    cuasidinero_otros_depositos     FLOAT          NULL,
    credito_sector_privado_total    FLOAT          NULL,
    credito_sector_privado_cartera  FLOAT          NULL,
    credito_sector_privado_otros    FLOAT          NULL,
    tasa_basica                     FLOAT          NULL,
    tasa_pasiva                     FLOAT          NULL,
    tasa_activa                     FLOAT          NULL,
    inflacion_mensual               FLOAT          NULL,
    inflacion_anual                 FLOAT          NULL,
    inflacion_acumulada             FLOAT          NULL,
    hash_registro                   NVARCHAR(64)   NOT NULL,
    fecha_carga                     DATETIME2      DEFAULT GETDATE(),
    CONSTRAINT PK_depositos_gobierno_ims1 PRIMARY KEY NONCLUSTERED (id)
)"""

_IMS1_IDX = """
CREATE CLUSTERED INDEX CIX_depositos_gobierno_ims1
ON depositos_gobierno_ims1 (anio, mes)"""

# DDL específico para IMS1.1 (formato ancho: Oferta Monetaria M1 y Liquidez Total M2)
_IMS1_1_TABLE = "depositos_gobierno_ims1_1"

_IMS1_1_DDL = """
CREATE TABLE depositos_gobierno_ims1_1 (
    id                               BIGINT IDENTITY(1,1) NOT NULL,
    anio                             INT            NULL,
    mes                              NVARCHAR(20)   NULL,
    especies_monetarias_circulacion  FLOAT          NULL,
    moneda_fraccionaria              FLOAT          NULL,
    dinero_electronico               FLOAT          NULL,
    depositos_vista                  FLOAT          NULL,
    oferta_monetaria_m1              FLOAT          NULL,
    cuasidinero                      FLOAT          NULL,
    liquidez_total_m2                FLOAT          NULL,
    reservas_bancarias               FLOAT          NULL,
    caja_bce                         FLOAT          NULL,
    caja_osd                         FLOAT          NULL,
    base_monetaria_bm                FLOAT          NULL,
    multiplicador_m1_bm              FLOAT          NULL,
    multiplicador_m2_bm              FLOAT          NULL,
    hash_registro                    NVARCHAR(64)   NOT NULL,
    fecha_carga                      DATETIME2      DEFAULT GETDATE(),
    CONSTRAINT PK_depositos_gobierno_ims1_1 PRIMARY KEY NONCLUSTERED (id)
)"""

_IMS1_1_IDX = """
CREATE CLUSTERED INDEX CIX_depositos_gobierno_ims1_1
ON depositos_gobierno_ims1_1 (anio, mes)"""

# ---------------------------------------------------------------------------
# Helpers de nombre de tabla y hoja
# ---------------------------------------------------------------------------

def _normalize_sheet(name: str) -> str:
    """'IMS1.1' -> 'ims1_1'  |  'IMS2 (2)' -> 'ims2_2'  |  'IMS14' -> 'ims14'"""
    s = name.lower()
    s = re.sub(r"[\s.()\[\]]+", "_", s)
    s = re.sub(r"[^a-z0-9_]", "", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def _table_name(sheet_name: str) -> str:
    return f"{TABLE_PREFIX}_{_normalize_sheet(sheet_name)}"


def _should_skip_sheet(name: str) -> bool:
    clean = name.lower().strip()
    if any(clean == p for p in _SKIP_SHEET_PATTERNS):
        return True
    if clean.startswith("hoja"):
        return True
    return False


def _is_ims_sheet(name: str) -> bool:
    return name.strip().upper().startswith("IMS")

# ---------------------------------------------------------------------------
# Apertura de libros
# ---------------------------------------------------------------------------

def _get_all_sheets_xls(path: Path) -> dict[str, list]:
    import xlrd
    wb = xlrd.open_workbook(str(path))
    result = {}
    for sname in wb.sheet_names():
        if _should_skip_sheet(sname) or not _is_ims_sheet(sname):
            continue
        sh = wb.sheet_by_name(sname)
        result[sname] = [sh.row_values(r) for r in range(sh.nrows)]
    return result


def _get_all_sheets_xlsx(path: Path) -> dict[str, list]:
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    result = {}
    for sname in wb.sheetnames:
        if _should_skip_sheet(sname) or not _is_ims_sheet(sname):
            continue
        ws = wb[sname]
        result[sname] = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return result


def _open_all_sheets(path: Path) -> dict[str, list]:
    suffix = path.suffix.lower()
    if suffix == ".xls":
        return _get_all_sheets_xls(path)
    elif suffix == ".xlsx":
        return _get_all_sheets_xlsx(path)
    return {}

# ---------------------------------------------------------------------------
# Helpers de valores
# ---------------------------------------------------------------------------

def _last_numeric_col(row: list) -> int | None:
    for i in range(len(row) - 1, -1, -1):
        v = row[i]
        if v is None or isinstance(v, bool) or v in ("", "-"):
            continue
        if isinstance(v, (int, float)):
            return i
        try:
            float(str(v).strip().replace(",", "."))
            return i
        except ValueError:
            continue
    return None


def _to_float(v) -> float | None:
    if v is None or isinstance(v, bool) or v in ("", "-"):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip().replace(",", "."))
    except ValueError:
        return None


def _clean_label(label) -> str | None:
    """Limpia y normaliza la etiqueta de la fila (para hojas largas)."""
    if label is None:
        return None
    s = str(label).strip()
    if not s or s == "-":
        return None
    # Footnote markers: "(1) text", "(*) text", "(**) text", etc.
    if re.match(r"^\([*\d]+\)", s):
        return None
    # Source / unit lines
    if s.lower().startswith(("fuente", "source", "nota:", "note:", "en millones", "in millions")):
        return None
    # Very long strings are footnotes that slipped through
    if len(s) > 295:
        return None
    # Strip leading outline numbering: "1. ", "1.2 ", "A. ", "a. ", "B.1.a. " etc.
    s = re.sub(r"^(?:(?:\d+|[A-Za-z])\.)+\d*\s+", "", s).strip()
    if not s:
        return None
    return s

# ---------------------------------------------------------------------------
# IMS1: detección de columnas y parser ancho
# ---------------------------------------------------------------------------

def _find_ims1_cols(rows: list) -> dict[str, int | None]:
    """
    Escanea las filas de cabecera (0-13) para mapear nombre-de-campo → índice de columna.
    Funciona tanto en archivos anteriores a Dinero Electrónico como en los actuales.
    """
    col_texts: dict[int, str] = {}
    for ri in range(min(14, len(rows))):
        for ci, v in enumerate(rows[ri]):
            s = str(v).strip().lower() if v is not None else ""
            if s:
                col_texts[ci] = col_texts.get(ci, "") + " " + s

    def find_col(*keywords, after: int = -1) -> int | None:
        for ci in sorted(col_texts):
            if ci <= after:
                continue
            t = col_texts[ci]
            if any(kw in t for kw in keywords):
                return ci
        return None

    cuasi_col = find_col("cuasidinero")
    cred_col  = find_col("crédito al sector privado", "credito al sector privado")
    infl_col  = find_col("inflaci", after=(cred_col or -1))

    return {
        "rild":                            find_col("rild", "reserva internacional libre", "reservas internacionales"),
        "pasivos_monetarios_pm":           find_col("pasivos monetarios"),
        "emision_monetaria_em":            find_col("emisión monetaria", "emision monetaria"),
        "reservas_bancarias_rb":           find_col("reservas bancarias"),
        "depositos_vista":                 find_col("a la vista"),
        "cuasidinero_total":               find_col("total",     after=(cuasi_col or 0) - 1),
        "cuasidinero_ahorro":              find_col("ahorro"),
        "cuasidinero_plazo":               find_col("plazo"),
        "cuasidinero_restringido":         find_col("restringido"),
        "cuasidinero_operaciones_reporto": find_col("reporto"),
        "cuasidinero_otros_depositos":     find_col("otros dep",  after=(cuasi_col or 0) - 1),
        "credito_sector_privado_total":    find_col("total",      after=(cred_col  or 0) - 1),
        "credito_sector_privado_cartera":  find_col("cartera"),
        "credito_sector_privado_otros":    find_col("otros",      after=(cred_col  or 0)),
        "tasa_basica":                     find_col("básica", "basica"),
        "tasa_pasiva":                     find_col("pasiva"),
        "tasa_activa":                     find_col("activa (", "activa("),
        "inflacion_mensual":               find_col("mensual",   after=(infl_col or 0) - 1),
        "inflacion_anual":                 find_col("anual",     after=(infl_col or 0) - 1),
        "inflacion_acumulada":             find_col("acumulada"),
    }


def _parse_ims1_sheet(rows: list) -> list[dict]:
    """
    Hoja IMS1 — formato ancho (una fila = un mes).
    Col 0: año (merged cell; None/'' en meses 2..12 del mismo año).
    Col 1: nombre del mes.
    Col 2: días del mes.
    Col 3+: indicadores según _find_ims1_cols.
    """
    cols = _find_ims1_cols(rows)

    # Encontrar primera fila de datos (col 1 es nombre de mes)
    data_start = None
    for ri, row in enumerate(rows):
        if row and len(row) > 1 and str(row[1]).strip() in _MESES:
            data_start = ri
            break
    if data_start is None:
        return []

    def get(row, field):
        ci = cols.get(field)
        if ci is None or ci >= len(row):
            return None
        return _to_float(row[ci])

    # Use dict keyed by (anio, mes) so that later weekly rows overwrite earlier ones.
    # Files may have multiple rows per month (weekly granularity); we keep the last
    # row, which corresponds to the end-of-month observation.
    monthly: dict[tuple, dict] = {}
    current_year: int | None = None

    for row in rows[data_start:]:
        if not row or len(row) < 3:
            continue

        # Año: llenar hacia adelante (celdas combinadas)
        year_raw = row[0]
        if year_raw not in (None, "", 0, 0.0):
            try:
                current_year = int(float(str(year_raw)))
            except (ValueError, TypeError):
                pass

        mes = str(row[1]).strip() if row[1] is not None else ""
        if mes not in _MESES or current_year is None:
            continue

        dias_raw = row[2]
        try:
            dias = int(float(str(dias_raw))) if dias_raw not in (None, "") else None
        except (ValueError, TypeError):
            dias = None

        # Overwrite — last row for (anio, mes) wins (end-of-month or latest week)
        monthly[(current_year, mes)] = {
            "anio":                            current_year,
            "mes":                             mes,
            "dias_mes":                        dias,
            "rild":                            get(row, "rild"),
            "pasivos_monetarios_pm":           get(row, "pasivos_monetarios_pm"),
            "emision_monetaria_em":            get(row, "emision_monetaria_em"),
            "reservas_bancarias_rb":           get(row, "reservas_bancarias_rb"),
            "depositos_vista":                 get(row, "depositos_vista"),
            "cuasidinero_total":               get(row, "cuasidinero_total"),
            "cuasidinero_ahorro":              get(row, "cuasidinero_ahorro"),
            "cuasidinero_plazo":               get(row, "cuasidinero_plazo"),
            "cuasidinero_restringido":         get(row, "cuasidinero_restringido"),
            "cuasidinero_operaciones_reporto": get(row, "cuasidinero_operaciones_reporto"),
            "cuasidinero_otros_depositos":     get(row, "cuasidinero_otros_depositos"),
            "credito_sector_privado_total":    get(row, "credito_sector_privado_total"),
            "credito_sector_privado_cartera":  get(row, "credito_sector_privado_cartera"),
            "credito_sector_privado_otros":    get(row, "credito_sector_privado_otros"),
            "tasa_basica":                     get(row, "tasa_basica"),
            "tasa_pasiva":                     get(row, "tasa_pasiva"),
            "tasa_activa":                     get(row, "tasa_activa"),
            "inflacion_mensual":               get(row, "inflacion_mensual"),
            "inflacion_anual":                 get(row, "inflacion_anual"),
            "inflacion_acumulada":             get(row, "inflacion_acumulada"),
        }

    return list(monthly.values())

# ---------------------------------------------------------------------------
# IMS1.1: detección de columnas y parser ancho
# ---------------------------------------------------------------------------

def _find_ims1_1_cols(rows: list) -> dict[str, int | None]:
    """Mapea nombre-de-campo → índice de columna para la hoja IMS1.1."""
    col_texts: dict[int, str] = {}
    # Start from row 5 to skip the sheet title rows (which also contain "OFERTA MONETARIA M1")
    for ri in range(5, min(12, len(rows))):
        for ci, v in enumerate(rows[ri]):
            s = str(v).strip().lower() if v is not None else ""
            if s:
                col_texts[ci] = col_texts.get(ci, "") + " " + s

    def find_col(*keywords, after: int = -1) -> int | None:
        for ci in sorted(col_texts):
            if ci <= after:
                continue
            t = col_texts[ci]
            if any(kw in t for kw in keywords):
                return ci
        return None

    m1_col = find_col("oferta monetaria m1", "oferta monetaria")

    return {
        "especies_monetarias_circulacion": find_col("especies monetarias"),
        "moneda_fraccionaria":             find_col("moneda fraccionaria"),
        "dinero_electronico":              find_col("dinero electrónico", "dinero electronico"),
        "depositos_vista":                 find_col("a la vista"),
        "oferta_monetaria_m1":             find_col("oferta monetaria m1", "oferta monetaria"),
        "cuasidinero":                     find_col("cuasidinero"),
        "liquidez_total_m2":               find_col("liquidez total m2", "liquidez total"),
        "reservas_bancarias":              find_col("reservas bancarias"),
        "caja_bce":                        find_col("caja bce"),
        "caja_osd":                        find_col("caja osd"),
        "base_monetaria_bm":               find_col("base monetaria"),
        "multiplicador_m1_bm":             find_col("multiplicador  m1", "multiplicador m1",
                                                     after=(m1_col or -1)),
        "multiplicador_m2_bm":             find_col("multiplicador m2"),
    }


def _parse_ims1_1_sheet(rows: list) -> list[dict]:
    """
    Hoja IMS1.1 — formato ancho (una fila = un mes).
    Col 0: año (merged cell). Col 1: nombre del mes. Col 2+: indicadores.
    """
    cols = _find_ims1_1_cols(rows)

    data_start = None
    for ri, row in enumerate(rows):
        if row and len(row) > 1 and str(row[1]).strip() in _MESES:
            data_start = ri
            break
    if data_start is None:
        return []

    def get(row, field):
        ci = cols.get(field)
        if ci is None or ci >= len(row):
            return None
        return _to_float(row[ci])

    monthly: dict[tuple, dict] = {}
    current_year: int | None = None

    for row in rows[data_start:]:
        if not row or len(row) < 2:
            continue

        year_raw = row[0]
        if year_raw not in (None, "", 0, 0.0):
            try:
                current_year = int(float(str(year_raw)))
            except (ValueError, TypeError):
                pass

        mes = str(row[1]).strip() if row[1] is not None else ""
        if mes not in _MESES or current_year is None:
            continue

        monthly[(current_year, mes)] = {
            "anio":                            current_year,
            "mes":                             mes,
            "especies_monetarias_circulacion": get(row, "especies_monetarias_circulacion"),
            "moneda_fraccionaria":             get(row, "moneda_fraccionaria"),
            "dinero_electronico":              get(row, "dinero_electronico"),
            "depositos_vista":                 get(row, "depositos_vista"),
            "oferta_monetaria_m1":             get(row, "oferta_monetaria_m1"),
            "cuasidinero":                     get(row, "cuasidinero"),
            "liquidez_total_m2":              get(row, "liquidez_total_m2"),
            "reservas_bancarias":              get(row, "reservas_bancarias"),
            "caja_bce":                        get(row, "caja_bce"),
            "caja_osd":                        get(row, "caja_osd"),
            "base_monetaria_bm":               get(row, "base_monetaria_bm"),
            "multiplicador_m1_bm":             get(row, "multiplicador_m1_bm"),
            "multiplicador_m2_bm":             get(row, "multiplicador_m2_bm"),
        }

    return list(monthly.values())

# ---------------------------------------------------------------------------
# Parser general (hojas largas)
# ---------------------------------------------------------------------------

def _parse_sheet(rows: list, fecha: date) -> list[dict]:
    """
    Extrae registros (indicador, valor) de una hoja larga.
    El valor es el de la última columna numérica de cada fila de datos.
    """
    if not rows:
        return []

    val_col = None
    best_count = 0
    for row in rows:
        nc = _last_numeric_col(row)
        if nc is not None and nc > best_count:
            best_count = nc
            val_col = nc

    if val_col is None:
        return []

    records = []
    seen_labels: set[str] = set()

    for row in rows:
        if not row:
            continue

        label = _clean_label(row[0] if row else None)
        if not label:
            continue

        # Skip year labels: "2006" (xlsx int) or "2006.0" (xls float via xlrd)
        if re.match(r"^\d{4}(?:\.0)?$", label) or label in (
            "Ene", "Feb", "Mar", "Abr", "May", "Jun",
            "Jul", "Ago", "Sep", "Oct", "Nov", "Dic",
            "Período", "Period",
        ):
            continue

        val = _to_float(row[val_col] if val_col < len(row) else None)

        # Skip header/subtitle rows that carry no numeric value
        if val is None:
            continue

        label_key = label.lower()
        if label_key in seen_labels:
            continue

        records.append({
            "fecha_semana": fecha,
            "anio":         fecha.year,
            "indicador":    label,
            "valor_millones": val,
        })
        seen_labels.add(label_key)

    return records

# ---------------------------------------------------------------------------
# Base de datos
# ---------------------------------------------------------------------------

def _ensure_table(engine, table: str) -> None:
    if not sa_inspect(engine).has_table(table):
        with engine.begin() as conn:
            conn.execute(text(_DDL_TEMPLATE.format(table=table)))
            conn.execute(text(_IDX_TEMPLATE.format(table=table)))
        print(f"  [db] Tabla '{table}' creada.")


def _ensure_ims1_table(engine) -> None:
    if not sa_inspect(engine).has_table(_IMS1_TABLE):
        with engine.begin() as conn:
            conn.execute(text(_IMS1_DDL))
            conn.execute(text(_IMS1_IDX))
        print(f"  [db] Tabla '{_IMS1_TABLE}' creada.")


def _existing_hashes(engine, table: str) -> set:
    if not sa_inspect(engine).has_table(table):
        return set()
    with engine.connect() as conn:
        rows = conn.execute(
            text(f"SELECT hash_registro FROM {table}")
        ).fetchall()
    return {r[0] for r in rows}


def _hash(rec: dict) -> str:
    key = "|".join(str(rec.get(k, "")) for k in
                   ("fecha_semana", "indicador", "valor_millones"))
    return hashlib.sha256(key.encode()).hexdigest()


def _ensure_ims1_1_table(engine) -> None:
    if not sa_inspect(engine).has_table(_IMS1_1_TABLE):
        with engine.begin() as conn:
            conn.execute(text(_IMS1_1_DDL))
            conn.execute(text(_IMS1_1_IDX))
        print(f"  [db] Tabla '{_IMS1_1_TABLE}' creada.")


def _hash_ims1(rec: dict) -> str:
    key = "|".join(str(rec.get(k, "")) for k in ("anio", "mes"))
    return hashlib.sha256(key.encode()).hexdigest()


def _hash_ims1_1(rec: dict) -> str:
    key = "|".join(str(rec.get(k, "")) for k in ("anio", "mes"))
    return hashlib.sha256(key.encode()).hexdigest()

# ---------------------------------------------------------------------------
# Helpers de fecha
# ---------------------------------------------------------------------------

def _date_from_filename(name: str) -> date | None:
    m = _FILE_DATE_RE.search(name)
    if not m:
        return None
    try:
        return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    except ValueError:
        return None

# ---------------------------------------------------------------------------
# Punto de entrada
# ---------------------------------------------------------------------------

def load(file_paths: list) -> None:
    """
    Parsea los XLS/XLSX y carga cada hoja IMS en su tabla correspondiente.
    Idempotente: usa hash_registro para no duplicar registros.
    IMS1 → tabla ancha (anio, mes). Resto → tabla larga (fecha_semana, indicador).
    """
    if not file_paths:
        print("[bce] Sin archivos para procesar.")
        return

    engine = get_master_engine()
    hash_cache: dict[str, set] = {}
    total_inserted = 0

    for path in sorted(file_paths):
        path = Path(path)
        if not path.exists() or path.name.startswith("~"):
            continue

        fecha = _date_from_filename(path.name)
        if not fecha:
            print(f"[bce] [warn] No se pudo extraer fecha de: {path.name}")
            continue

        print(f"[bce] Procesando: {path.name}  ({fecha})")

        sheets = _open_all_sheets(path)
        if not sheets:
            print(f"  [warn] Sin hojas IMS en {path.name}")
            continue

        for sheet_name, rows in sheets.items():

            # ---- IMS1: parser ancho especializado ----
            if sheet_name.strip().upper() == "IMS1":
                _ensure_ims1_table(engine)
                if _IMS1_TABLE not in hash_cache:
                    hash_cache[_IMS1_TABLE] = _existing_hashes(engine, _IMS1_TABLE)

                records = _parse_ims1_sheet(rows)
                if not records:
                    continue

                new_rows = []
                for rec in records:
                    h = _hash_ims1(rec)
                    rec["hash_registro"] = h
                    if h not in hash_cache[_IMS1_TABLE]:
                        new_rows.append(rec)
                        hash_cache[_IMS1_TABLE].add(h)

                if not new_rows:
                    print(f"  IMS1: ya cargado")
                    continue

                df = pd.DataFrame(new_rows)
                df.to_sql(_IMS1_TABLE, engine, if_exists="append", index=False)
                print(f"  IMS1 -> {_IMS1_TABLE}: {len(df)} filas ({fecha.year})")
                total_inserted += len(df)
                continue

            # ---- IMS1.1: parser ancho especializado ----
            if sheet_name.strip().upper() == "IMS1.1":
                _ensure_ims1_1_table(engine)
                if _IMS1_1_TABLE not in hash_cache:
                    hash_cache[_IMS1_1_TABLE] = _existing_hashes(engine, _IMS1_1_TABLE)

                records = _parse_ims1_1_sheet(rows)
                if not records:
                    continue

                new_rows = []
                for rec in records:
                    h = _hash_ims1_1(rec)
                    rec["hash_registro"] = h
                    if h not in hash_cache[_IMS1_1_TABLE]:
                        new_rows.append(rec)
                        hash_cache[_IMS1_1_TABLE].add(h)

                if not new_rows:
                    print(f"  IMS1.1: ya cargado")
                    continue

                df = pd.DataFrame(new_rows)
                df.to_sql(_IMS1_1_TABLE, engine, if_exists="append", index=False)
                print(f"  IMS1.1 -> {_IMS1_1_TABLE}: {len(df)} filas ({fecha.year})")
                total_inserted += len(df)
                continue

            # ---- Resto de hojas IMS: parser largo ----
            table = _table_name(sheet_name)
            _ensure_table(engine, table)

            if table not in hash_cache:
                hash_cache[table] = _existing_hashes(engine, table)

            records = _parse_sheet(rows, fecha)
            if not records:
                continue

            new_rows = []
            for rec in records:
                rec["hash_registro"] = _hash(rec)
                if rec["hash_registro"] not in hash_cache[table]:
                    new_rows.append(rec)
                    hash_cache[table].add(rec["hash_registro"])

            if not new_rows:
                print(f"  {sheet_name}: ya cargado")
                continue

            df = pd.DataFrame(new_rows)
            df["fecha_semana"] = pd.to_datetime(df["fecha_semana"]).dt.date
            df.to_sql(table, engine, if_exists="append", index=False)
            print(f"  {sheet_name} -> {table}: {len(df)} filas")
            total_inserted += len(df)

    print(f"[bce] Carga completada - {total_inserted} filas nuevas en total.")
