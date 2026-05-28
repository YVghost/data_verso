"""
ETL loader: IMAEc — Índice de Actividad Económica Coyuntural (BCE)

Fuente: IMAEc_{YYYYMM}.xlsx  (archivo mensual, contiene datos desde 2018.ene)

Tablas destino:
  imaec_bruto     — hojas _brut_ (datos brutos por actividad y por
                    separación petrolero/no petrolero)
  imaec_ajustado  — hojas _ajus_ (ajustados de estacionalidad,
                    variaciones, contribuciones, acumulados × 2 grupos)

Estructura Excel (todas las hojas de datos):
  Fila 9  : encabezados de actividad económica
  Fila 10+: datos — col A = "YYYY.mes" (ej: 2026.mar), cols B+ = valores

Columna tipo_indice (diferenciadora dentro de cada tabla):
  Bruto   : indica el índice sin mencionar "Datos Brutos"
  Ajustado: indica índice + tipo de ajuste/variación

Deduplicación: SHA-256 sobre (anio, mes, actividad, tipo_indice).
"""

import hashlib
import re
import sys
from pathlib import Path

import openpyxl
from sqlalchemy import text

sys.path.append(str(Path(__file__).resolve().parents[2]))
from utils.base_engine import get_master_engine

_TABLE_BRUTO = "imaec_bruto"
_TABLE_AJUS  = "imaec_ajustado"

_DDL_BRUTO = """
CREATE TABLE imaec_bruto (
    id            BIGINT IDENTITY(1,1) NOT NULL,
    anio          SMALLINT       NOT NULL,
    mes           NVARCHAR(3)    NOT NULL,
    actividad     NVARCHAR(300)  NOT NULL,
    valor         FLOAT          NULL,
    tipo_indice   NVARCHAR(300)  NOT NULL,
    nombre_hoja   NVARCHAR(50)   NOT NULL,
    hash_registro NVARCHAR(64)   NOT NULL,
    fecha_carga   DATETIME2      DEFAULT GETDATE(),
    CONSTRAINT PK_imaec_bruto PRIMARY KEY NONCLUSTERED (id)
)"""

_IDX_BRUTO = (
    "CREATE CLUSTERED INDEX CIX_imaec_bruto "
    "ON imaec_bruto (anio, mes)"
)

_DDL_AJUS = """
CREATE TABLE imaec_ajustado (
    id            BIGINT IDENTITY(1,1) NOT NULL,
    anio          SMALLINT       NOT NULL,
    mes           NVARCHAR(3)    NOT NULL,
    actividad     NVARCHAR(300)  NOT NULL,
    valor         FLOAT          NULL,
    tipo_indice   NVARCHAR(300)  NOT NULL,
    nombre_hoja   NVARCHAR(50)   NOT NULL,
    hash_registro NVARCHAR(64)   NOT NULL,
    fecha_carga   DATETIME2      DEFAULT GETDATE(),
    CONSTRAINT PK_imaec_ajustado PRIMARY KEY NONCLUSTERED (id)
)"""

_IDX_AJUS = (
    "CREATE CLUSTERED INDEX CIX_imaec_ajustado "
    "ON imaec_ajustado (anio, mes)"
)

# ---------------------------------------------------------------------------
# Mapping hoja → (tabla, tipo_indice)
# Las hojas _brut_ no incluyen "Datos Brutos" en el tipo_indice
# (ya está implícito en la tabla imaec_bruto).
# Las hojas _pet son la misma métrica pero con desagregación
# Petrolero / No Petrolero en las columnas.
# ---------------------------------------------------------------------------
_SHEET_MAP: dict[str, tuple[str, str]] = {
    # ── Datos Brutos ─────────────────────────────────────────────────────────
    "Of_Ivol_brut_clas":        (_TABLE_BRUTO, "Índices de Volumen Encadenados, 2018=100"),
    "Of_Ivol_brut_clas_pet":    (_TABLE_BRUTO, "Índices de Volumen Encadenados, 2018=100, Petrolero/No Petrolero"),
    # ── Datos Ajustados de Estacionalidad ────────────────────────────────────
    "Of_Ivol_ajus_clas":        (_TABLE_AJUS,  "Índices de Volumen Encadenados, 2018=100, Datos Ajustados de Estacionalidad"),
    "Of_Ivol_ajus_clas_vY":     (_TABLE_AJUS,  "Índices de Volumen Encadenados, 2018=100, Datos Ajustados de Estacionalidad, Variación Interanual"),
    "Of_Ivol_ajus_cont_vY":     (_TABLE_AJUS,  "Contribución al Crecimiento, Datos Ajustados de Estacionalidad, Variación Interanual"),
    "Of_Ivol_ajus_acum_vY":     (_TABLE_AJUS,  "Índices de Volumen Encadenados, 2018=100, Datos Ajustados de Estacionalidad, Variación Acumulada Interanual"),
    "Of_Ivol_ajus_clas_pet":    (_TABLE_AJUS,  "Índices de Volumen Encadenados, 2018=100, Datos Ajustados de Estacionalidad, Petrolero/No Petrolero"),
    "Of_Ivol_ajus_clas_vY_pet": (_TABLE_AJUS,  "Índices de Volumen Encadenados, 2018=100, Datos Ajustados de Estacionalidad, Variación Interanual, Petrolero/No Petrolero"),
    "Of_Ivol_ajus_cont_vY_pet": (_TABLE_AJUS,  "Contribución al Crecimiento, Datos Ajustados de Estacionalidad, Variación Interanual, Petrolero/No Petrolero"),
    "Of_Ivol_ajus_acum_vY_pet": (_TABLE_AJUS,  "Índices de Volumen Encadenados, 2018=100, Datos Ajustados de Estacionalidad, Variación Acumulada Interanual, Petrolero/No Petrolero"),
}

# YYYY.mes  (ej: 2026.mar)
_FECHA_RE = re.compile(r"^(\d{4})\.(ene|feb|mar|abr|may|jun|jul|ago|sep|oct|nov|dic)$",
                       re.IGNORECASE)

_HDR_ROW        = 9
_DATA_ROW_START = 10
_BATCH          = 2000


# ---------------------------------------------------------------------------
# Punto de entrada
# ---------------------------------------------------------------------------

def load(files: list[Path]) -> None:
    if not files:
        print("[imaec] Sin archivos para cargar.")
        return

    engine = get_master_engine()
    _ensure_tables(engine)

    ex_bruto = _get_existing_hashes(engine, _TABLE_BRUTO)
    ex_ajus  = _get_existing_hashes(engine, _TABLE_AJUS)

    tot_bruto = tot_ajus = 0

    for path in sorted(files):
        nb, na = _load_file(path, engine, ex_bruto, ex_ajus)
        tot_bruto += nb
        tot_ajus  += na

    print(
        f"[imaec] Total: {tot_bruto} en {_TABLE_BRUTO}, "
        f"{tot_ajus} en {_TABLE_AJUS}."
    )


# ---------------------------------------------------------------------------
# DDL
# ---------------------------------------------------------------------------

def _ensure_tables(engine) -> None:
    from sqlalchemy import inspect as sa_inspect
    insp = sa_inspect(engine)
    if not insp.has_table(_TABLE_BRUTO):
        with engine.begin() as conn:
            conn.execute(text(_DDL_BRUTO))
            conn.execute(text(_IDX_BRUTO))
        print(f"[imaec] Tabla {_TABLE_BRUTO} creada.")
    if not insp.has_table(_TABLE_AJUS):
        with engine.begin() as conn:
            conn.execute(text(_DDL_AJUS))
            conn.execute(text(_IDX_AJUS))
        print(f"[imaec] Tabla {_TABLE_AJUS} creada.")


def _get_existing_hashes(engine, table: str) -> set[str]:
    with engine.connect() as conn:
        rows = conn.execute(
            text(f"SELECT hash_registro FROM {table}")
        ).fetchall()
    return {r[0] for r in rows}


# ---------------------------------------------------------------------------
# Por archivo
# ---------------------------------------------------------------------------

def _load_file(
    path: Path,
    engine,
    ex_bruto: set[str],
    ex_ajus:  set[str],
) -> tuple[int, int]:
    print(f"[imaec] Procesando {path.name}...")
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as ex:
        print(f"[imaec] Error abriendo {path.name}: {ex}")
        return 0, 0

    n_bruto = n_ajus = 0

    for sh_name in wb.sheetnames:
        if sh_name not in _SHEET_MAP:
            continue
        table, tipo_indice = _SHEET_MAP[sh_name]
        ws = wb[sh_name]

        records = _parse_sheet(ws, tipo_indice, sh_name)
        if not records:
            continue

        existing = ex_bruto if table == _TABLE_BRUTO else ex_ajus
        new = [r for r in records if r["hash_registro"] not in existing]
        if not new:
            continue

        for i in range(0, len(new), _BATCH):
            _insert(new[i : i + _BATCH], table, engine)
        existing.update(r["hash_registro"] for r in new)

        if table == _TABLE_BRUTO:
            n_bruto += len(new)
        else:
            n_ajus += len(new)
        print(f"[imaec]   {sh_name}: {len(new)} filas nuevas.")

    wb.close()
    return n_bruto, n_ajus


# ---------------------------------------------------------------------------
# Parsing de hoja
# ---------------------------------------------------------------------------

def _parse_sheet(ws, tipo_indice: str, sh_name: str) -> list[dict]:
    hdr = next(ws.iter_rows(min_row=_HDR_ROW, max_row=_HDR_ROW, values_only=True))
    actividades = [str(v).strip() for v in hdr[1:] if v is not None]

    records: list[dict] = []

    for row in ws.iter_rows(min_row=_DATA_ROW_START, values_only=True):
        fecha_raw = row[0]
        if fecha_raw is None:
            continue

        m = _FECHA_RE.match(str(fecha_raw).strip())
        if not m:
            continue

        anio = int(m.group(1))
        mes  = m.group(2).lower()
        values = row[1:]

        for i, act in enumerate(actividades):
            if i >= len(values):
                break
            v = values[i]
            if v is None:
                continue
            try:
                valor = float(v)
            except (TypeError, ValueError):
                continue

            records.append({
                "anio":          anio,
                "mes":           mes,
                "actividad":     act,
                "valor":         valor,
                "tipo_indice":   tipo_indice,
                "nombre_hoja":   sh_name,
                "hash_registro": _hash(anio, mes, act, tipo_indice),
            })

    return records


# ---------------------------------------------------------------------------
# Inserción e helpers
# ---------------------------------------------------------------------------

def _insert(records: list[dict], table: str, engine) -> None:
    cols         = list(records[0].keys())
    col_list     = ", ".join(cols)
    placeholders = ", ".join(f":{c}" for c in cols)
    with engine.begin() as conn:
        conn.execute(
            text(f"INSERT INTO {table} ({col_list}) VALUES ({placeholders})"),
            records,
        )


def _hash(anio: int, mes: str, actividad: str, tipo_indice: str) -> str:
    key = f"{anio}|{mes}|{actividad}|{tipo_indice}"
    return hashlib.sha256(key.encode()).hexdigest()
