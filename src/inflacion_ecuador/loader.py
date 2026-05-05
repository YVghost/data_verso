"""
Parser de XLS/XLSX de inflación IPC — INEC

Archivos y tablas:
  SERIE HISTORICA IPC_*.xls        (layout año×mes pivot)
    '2. VARIACIÓN MENSUAL'  → inflacion_ecuador_variacion_mensual
    '3. VARIACIÓN ANUAL'    → inflacion_ecuador_variacion_anual

  ipc_indicadores_descriptivos_*.xlsx  (layout mes×categoría wide)
    '2. VARIA.MENSUAL'      → inflacion_ecuador_indicadores_variacion_mensual
    '3. VARIA.ANUAL'        → inflacion_ecuador_indicadores_variacion_anual

  ipc_incid_nac_div_*.xlsx  (layout división×período long)
    '1.INCID. MENSUAL'      → inflacion_ecuador_series_incidencias_mensual
    '2.INCID. ANUAL'        → inflacion_ecuador_series_incidencias_anual

  ipc_var_men_nac_reg_ciud[_emp]_*.xlsx  (layout CCIF×período, multi-hoja por región)
  ipc_var_anu_nac_reg_ciud[_emp]_*.xlsx
    todas las hojas (excl. Contenido, Esmeraldas, Machala, Sto. Domingo)
    → inflacion_ecuador_series_ipc_mensual  (col es_empalmada BIT)
    → inflacion_ecuador_series_ipc_anual    (col es_empalmada BIT)
"""

import re
import sys
import hashlib
import datetime as dt
import pandas as pd
from datetime import date
from pathlib import Path
from sqlalchemy import text, inspect as sa_inspect

sys.path.append(str(Path(__file__).resolve().parents[2]))
from utils.base_engine import get_master_engine

# ---------------------------------------------------------------------------
# Routing
# ---------------------------------------------------------------------------

_SERIE_HIST_RE  = re.compile(r"SERIE\s*HISTORICA\s*IPC.*\.(xls|xlsx)$",       re.IGNORECASE)
_INDICADORES_RE = re.compile(r"ipc_indicadores_descriptivos.*\.(xls|xlsx)$",  re.IGNORECASE)
_INCIDENCIAS_RE = re.compile(r"ipc_incid_nac_div.*\.(xls|xlsx)$",             re.IGNORECASE)
_VAR_MEN_RE     = re.compile(r"ipc_var_men_nac_reg_ciud.*\.(xls|xlsx)$",      re.IGNORECASE)
_VAR_ANU_RE     = re.compile(r"ipc_var_anu_nac_reg_ciud.*\.(xls|xlsx)$",      re.IGNORECASE)

# Hojas a excluir en archivos de series IPC por región
_SKIP_SHEET_PATTERNS = ["contenido", "esmeraldas", "machala", "sto. domingo", "sto domingo"]

_FILE_SHEETS = {
    "serie": [
        ("VARIACIÓN MENSUAL", "serie_mensual"),
        ("VARIACIÓN ANUAL",   "serie_anual"),
    ],
    "indicadores": [
        ("VARIA.MENSUAL", "ind_mensual"),
        ("VARIA.ANUAL",   "ind_anual"),
    ],
    "incidencias": [
        ("INCID. MENSUAL", "incid_mensual"),
        ("INCID. ANUAL",   "incid_anual"),
    ],
}

# ---------------------------------------------------------------------------
# DDL
# ---------------------------------------------------------------------------

_DDL_SERIE = """
CREATE TABLE {table} (
    id              BIGINT IDENTITY(1,1) NOT NULL,
    fecha           DATE           NOT NULL,
    anio            INT            NOT NULL,
    mes             SMALLINT       NOT NULL,
    nombre_mes      NVARCHAR(20)   NOT NULL,
    variacion       FLOAT          NULL,
    promedio_anual  FLOAT          NULL,
    hash_registro   NVARCHAR(64)   NOT NULL,
    fecha_carga     DATETIME2      DEFAULT GETDATE(),
    CONSTRAINT PK_{table} PRIMARY KEY NONCLUSTERED (id)
)
"""

_DDL_INDICADORES = """
CREATE TABLE {table} (
    id                  BIGINT IDENTITY(1,1) NOT NULL,
    fecha               DATE           NOT NULL,
    anio                INT            NOT NULL,
    mes                 SMALLINT       NOT NULL,
    general             FLOAT          NULL,
    alimentos           FLOAT          NULL,
    no_alimentos        FLOAT          NULL,
    prod_transables     FLOAT          NULL,
    prod_no_transables  FLOAT          NULL,
    prod_con_iva        FLOAT          NULL,
    prod_sin_iva        FLOAT          NULL,
    bienes              FLOAT          NULL,
    servicios           FLOAT          NULL,
    hash_registro       NVARCHAR(64)   NOT NULL,
    fecha_carga         DATETIME2      DEFAULT GETDATE(),
    CONSTRAINT PK_{table} PRIMARY KEY NONCLUSTERED (id)
)
"""

# Long: una fila por (período × división CCIF)
_DDL_INCIDENCIAS = """
CREATE TABLE {table} (
    id              BIGINT IDENTITY(1,1) NOT NULL,
    fecha           DATE           NOT NULL,
    anio            INT            NOT NULL,
    mes             SMALLINT       NOT NULL,
    periodo         NVARCHAR(10)   NOT NULL,
    nivel           NVARCHAR(50)   NULL,
    cod_ccif        NVARCHAR(5)    NOT NULL,
    descripcion     NVARCHAR(200)  NOT NULL,
    ponderacion     FLOAT          NULL,
    incidencia      FLOAT          NULL,
    hash_registro   NVARCHAR(64)   NOT NULL,
    fecha_carga     DATETIME2      DEFAULT GETDATE(),
    CONSTRAINT PK_{table} PRIMARY KEY NONCLUSTERED (id)
)
"""

_DDL_SERIES_IPC = """
CREATE TABLE {table} (
    id              BIGINT IDENTITY(1,1) NOT NULL,
    fecha           DATE           NOT NULL,
    anio            INT            NOT NULL,
    mes             SMALLINT       NOT NULL,
    periodo         NVARCHAR(10)   NOT NULL,
    region          NVARCHAR(60)   NOT NULL,
    nivel           NVARCHAR(20)   NULL,
    cod_ccif        NVARCHAR(20)   NOT NULL,
    descripcion     NVARCHAR(300)  NOT NULL,
    variacion       FLOAT          NULL,
    es_empalmada    NVARCHAR(2)    NOT NULL,
    hash_registro   NVARCHAR(64)   NOT NULL,
    fecha_carga     DATETIME2      DEFAULT GETDATE(),
    CONSTRAINT PK_{table} PRIMARY KEY NONCLUSTERED (id)
)
"""

_IDX_DDL         = "CREATE CLUSTERED INDEX IX_{table}_anio_mes ON {table} (anio, mes)"
_IDX_DDL_INCID   = "CREATE CLUSTERED INDEX IX_{table}_anio_mes_ccif ON {table} (anio, mes, cod_ccif)"
_IDX_DDL_IPC     = "CREATE CLUSTERED INDEX IX_{table}_anio_mes ON {table} (anio, mes)"

_TABLE_INFO = {
    "serie_mensual":  ("inflacion_ecuador_variacion_mensual",               _DDL_SERIE,        _IDX_DDL),
    "serie_anual":    ("inflacion_ecuador_variacion_anual",                 _DDL_SERIE,        _IDX_DDL),
    "ind_mensual":    ("inflacion_ecuador_indicadores_variacion_mensual",   _DDL_INDICADORES,  _IDX_DDL),
    "ind_anual":      ("inflacion_ecuador_indicadores_variacion_anual",     _DDL_INDICADORES,  _IDX_DDL),
    "incid_mensual":  ("inflacion_ecuador_series_incidencias_mensual",      _DDL_INCIDENCIAS,  _IDX_DDL_INCID),
    "incid_anual":    ("inflacion_ecuador_series_incidencias_anual",        _DDL_INCIDENCIAS,  _IDX_DDL_INCID),
    "ipc_mensual":    ("inflacion_ecuador_series_ipc_mensual",              _DDL_SERIES_IPC,   _IDX_DDL_IPC),
    "ipc_anual":      ("inflacion_ecuador_series_ipc_anual",                _DDL_SERIES_IPC,   _IDX_DDL_IPC),
}

# Mapeo de columna Excel → nombre BD (indicadores)
_COL_MAP = {
    "general":                 "general",
    "alimentos":               "alimentos",
    "no alimentos":            "no_alimentos",
    "productos transables":    "prod_transables",
    "productos no transables": "prod_no_transables",
    "productos con iva":       "prod_con_iva",
    "productos sin iva":       "prod_sin_iva",
    "bienes":                  "bienes",
    "servicios":               "servicios",
}
_IND_COLS = list(_COL_MAP.values())

# Abreviaturas de mes en español
_MES_ABR = {
    "ene": 1, "feb": 2, "mar": 3, "abr": 4,
    "may": 5, "jun": 6, "jul": 7, "ago": 8,
    "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dic": 12,
}


# ---------------------------------------------------------------------------
# Punto de entrada
# ---------------------------------------------------------------------------

def load(xls_paths: list) -> None:
    if not xls_paths:
        print("[inflacion] Sin archivos para procesar.")
        return

    engine = get_master_engine()
    _ensure_tables(engine)

    for xls_path in xls_paths:
        _process_file(Path(xls_path), engine)


# ---------------------------------------------------------------------------
# Creación de tablas
# ---------------------------------------------------------------------------

def _ensure_tables(engine) -> None:
    inspector = sa_inspect(engine)
    with engine.begin() as conn:
        for key, (table, ddl, idx_ddl) in _TABLE_INFO.items():
            if inspector.has_table(table):
                if key.startswith("ind_"):
                    existing_cols = {c["name"] for c in inspector.get_columns(table)}
                    if "general" not in existing_cols:
                        conn.execute(text(f"DROP TABLE {table}"))
                        print(f"[inflacion] Tabla antigua eliminada (schema incorrecto): {table}")
                        conn.execute(text(ddl.format(table=table)))
                        conn.execute(text(idx_ddl.format(table=table)))
                        print(f"[inflacion] Tabla recreada: {table}")
            else:
                conn.execute(text(ddl.format(table=table)))
                conn.execute(text(idx_ddl.format(table=table)))
                print(f"[inflacion] Tabla creada: {table}")


# ---------------------------------------------------------------------------
# Procesamiento de un archivo
# ---------------------------------------------------------------------------

def _file_type(path: Path) -> str | None:
    name = path.name
    if _SERIE_HIST_RE.match(name):   return "serie"
    if _INDICADORES_RE.match(name):  return "indicadores"
    if _INCIDENCIAS_RE.match(name):  return "incidencias"
    if _VAR_MEN_RE.match(name):      return "var_men"
    if _VAR_ANU_RE.match(name):      return "var_anu"
    return None


def _process_file(path: Path, engine) -> None:
    ftype = _file_type(path)
    if ftype is None:
        print(f"[inflacion] [skip] No reconocido: {path.name}")
        return

    print(f"[inflacion] Procesando ({ftype}): {path.name}")

    # Series IPC por región — flujo multi-hoja propio
    if ftype in ("var_men", "var_anu"):
        _process_series_ipc(path, ftype, engine)
        return

    try:
        sheets = _open_workbook(path)
    except Exception as ex:
        print(f"  Error abriendo archivo: {ex}")
        return

    parse_fn = {
        "serie":       _parse_serie_rows,
        "indicadores": _parse_indicadores_rows,
        "incidencias": _parse_incidencias_rows,
    }[ftype]

    for sheet_pattern, table_key in _FILE_SHEETS[ftype]:
        rows = _find_sheet_rows(sheets, sheet_pattern)
        if rows is None:
            print(f"  [warn] Hoja '{sheet_pattern}' no encontrada")
            continue

        df = parse_fn(rows)
        if df.empty:
            print(f"  [skip] Sin datos en hoja '{sheet_pattern}'")
            continue

        table, _, _ = _TABLE_INFO[table_key]
        _insert_new(engine, table, df)


def _process_series_ipc(path: Path, ftype: str, engine) -> None:
    """Procesa archivos de variación IPC por región (mensual o anual, normal o empalmado)."""
    es_empalmada = "Si" if "emp" in path.name.lower() else "No"
    table_key    = "ipc_mensual" if ftype == "var_men" else "ipc_anual"
    table, _, _  = _TABLE_INFO[table_key]

    try:
        sheets = _open_workbook(path)
    except Exception as ex:
        print(f"  Error abriendo archivo: {ex}")
        return

    all_dfs = []
    for sheet_name, rows in sheets.items():
        if _skip_ipc_sheet(sheet_name):
            continue

        region = _sheet_to_region(sheet_name)
        df = _parse_series_ipc_rows(rows, region, es_empalmada)
        if not df.empty:
            print(f"  {sheet_name}: {len(df)} registros")
            all_dfs.append(df)

    if not all_dfs:
        print(f"  [skip] Sin datos")
        return

    combined = pd.concat(all_dfs, ignore_index=True)
    _insert_new(engine, table, combined)


# ---------------------------------------------------------------------------
# Lectura agnóstica de formato
# ---------------------------------------------------------------------------

def _open_workbook(path: Path) -> dict:
    """Devuelve {sheet_name: [[valores], ...]} para todas las hojas."""
    suffix = path.suffix.lower()

    if suffix == ".xls":
        import xlrd
        wb = xlrd.open_workbook(str(path))
        result = {}
        for sh in wb.sheets():
            result[sh.name] = [
                [_xlrd_val(sh, r, c) for c in range(sh.ncols)]
                for r in range(sh.nrows)
            ]
        return result

    if suffix == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        result = {}
        for name in wb.sheetnames:
            result[name] = [[cell.value for cell in row]
                            for row in wb[name].iter_rows()]
        wb.close()
        return result

    raise ValueError(f"Formato no soportado: {path.suffix}")


def _xlrd_val(sh, row: int, col: int):
    import xlrd
    ctype = sh.cell_type(row, col)
    if ctype in (xlrd.XL_CELL_NUMBER, xlrd.XL_CELL_TEXT):
        return sh.cell_value(row, col)
    return None


def _find_sheet_rows(sheets: dict, pattern: str) -> list | None:
    pattern_upper = pattern.upper()
    for name, rows in sheets.items():
        if pattern_upper in name.upper():
            return rows
    return None


# ---------------------------------------------------------------------------
# Parser 1: SERIE HISTORICA — pivot año×mes
# ---------------------------------------------------------------------------

def _parse_serie_rows(rows: list) -> pd.DataFrame:
    if len(rows) < 7:
        return pd.DataFrame()

    header = rows[4]
    month_names = []
    for c in range(1, 13):
        raw = str(header[c]).strip() if c < len(header) and header[c] is not None else ""
        month_names.append(raw.capitalize() if raw and raw != "None" else f"Mes{c}")

    records = []
    for row in rows[6:]:
        if not row or row[0] is None:
            continue
        try:
            year = int(float(row[0]))
        except (ValueError, TypeError):
            continue
        if not (1900 <= year <= 2100):
            continue

        promedio_anual = _num_or_none(row[13] if len(row) > 13 else None)

        for mes_idx, nombre_mes in enumerate(month_names, start=1):
            records.append({
                "fecha":          date(year, mes_idx, 1),
                "anio":           year,
                "mes":            mes_idx,
                "nombre_mes":     nombre_mes,
                "variacion":      _num_or_none(row[mes_idx] if mes_idx < len(row) else None),
                "promedio_anual": promedio_anual,
            })

    if not records:
        return pd.DataFrame()

    df = pd.DataFrame(records)
    df["hash_registro"] = df.apply(
        lambda r: hashlib.sha256(
            f"{r['anio']}|{r['mes']}|{r['variacion']}|{r['promedio_anual']}".encode()
        ).hexdigest(), axis=1,
    )
    return df


# ---------------------------------------------------------------------------
# Parser 2: INDICADORES — wide mes×categoría
# ---------------------------------------------------------------------------

def _parse_indicadores_rows(rows: list) -> pd.DataFrame:
    header_idx = None
    for i, row in enumerate(rows):
        if row and isinstance(row[0], str) and row[0].strip().lower() == "mes":
            header_idx = i
            break

    if header_idx is None:
        return pd.DataFrame()

    header = rows[header_idx]
    col_mapping = {}
    for c in range(1, len(header)):
        val = header[c]
        if val is None:
            continue
        normalized = re.sub(r'\*+', '', str(val)).strip().lower()
        db_name = _COL_MAP.get(normalized)
        if db_name:
            col_mapping[c] = db_name

    if not col_mapping:
        return pd.DataFrame()

    records = []
    for row in rows[header_idx + 1:]:
        if not row or row[0] is None:
            break
        if not isinstance(row[0], dt.datetime):
            continue

        d = row[0]
        record = {"fecha": date(d.year, d.month, 1), "anio": d.year, "mes": d.month}
        for col_idx, db_col in col_mapping.items():
            record[db_col] = _num_or_none(row[col_idx] if col_idx < len(row) else None)
        for col in _IND_COLS:
            record.setdefault(col, None)

        records.append(record)

    if not records:
        return pd.DataFrame()

    df = pd.DataFrame(records)
    df["hash_registro"] = df.apply(
        lambda r: hashlib.sha256(
            "|".join([str(r["anio"]), str(r["mes"])] +
                     [str(r.get(c)) for c in _IND_COLS]).encode()
        ).hexdigest(), axis=1,
    )
    return df


# ---------------------------------------------------------------------------
# Parser 3: INCIDENCIAS — pivot división×período → long
#   Fila 4: "Nivel", "Ponderación", "Cód. CCIF", "Descripción", periodo1, periodo2, ...
#   Filas 5+: una fila por división CCIF con sus valores por período
# ---------------------------------------------------------------------------

def _parse_incidencias_rows(rows: list) -> pd.DataFrame:
    if len(rows) < 6:
        return pd.DataFrame()

    # Fila 4 es el encabezado
    header = rows[4]

    # Recolectar columnas de período (col 4 en adelante)
    period_cols = []  # [(col_idx, periodo_str, year, mes_num)]
    for c in range(4, len(header)):
        val = header[c]
        if val is None:
            continue
        parsed = _parse_period(str(val).strip())
        if parsed[0] is not None:
            period_cols.append((c, str(val).strip(), parsed[0], parsed[1]))

    if not period_cols:
        return pd.DataFrame()

    records = []
    for row in rows[5:]:
        if not row or row[0] is None:
            continue
        # Filtrar filas de notas/fuente (sin código CCIF)
        cod = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        desc = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
        if not cod or not desc:
            continue

        nivel      = str(row[0]).strip() if row[0] is not None else None
        ponderacion = _num_or_none(row[1] if len(row) > 1 else None)

        for col_idx, periodo, year, mes_num in period_cols:
            records.append({
                "fecha":       date(year, mes_num, 1),
                "anio":        year,
                "mes":         mes_num,
                "periodo":     periodo,
                "nivel":       nivel,
                "cod_ccif":    cod,
                "descripcion": desc,
                "ponderacion": ponderacion,
                "incidencia":  _num_or_none(row[col_idx] if col_idx < len(row) else None),
            })

    if not records:
        return pd.DataFrame()

    df = pd.DataFrame(records)
    df["hash_registro"] = df.apply(
        lambda r: hashlib.sha256(
            f"{r['anio']}|{r['mes']}|{r['cod_ccif']}|{r['incidencia']}".encode()
        ).hexdigest(), axis=1,
    )
    return df


# ---------------------------------------------------------------------------
# Parser 4: SERIES IPC POR REGIÓN — pivot CCIF×período → long
#   Fila 4: "Nivel", "Cód. CCIF", "Descripción CCIF", periodo1, periodo2, ...
#   Filas 5+: una fila por categoría CCIF con sus valores en cada período
# ---------------------------------------------------------------------------

def _skip_ipc_sheet(sheet_name: str) -> bool:
    lower = sheet_name.lower()
    return any(p in lower for p in _SKIP_SHEET_PATTERNS)


def _sheet_to_region(sheet_name: str) -> str:
    """'2. REGION SIERRA' → 'Region Sierra' | '9. Quito' → 'Quito'"""
    name = re.sub(r'^\d+\.\s*', '', sheet_name).strip()
    return name.title() if name.isupper() else name


def _parse_series_ipc_rows(rows: list, region: str, es_empalmada: int) -> pd.DataFrame:
    """
    Convierte el pivot CCIF×período en formato long.
    Fila 4: encabezado con períodos en cols 3+
    Filas 5+: datos por categoría CCIF
    """
    if len(rows) < 6:
        return pd.DataFrame()

    header = rows[4]

    # Recolectar columnas de período (col 3 en adelante)
    period_cols = []
    for c in range(3, len(header)):
        val = header[c]
        if val is None:
            continue
        parsed = _parse_period(str(val).strip())
        if parsed[0] is not None:
            period_cols.append((c, str(val).strip(), parsed[0], parsed[1]))

    if not period_cols:
        return pd.DataFrame()

    records = []
    for row in rows[5:]:
        if not row or row[0] is None:
            continue
        cod  = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        desc = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        if not cod or not desc:
            continue

        nivel = str(row[0]).strip() if row[0] is not None else None

        for col_idx, periodo, year, mes_num in period_cols:
            records.append({
                "fecha":        date(year, mes_num, 1),
                "anio":         year,
                "mes":          mes_num,
                "periodo":      periodo,
                "region":       region,
                "nivel":        nivel,
                "cod_ccif":     cod,
                "descripcion":  desc,
                "variacion":    _num_or_none(row[col_idx] if col_idx < len(row) else None),
                "es_empalmada": es_empalmada,
            })

    if not records:
        return pd.DataFrame()

    df = pd.DataFrame(records)
    df["hash_registro"] = df.apply(
        lambda r: hashlib.sha256(
            f"{r['anio']}|{r['mes']}|{r['region']}|{r['cod_ccif']}|{r['es_empalmada']}|{r['variacion']}".encode()
        ).hexdigest(), axis=1,
    )
    return df


def _parse_period(s: str):
    """'feb-06' → (2006, 2) | 'ene-15' → (2015, 1). Retorna (None, None) si no parsea."""
    m = re.match(r"([a-z]+)-(\d{2})$", s.strip().lower())
    if not m:
        return None, None
    mes_str = m.group(1)
    mes_num = _MES_ABR.get(mes_str) or _MES_ABR.get(mes_str[:3])
    if not mes_num:
        return None, None
    yr = int(m.group(2))
    year = 2000 + yr if yr <= 30 else 1900 + yr
    return year, mes_num


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _num_or_none(value):
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# Inserción deduplicada
# ---------------------------------------------------------------------------

def _insert_new(engine, table: str, df: pd.DataFrame) -> None:
    with engine.connect() as conn:
        existing = pd.read_sql(
            text(f"SELECT hash_registro FROM {table}"), conn
        )

    existing_hashes = set(existing["hash_registro"].tolist())
    new_rows = df[~df["hash_registro"].isin(existing_hashes)].copy()

    if new_rows.empty:
        print(f"  [{table}] Sin registros nuevos.")
        return

    new_rows = new_rows.sort_values(["anio", "mes"]).reset_index(drop=True)
    new_rows.to_sql(table, engine, if_exists="append", index=False)
    print(f"  [{table}] {len(new_rows)} filas nuevas cargadas.")
