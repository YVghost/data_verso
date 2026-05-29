"""
ETL loader: Captaciones Mutualistas — SEPS Ecuador

Tablas:
  mutualistas_captaciones          — reportes Mut  (xlsm/xlsx/txt/csv)
  mutualistas_captaciones_sectores — reportes S1/S2/S3 + columna sector
  mutualistas_captaciones_bruto    — bases granulares (xlsm/xlsx/txt/csv)

Cada ZIP se extrae primero a disco (downloads/.../extracted/{nombre}/)
y luego se procesan los archivos extraídos.  Los archivos ya extraídos
no se vuelven a extraer en ejecuciones posteriores.

ZIPs anidados (anterior.zip) se extraen en dos pasadas:
  1. anterior.zip → extracted/anterior/
  2. 2016.zip, 2017.zip … dentro → extracted/anterior/2016/ etc.

Deduplicacion: hash SHA-256 por registro. Re-ejecutar es idempotente.
"""

import hashlib
import re
import shutil
import sys
import unicodedata
import zipfile
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from sqlalchemy import text

sys.path.append(str(Path(__file__).resolve().parents[2]))
from utils.base_engine import get_master_engine

_CHUNKSIZE = 50_000

# ---------------------------------------------------------------------------
# DDL — mutualistas_captaciones
# ---------------------------------------------------------------------------

_TABLE_MUT = "mutualistas_captaciones"

_DDL_MUT = """
CREATE TABLE mutualistas_captaciones (
    id               BIGINT IDENTITY(1,1) NOT NULL,
    anio             SMALLINT       NOT NULL,
    saldo            FLOAT          NULL,
    num_clientes     INT            NULL,
    num_cuentas      INT            NULL,
    fecha_corte      DATE           NOT NULL,
    region           NVARCHAR(50)   NULL,
    provincia        NVARCHAR(100)  NULL,
    canton           NVARCHAR(200)  NULL,
    tipo_deposito    NVARCHAR(100)  NULL,
    estado_operacion NVARCHAR(200)  NULL,
    ruc              NVARCHAR(15)   NULL,
    razon_social     NVARCHAR(200)  NULL,
    hash_registro    NVARCHAR(64)   NOT NULL,
    fecha_carga      DATETIME2      DEFAULT GETDATE(),
    CONSTRAINT PK_mutualistas_captaciones PRIMARY KEY NONCLUSTERED (id)
)"""
_IDX_MUT = ("CREATE CLUSTERED INDEX CIX_mutualistas_captaciones "
            "ON mutualistas_captaciones (anio, fecha_corte)")

# ---------------------------------------------------------------------------
# DDL — mutualistas_captaciones_sectores
# ---------------------------------------------------------------------------

_TABLE_SEC = "mutualistas_captaciones_sectores"

_DDL_SEC = """
CREATE TABLE mutualistas_captaciones_sectores (
    id               BIGINT IDENTITY(1,1) NOT NULL,
    anio             SMALLINT       NOT NULL,
    sector           TINYINT        NOT NULL,
    saldo            FLOAT          NULL,
    num_clientes     INT            NULL,
    num_cuentas      INT            NULL,
    fecha_corte      DATE           NOT NULL,
    region           NVARCHAR(50)   NULL,
    provincia        NVARCHAR(100)  NULL,
    canton           NVARCHAR(200)  NULL,
    tipo_deposito    NVARCHAR(100)  NULL,
    estado_operacion NVARCHAR(200)  NULL,
    ruc              NVARCHAR(15)   NULL,
    razon_social     NVARCHAR(200)  NULL,
    hash_registro    NVARCHAR(64)   NOT NULL,
    fecha_carga      DATETIME2      DEFAULT GETDATE(),
    CONSTRAINT PK_mutualistas_captaciones_sectores PRIMARY KEY NONCLUSTERED (id)
)"""
_IDX_SEC = ("CREATE CLUSTERED INDEX CIX_mutualistas_captaciones_sectores "
            "ON mutualistas_captaciones_sectores (anio, sector, fecha_corte)")

# ---------------------------------------------------------------------------
# DDL — mutualistas_captaciones_bruto
# ---------------------------------------------------------------------------

_TABLE_BRUTO = "mutualistas_captaciones_bruto"

_DDL_BRUTO = """
CREATE TABLE mutualistas_captaciones_bruto (
    id                BIGINT IDENTITY(1,1) NOT NULL,
    anio              SMALLINT       NOT NULL,
    tipo_persona      NVARCHAR(50)   NULL,
    fecha_corte       DATE           NOT NULL,
    segmento          NVARCHAR(50)   NULL,
    provincia         NVARCHAR(100)  NULL,
    canton            NVARCHAR(200)  NULL,
    parroquia         NVARCHAR(200)  NULL,
    dpaparroquia      NVARCHAR(20)   NULL,
    estado_operacion  NVARCHAR(100)  NULL,
    tipo_cuenta       NVARCHAR(100)  NULL,
    banda_maduracion  NVARCHAR(100)  NULL,
    sexo              NVARCHAR(20)   NULL,
    rango_edad        NVARCHAR(50)   NULL,
    nivel_instruccion NVARCHAR(100)  NULL,
    rango_saldo       NVARCHAR(200)  NULL,
    nro_cuentas       INT            NULL,
    saldo_usd         FLOAT          NULL,
    nro_depositantes  INT            NULL,
    hash_registro     NVARCHAR(64)   NOT NULL,
    fecha_carga       DATETIME2      DEFAULT GETDATE(),
    CONSTRAINT PK_mutualistas_captaciones_bruto PRIMARY KEY NONCLUSTERED (id)
)"""
_IDX_BRUTO = ("CREATE CLUSTERED INDEX CIX_mutualistas_captaciones_bruto "
              "ON mutualistas_captaciones_bruto (anio, fecha_corte)")


# ---------------------------------------------------------------------------
# Punto de entrada
# ---------------------------------------------------------------------------

def load(files: dict) -> None:
    """
    files = {"reportes": list[Path], "bases": list[Path]}
    """
    engine = get_master_engine()
    _ensure_tables(engine)

    existing_mut   = _get_existing_hashes(engine, _TABLE_MUT)
    existing_sec   = _get_existing_hashes(engine, _TABLE_SEC)
    existing_bruto = _get_existing_hashes(engine, _TABLE_BRUTO)

    total_mut = total_sec = total_bruto = 0

    for zip_path in files.get("reportes", []):
        m, s = _load_reportes_zip(zip_path, engine, existing_mut, existing_sec)
        total_mut += m
        total_sec += s

    min_year = files.get("min_year", 2017)
    for zip_path in files.get("bases", []):
        total_bruto += _load_bases_zip(zip_path, engine, existing_bruto,
                                       min_year=min_year)

    print(f"[mutualistas] Resumen: "
          f"mutualistas_captaciones={total_mut:,}  "
          f"mutualistas_captaciones_sectores={total_sec:,}  "
          f"mutualistas_captaciones_bruto={total_bruto:,} filas nuevas.")


# ---------------------------------------------------------------------------
# DDL helpers
# ---------------------------------------------------------------------------

def _ensure_tables(engine) -> None:
    from sqlalchemy import inspect as sa_inspect
    insp = sa_inspect(engine)
    for table, ddl, idx in [
        (_TABLE_MUT,   _DDL_MUT,   _IDX_MUT),
        (_TABLE_SEC,   _DDL_SEC,   _IDX_SEC),
        (_TABLE_BRUTO, _DDL_BRUTO, _IDX_BRUTO),
    ]:
        if not insp.has_table(table):
            with engine.begin() as conn:
                conn.execute(text(ddl))
                conn.execute(text(idx))
            print(f"[mutualistas] Tabla {table} creada.")


def _get_existing_hashes(engine, table: str) -> set[str]:
    with engine.connect() as conn:
        rows = conn.execute(text(f"SELECT hash_registro FROM {table}")).fetchall()
    return {r[0] for r in rows}


# ---------------------------------------------------------------------------
# Extracción a disco
# ---------------------------------------------------------------------------

def _ensure_extracted(zip_path: Path) -> Path:
    """
    Extrae zip_path a <parent>/extracted/<stem>/.
    Si el directorio ya existe con contenido, no re-extrae.
    Si el ZIP fue re-descargado (más reciente que el dir), re-extrae.
    """
    out = zip_path.parent / "extracted" / zip_path.stem
    if out.exists() and any(out.rglob("*")):
        # Re-extraer solo si el ZIP es más nuevo que la extracción
        if zip_path.stat().st_mtime <= out.stat().st_mtime:
            return out
        shutil.rmtree(out)
    out.mkdir(parents=True, exist_ok=True)
    print(f"[mutualistas] Extrayendo {zip_path.name}...")
    try:
        with zipfile.ZipFile(zip_path) as zf:
            zf.extractall(out)
        count = sum(1 for p in out.rglob("*") if p.is_file())
        print(f"[mutualistas] {zip_path.name}: {count} archivos extraídos.")
    except Exception as ex:
        print(f"[mutualistas] Error extrayendo {zip_path.name}: {ex}")
    return out


def _extract_nested_zips(outer_dir: Path, min_year: int) -> None:
    """Extrae los ZIPs anuales dentro de un directorio (ej: anterior/)."""
    for inner_zip in sorted(outer_dir.glob("**/*.zip")):
        m = re.match(r"^(\d{4})\.zip$", inner_zip.name)
        if not m:
            continue
        year = int(m.group(1))
        if year < min_year:
            continue
        inner_out = inner_zip.parent / inner_zip.stem
        if inner_out.exists() and any(inner_out.rglob("*")):
            continue
        inner_out.mkdir(parents=True, exist_ok=True)
        try:
            with zipfile.ZipFile(inner_zip) as zf:
                zf.extractall(inner_out)
            print(f"[mutualistas] {inner_zip.name}: extraído en {inner_out.name}/")
        except Exception as ex:
            print(f"[mutualistas] Error extrayendo {inner_zip.name}: {ex}")


# ---------------------------------------------------------------------------
# Carga de reportes (ZIP con xlsm)
# ---------------------------------------------------------------------------

def _load_reportes_zip(zip_path: Path, engine,
                       existing_mut: set, existing_sec: set,
                       min_year: int = 2017) -> tuple[int, int]:
    print(f"[mutualistas] Reportes: {zip_path.name}")
    extract_dir = _ensure_extracted(zip_path)
    new_mut = new_sec = 0

    _XLSM = {".xlsm", ".xlsx", ".xltm"}
    _CSV  = {".txt", ".csv"}

    for fpath in sorted(extract_dir.rglob("*")):
        if not fpath.is_file():
            continue
        ext = fpath.suffix.lower()
        if ext not in _XLSM and ext not in _CSV:
            continue

        # Filtrar por año de directorio padre si aplica
        dir_year = _year_from_dirpath(fpath, extract_dir)
        if dir_year is not None and dir_year < min_year:
            continue

        basename = fpath.name
        if not _is_captaciones_file(basename):
            continue

        sector = _detect_sector(basename)
        is_mut = _is_mutualista(basename)
        if not is_mut and sector is None:
            continue

        tbl      = _TABLE_MUT if is_mut else _TABLE_SEC
        existing = existing_mut if is_mut else existing_sec

        if ext in _XLSM:
            try:
                wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
                _pref = [s for s in wb.sheetnames if s == "Base_captaciones"]
                _pref = _pref or [s for s in wb.sheetnames
                                  if "base" in s.lower() or "cap" in s.lower()]
                _pref = _pref or (wb.sheetnames[:1])
                if not _pref:
                    wb.close()
                    continue
                ws   = wb[_pref[0]]
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
            except Exception as ex:
                print(f"[mutualistas] Error leyendo {basename}: {ex}")
                continue

            if not rows or len(rows) < 2:
                continue

            header  = [_norm_col(str(c)) for c in rows[0]]
            records = []
            for row in rows[1:]:
                rec = _parse_reportes_row(header, row, sector, is_mut)
                if rec is None:
                    continue
                h = _hash_rec(rec)
                rec["hash_registro"] = h
                if h not in existing:
                    records.append(rec)
                    existing.add(h)

            if records:
                _insert(records, tbl, engine)
                if is_mut:
                    new_mut += len(records)
                else:
                    new_sec += len(records)
                print(f"[mutualistas] {basename}: {len(records):,} filas nuevas -> {tbl}")
            else:
                print(f"[mutualistas] {basename}: sin filas nuevas.")

        else:  # TXT / CSV
            try:
                enc = _detect_enc(fpath.read_bytes()[:4096])
                sep = _detect_sep_bytes(fpath.read_bytes()[:4096], enc)
                inserted = 0
                reader = pd.read_csv(
                    fpath, sep=sep, encoding=enc, dtype=str,
                    chunksize=_CHUNKSIZE, on_bad_lines="skip", low_memory=False,
                )
                for chunk in reader:
                    header  = [_norm_col(c) for c in chunk.columns]
                    records = []
                    for _, df_row in chunk.iterrows():
                        rec = _parse_reportes_row(
                            header, tuple(df_row.values), sector, is_mut)
                        if rec is None:
                            continue
                        h = _hash_rec(rec)
                        rec["hash_registro"] = h
                        if h not in existing:
                            records.append(rec)
                            existing.add(h)
                    if records:
                        _insert(records, tbl, engine)
                        inserted += len(records)
                        if is_mut:
                            new_mut += len(records)
                        else:
                            new_sec += len(records)
            except Exception as ex:
                print(f"[mutualistas] Error procesando {basename}: {ex}")
                continue

            if inserted:
                print(f"[mutualistas] {basename}: {inserted:,} filas nuevas -> {tbl}")
            else:
                print(f"[mutualistas] {basename}: sin filas nuevas.")

    return new_mut, new_sec


# ---------------------------------------------------------------------------
# Carga de bases (ZIP con TXT/CSV — incluyendo ZIPs anidados en "anterior")
# ---------------------------------------------------------------------------

def _load_bases_zip(zip_path: Path, engine, existing: set,
                    min_year: int = 2017) -> int:
    print(f"[mutualistas] Bases: {zip_path.name}")
    extract_dir = _ensure_extracted(zip_path)

    # Extraer ZIPs anuales anidados (caso anterior.zip)
    _extract_nested_zips(extract_dir, min_year)

    total_new  = 0
    found_data = False

    for fpath in sorted(extract_dir.rglob("*")):
        if not fpath.is_file():
            continue
        ext = fpath.suffix.lower()

        # Saltar ZIPs (ya se extrajeron en _extract_nested_zips)
        if ext == ".zip":
            continue

        if ext not in {".xlsm", ".xlsx", ".txt", ".csv"}:
            continue

        # Filtrar por año de directorio si aplica
        dir_year = _year_from_dirpath(fpath, extract_dir)
        if dir_year is not None and dir_year < min_year:
            continue

        found_data = True
        if ext in {".xlsm", ".xlsx"}:
            total_new += _process_bases_xlsm(fpath, engine, existing)
        else:
            total_new += _process_bases_stream(fpath, engine, existing)

    if not found_data:
        print(f"[mutualistas] {zip_path.name}: sin archivos de datos reconocidos.")

    return total_new


def _process_bases_xlsm(fpath: Path, engine, existing: set) -> int:
    """Lee un XLSM/XLSX de bases y carga filas en mutualistas_captaciones_bruto."""
    try:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    except Exception as ex:
        print(f"[mutualistas] Error abriendo XLSM {fpath.name}: {ex}")
        return 0

    candidates = [s for s in wb.sheetnames
                  if "base" in s.lower() or "cap" in s.lower()]
    sheet_name = candidates[0] if candidates else (wb.sheetnames[0] if wb.sheetnames else None)
    if sheet_name is None:
        wb.close()
        return 0

    ws   = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows or len(rows) < 2:
        return 0

    header   = [_norm_col(c) for c in rows[0]]
    batch    = []
    inserted = 0
    for row in rows[1:]:
        series = pd.Series(dict(zip(header, row)))
        rec = _parse_bruto_row(series)
        if rec is None:
            continue
        h = _hash_rec(rec)
        rec["hash_registro"] = h
        if h not in existing:
            batch.append(rec)
            existing.add(h)
        if len(batch) >= _CHUNKSIZE:
            _insert(batch, _TABLE_BRUTO, engine)
            inserted += len(batch)
            batch = []
    if batch:
        _insert(batch, _TABLE_BRUTO, engine)
        inserted += len(batch)

    if inserted:
        print(f"[mutualistas] {fpath.name}: {inserted:,} filas nuevas -> {_TABLE_BRUTO}")
    else:
        print(f"[mutualistas] {fpath.name}: sin filas nuevas.")
    return inserted


def _process_bases_stream(fpath: Path, engine, existing: set) -> int:
    """Lee un TXT/CSV desde disco (seekable) en modo streaming por chunks."""
    try:
        sample = fpath.read_bytes()[:4096]
        enc = _detect_enc(sample)
        sep = _detect_sep_bytes(sample, enc)
        inserted = 0
        reader = pd.read_csv(
            fpath, sep=sep, encoding=enc, dtype=str,
            chunksize=_CHUNKSIZE, on_bad_lines="skip", low_memory=False,
        )
        for chunk in reader:
            chunk.columns = [_norm(c) for c in chunk.columns]
            records = []
            for _, row in chunk.iterrows():
                rec = _parse_bruto_row(row)
                if rec is None:
                    continue
                h = _hash_rec(rec)
                rec["hash_registro"] = h
                if h not in existing:
                    records.append(rec)
                    existing.add(h)
            if records:
                _insert(records, _TABLE_BRUTO, engine)
                inserted += len(records)
    except Exception as ex:
        print(f"[mutualistas] Error procesando {fpath.name}: {ex}")
        return 0

    if inserted:
        print(f"[mutualistas] {fpath.name}: {inserted:,} filas nuevas -> {_TABLE_BRUTO}")
    else:
        print(f"[mutualistas] {fpath.name}: sin filas nuevas.")
    return inserted


# ---------------------------------------------------------------------------
# Helpers de detección
# ---------------------------------------------------------------------------

def _detect_sector(fname: str) -> int | None:
    m = re.search(r"[_\-]S(\d)[_\-\.]", fname, re.IGNORECASE)
    if m:
        return int(m.group(1))
    m = re.search(r"S([123])", fname, re.IGNORECASE)
    return int(m.group(1)) if m else None


def _is_mutualista(fname: str) -> bool:
    return bool(re.search(r"Mut", fname, re.IGNORECASE))


def _is_captaciones_file(basename: str) -> bool:
    name = basename.lower()
    return "captacion" in name or bool(
        re.search(r"[_\-](mut|s[123])[_\-\.]", name, re.IGNORECASE))


def _year_from_dirpath(fpath: Path, root: Path) -> int | None:
    """Extrae año del primer componente del path relativo al directorio raíz."""
    try:
        rel = fpath.relative_to(root)
        first = rel.parts[0] if len(rel.parts) > 1 else None
        if first:
            m = re.match(r"^(\d{4})", first)
            if m:
                return int(m.group(1))
    except ValueError:
        pass
    return None


# ---------------------------------------------------------------------------
# Parsers de fila
# ---------------------------------------------------------------------------

def _parse_reportes_row(header: list[str], row: tuple,
                        sector: int | None, is_mut: bool) -> dict | None:
    d = dict(zip(header, row))

    fecha_corte = _to_date(d.get("FECHA DE CORTE") or d.get("FECHADECORTE"))
    if fecha_corte is None:
        return None

    rec = {
        "anio":             int(fecha_corte[:4]),
        "saldo":            _to_float(d.get("SALDO")),
        "num_clientes":     _to_int(d.get("NUMERO DE CLIENTES") or d.get("NUMERODECLIENTES")),
        "num_cuentas":      _to_int(d.get("NUMERO DE CUENTAS") or d.get("NUMERODECUENTAS")),
        "fecha_corte":      fecha_corte,
        "region":           _clean(d.get("REGION")),
        "provincia":        _clean(d.get("PROVINCIA")),
        "canton":           _clean(d.get("CANTON")),
        "tipo_deposito":    _clean(d.get("TIPO DE DEPOSITO") or d.get("TIPODEDEPOSITO")),
        "estado_operacion": _clean(d.get("ESTADO OPERACION")
                                   or d.get("ESTADO OPERACIÓN")
                                   or d.get("ESTADOOPERACION")),
        "ruc":              _clean(d.get("RUC")),
        "razon_social":     _clean(d.get("RAZON SOCIAL") or d.get("RAZONSOCIAL")),
    }
    if not is_mut:
        rec["sector"] = sector
    return rec


def _parse_bruto_row(row: pd.Series) -> dict | None:
    def g(*keys):
        for k in keys:
            v = row.get(k)
            if v is not None and str(v).strip().lower() not in ("nan", "none", ""):
                return v
        return None

    fecha_corte = _to_date(_strip_quotes(
        g("FECHA_CORTE", "FECHA DE CORTE", "FECHADECORTE")))
    if fecha_corte is None:
        return None
    return {
        "anio":              int(fecha_corte[:4]),
        "tipo_persona":      _strip_quotes(g("TIPO_PERSONA", "TIPO DE PERSONA", "TIPO PERSONA")),
        "fecha_corte":       fecha_corte,
        "segmento":          _strip_quotes(g("SEGMENTO")),
        "provincia":         _strip_quotes(g("PROVINCIA")),
        "canton":            _strip_quotes(g("CANTON")),
        "parroquia":         _strip_quotes(g("PARROQUIA")),
        "dpaparroquia":      _strip_quotes(g("DPAPARROQUIA", "DPA PARROQUIA",
                                             "COD DPA", "COD_DPA")),
        "estado_operacion":  _strip_quotes(g("ESTADO_OPERACION", "ESTADO OPERACION",
                                             "ESTADO DE OPERACION")),
        "tipo_cuenta":       _strip_quotes(g("TIPO_CUENTA", "TIPO DE CUENTA", "TIPO CUENTA")),
        "banda_maduracion":  _strip_quotes(g("BANDAMADURACION", "BANDA MADURACION",
                                             "BANDA DE MADURACION")),
        "sexo":              _strip_quotes(g("SEXO")),
        "rango_edad":        _strip_quotes(g("RANGOEDAD", "RANGO EDAD", "RANGO DE EDAD")),
        "nivel_instruccion": _strip_quotes(g("NIVELINSTRUCCION", "NIVEL INSTRUCCION",
                                             "NIVEL DE INSTRUCCION")),
        "rango_saldo":       _strip_quotes(g("RANGOSALDO", "RANGO SALDO", "RANGO DE SALDO")),
        "nro_cuentas":       _to_int(g("NROCUENTAS", "NRO CUENTAS",
                                       "NUMERO CUENTAS", "NUMERO DE CUENTAS")),
        "saldo_usd":         _to_float(g("SALDO (USD)", "SALDO_USD", "SALDO USD")),
        "nro_depositantes":  _to_int(g("NRODEPOSITANTES", "NRO DEPOSITANTES",
                                       "NUMERO DEPOSITANTES")),
    }


# ---------------------------------------------------------------------------
# Insercion
# ---------------------------------------------------------------------------

def _insert(records: list[dict], table: str, engine) -> None:
    if not records:
        return
    cols         = list(records[0].keys())
    col_list     = ", ".join(cols)
    placeholders = ", ".join(f":{c}" for c in cols)
    with engine.begin() as conn:
        conn.execute(
            text(f"INSERT INTO {table} ({col_list}) VALUES ({placeholders})"),
            records,
        )


# ---------------------------------------------------------------------------
# Helpers de texto / tipo
# ---------------------------------------------------------------------------

def _norm(c: str) -> str:
    c = re.sub(r"^[^\x20-\x7E]+", "", str(c).strip())
    return c.strip().upper()


def _norm_col(c: str) -> str:
    s = str(c).strip().replace("▓", "")
    s = re.sub(r"^[^\x20-\x7E]+", "", s)
    nfkd = unicodedata.normalize("NFKD", s)
    ascii_s = "".join(ch for ch in nfkd if unicodedata.category(ch) != "Mn")
    ascii_s = re.sub(r"[^A-Za-z0-9 _]", " ", ascii_s)
    return re.sub(r"\s+", " ", ascii_s).strip().upper()


def _clean(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s.lower() not in ("nan", "none") else None


def _strip_quotes(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip().strip('"').strip()
    return s if s and s.lower() not in ("nan", "none") else None


def _to_float(v) -> float | None:
    if v is None:
        return None
    try:
        f = float(str(v).strip().replace(",", ""))
        return None if (np.isnan(f) or np.isinf(f)) else f
    except (ValueError, TypeError):
        return None


def _to_int(v) -> int | None:
    if v is None:
        return None
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        return None


def _to_date(v) -> str | None:
    if v is None:
        return None
    from datetime import datetime, date
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip().rstrip(",").strip()
    if not s or s.lower() in ("nan", "none", ""):
        return None
    m = re.match(r"(\d{4}-\d{2}-\d{2})", s)
    if m:
        return m.group(1)
    return None


def _hash_rec(rec: dict) -> str:
    key = "|".join(str(v) if v is not None else "" for v in rec.values())
    return hashlib.sha256(key.encode()).hexdigest()


def _detect_enc(raw: bytes) -> str:
    try:
        raw.decode("utf-8")
        return "utf-8-sig"
    except UnicodeDecodeError:
        return "latin-1"


def _detect_sep_bytes(raw: bytes, enc: str) -> str:
    first_line = raw.decode(enc, errors="replace").split("\n")[0]
    counts = {"\t": first_line.count("\t"),
              "|":  first_line.count("|"),
              ";":  first_line.count(";")}
    return max(counts, key=counts.get)
