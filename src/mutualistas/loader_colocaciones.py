"""
ETL loader: Colocaciones Mutualistas — SEPS Ecuador

Tablas (8):
  mutualistas_colocaciones_volumen_credito  — Base_vcredito, archivos Mut
  mutualistas_colocaciones_volumen_credito_sectores     — Base_vcredito, archivos S1/S2/S3 + sector
  mutualistas_colocaciones                  — Base_colocaciones, archivos MUT
  mutualistas_colocaciones_sectores                     — Base_colocaciones, archivos S1/S2/S3 + sector
  mutualistas_colocaciones_volumen_credito_bruto        — TXT tab-sep (volumen_bruto ZIPs)
  mutualistas_colocaciones_mensual_bruto                — TXT Deflate64 ~2.8 GB (col_bruto ZIPs)
  mutualistas_colocaciones_tarjetas_con_forma_pago      — TXT "Con forma de pago" (tarjetas ZIPs)
  mutualistas_colocaciones_tarjetas_sin_forma_pago      — TXT "Sin forma de pago" (tarjetas ZIPs)

Deflate64: col_bruto usa compress_type=9. Requiere el paquete 'inflate64'.
  pip install inflate64
"""

import hashlib
import io
import re
import sys
import unicodedata
import zipfile
from datetime import datetime as _dt, date as _date, timedelta as _td
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from sqlalchemy import text

sys.path.append(str(Path(__file__).resolve().parents[2]))
from utils.base_engine import get_master_engine

_CHUNKSIZE = 50_000

# ---------------------------------------------------------------------------
# Soporte Deflate64 (compress_type=9) via inflate64
# ---------------------------------------------------------------------------

_DEFLATE64_OK = False
try:
    import inflate64 as _inf64

    _orig_get_decomp = zipfile._get_decompressor

    class _D64Decompressor:
        def __init__(self):
            self._inflater = _inf64.Inflater()

        def decompress(self, data, max_length=-1):
            if self._inflater.eof:
                return b""
            return self._inflater.inflate(data)

        @property
        def eof(self):
            return self._inflater.eof

        @property
        def unused_data(self):
            return b""

    def _patched_get_decomp(compress_type):
        return (_D64Decompressor() if compress_type == 9
                else _orig_get_decomp(compress_type))

    zipfile._get_decompressor = _patched_get_decomp
    _DEFLATE64_OK = True
except ImportError:
    pass

# ---------------------------------------------------------------------------
# DDL — mutualistas_colocaciones_volumen_credito
# ---------------------------------------------------------------------------

_TABLE_VOL_MUT = "mutualistas_colocaciones_volumen_credito"

_DDL_VOL_MUT = """
CREATE TABLE mutualistas_colocaciones_volumen_credito (
    id               BIGINT IDENTITY(1,1) NOT NULL,
    anio             SMALLINT       NOT NULL,
    fecha_corte      DATE           NOT NULL,
    num_ruc          NVARCHAR(20)   NULL,
    razon_social     NVARCHAR(300)  NULL,
    segmento         NVARCHAR(50)   NULL,
    region           NVARCHAR(50)   NULL,
    provincia        NVARCHAR(100)  NULL,
    canton           NVARCHAR(200)  NULL,
    tipo_credito_gral NVARCHAR(200) NULL,
    estado_operacion NVARCHAR(100)  NULL,
    actividad_economica NVARCHAR(300) NULL,
    destino_financiero  NVARCHAR(200) NULL,
    tipo_operacion   NVARCHAR(100)  NULL,
    val_operacion    FLOAT          NULL,
    sujetos_credito  INT            NULL,
    operaciones      INT            NULL,
    hash_registro    NVARCHAR(64)   NOT NULL,
    fecha_carga      DATETIME2      DEFAULT GETDATE(),
    CONSTRAINT PK_mutualistas_colocaciones_volumen_credito
        PRIMARY KEY NONCLUSTERED (id)
)"""
_IDX_VOL_MUT = ("CREATE CLUSTERED INDEX CIX_mutualistas_colocaciones_volumen_credito "
                "ON mutualistas_colocaciones_volumen_credito (anio, fecha_corte)")

# ---------------------------------------------------------------------------
# DDL — mutualistas_colocaciones_volumen_credito_sectores
# ---------------------------------------------------------------------------

_TABLE_VOL_SEC = "mutualistas_colocaciones_volumen_credito_sectores"

_DDL_VOL_SEC = """
CREATE TABLE mutualistas_colocaciones_volumen_credito_sectores (
    id               BIGINT IDENTITY(1,1) NOT NULL,
    anio             SMALLINT       NOT NULL,
    sector           TINYINT        NOT NULL,
    fecha_corte      DATE           NOT NULL,
    num_ruc          NVARCHAR(20)   NULL,
    razon_social     NVARCHAR(300)  NULL,
    segmento         NVARCHAR(50)   NULL,
    region           NVARCHAR(50)   NULL,
    provincia        NVARCHAR(100)  NULL,
    canton           NVARCHAR(200)  NULL,
    tipo_credito_gral NVARCHAR(200) NULL,
    estado_operacion NVARCHAR(100)  NULL,
    actividad_economica NVARCHAR(300) NULL,
    destino_financiero  NVARCHAR(200) NULL,
    tipo_operacion   NVARCHAR(100)  NULL,
    val_operacion    FLOAT          NULL,
    sujetos_credito  INT            NULL,
    operaciones      INT            NULL,
    hash_registro    NVARCHAR(64)   NOT NULL,
    fecha_carga      DATETIME2      DEFAULT GETDATE(),
    CONSTRAINT PK_mutualistas_colocaciones_volumen_credito_sectores
        PRIMARY KEY NONCLUSTERED (id)
)"""
_IDX_VOL_SEC = ("CREATE CLUSTERED INDEX CIX_mutualistas_colocaciones_volumen_credito_sectores "
                "ON mutualistas_colocaciones_volumen_credito_sectores (anio, sector, fecha_corte)")

# ---------------------------------------------------------------------------
# DDL — mutualistas_colocaciones
# ---------------------------------------------------------------------------

_TABLE_COL_MUT = "mutualistas_colocaciones"

_DDL_COL_MUT = """
CREATE TABLE mutualistas_colocaciones (
    id                  BIGINT IDENTITY(1,1) NOT NULL,
    anio                SMALLINT       NOT NULL,
    cartera_por_vencer  FLOAT          NULL,
    cartera_no_devenga  FLOAT          NULL,
    cartera_vencida     FLOAT          NULL,
    cartera_total       FLOAT          NULL,
    num_operaciones     INT            NULL,
    num_sujetos_credito INT            NULL,
    fecha_corte         DATE           NOT NULL,
    region              NVARCHAR(50)   NULL,
    provincia           NVARCHAR(100)  NULL,
    canton              NVARCHAR(200)  NULL,
    subtipo_credito     NVARCHAR(200)  NULL,
    origen_operacion    NVARCHAR(100)  NULL,
    linea_credito       NVARCHAR(200)  NULL,
    estado_operacion    NVARCHAR(100)  NULL,
    clase_credito       NVARCHAR(100)  NULL,
    actividad_economica NVARCHAR(300)  NULL,
    ruc                 NVARCHAR(20)   NULL,
    razon_social        NVARCHAR(300)  NULL,
    hash_registro       NVARCHAR(64)   NOT NULL,
    fecha_carga         DATETIME2      DEFAULT GETDATE(),
    CONSTRAINT PK_mutualistas_colocaciones PRIMARY KEY NONCLUSTERED (id)
)"""
_IDX_COL_MUT = ("CREATE CLUSTERED INDEX CIX_mutualistas_colocaciones "
                "ON mutualistas_colocaciones (anio, fecha_corte)")

# ---------------------------------------------------------------------------
# DDL — mutualistas_colocaciones_sectores
# ---------------------------------------------------------------------------

_TABLE_COL_SEC = "mutualistas_colocaciones_sectores"

_DDL_COL_SEC = """
CREATE TABLE mutualistas_colocaciones_sectores (
    id                  BIGINT IDENTITY(1,1) NOT NULL,
    anio                SMALLINT       NOT NULL,
    sector              TINYINT        NOT NULL,
    cartera_por_vencer  FLOAT          NULL,
    cartera_no_devenga  FLOAT          NULL,
    cartera_vencida     FLOAT          NULL,
    cartera_total       FLOAT          NULL,
    num_operaciones     INT            NULL,
    num_sujetos_credito INT            NULL,
    fecha_corte         DATE           NOT NULL,
    region              NVARCHAR(50)   NULL,
    provincia           NVARCHAR(100)  NULL,
    canton              NVARCHAR(200)  NULL,
    subtipo_credito     NVARCHAR(200)  NULL,
    origen_operacion    NVARCHAR(100)  NULL,
    linea_credito       NVARCHAR(200)  NULL,
    estado_operacion    NVARCHAR(100)  NULL,
    clase_credito       NVARCHAR(100)  NULL,
    actividad_economica NVARCHAR(300)  NULL,
    ruc                 NVARCHAR(20)   NULL,
    razon_social        NVARCHAR(300)  NULL,
    hash_registro       NVARCHAR(64)   NOT NULL,
    fecha_carga         DATETIME2      DEFAULT GETDATE(),
    CONSTRAINT PK_mutualistas_colocaciones_sectores PRIMARY KEY NONCLUSTERED (id)
)"""
_IDX_COL_SEC = ("CREATE CLUSTERED INDEX CIX_mutualistas_colocaciones_sectores "
                "ON mutualistas_colocaciones_sectores (anio, sector, fecha_corte)")

# ---------------------------------------------------------------------------
# DDL — mutualistas_colocaciones_volumen_credito_bruto
# ---------------------------------------------------------------------------

_TABLE_VOL_BRUTO = "mutualistas_colocaciones_volumen_credito_bruto"

_DDL_VOL_BRUTO = """
CREATE TABLE mutualistas_colocaciones_volumen_credito_bruto (
    id                   BIGINT IDENTITY(1,1) NOT NULL,
    anio                 SMALLINT       NOT NULL,
    tipo_operacion       NVARCHAR(200)  NULL,
    estado_operacion     NVARCHAR(100)  NULL,
    segmento             NVARCHAR(50)   NULL,
    cod_actividad_eco    NVARCHAR(20)   NULL,
    acteco_seccion       NVARCHAR(500)  NULL,
    acteco_division      NVARCHAR(500)  NULL,
    acteco_grupo         NVARCHAR(500)  NULL,
    acteco_clase         NVARCHAR(500)  NULL,
    acteco_subclase      NVARCHAR(500)  NULL,
    actividad_economica  NVARCHAR(2000) NULL,
    canton               NVARCHAR(200)  NULL,
    cod_dpa              NVARCHAR(20)   NULL,
    parroquia            NVARCHAR(200)  NULL,
    provincia            NVARCHAR(100)  NULL,
    fecha_corte          DATE           NOT NULL,
    tipo_credito_general NVARCHAR(200)  NULL,
    instruccion          NVARCHAR(100)  NULL,
    sexo                 NVARCHAR(20)   NULL,
    destino_financiero   NVARCHAR(200)  NULL,
    rango_edad           NVARCHAR(50)   NULL,
    rango_monto_credito  NVARCHAR(100)  NULL,
    rango_plazo_original NVARCHAR(100)  NULL,
    tipo_persona         NVARCHAR(50)   NULL,
    tipo_credito_espec   NVARCHAR(200)  NULL,
    nro_sujetos_credito  INT            NULL,
    nro_operaciones      INT            NULL,
    monto_concedido_usd  FLOAT          NULL,
    hash_registro        NVARCHAR(64)   NOT NULL,
    fecha_carga          DATETIME2      DEFAULT GETDATE(),
    CONSTRAINT PK_mutualistas_colocaciones_volumen_credito_bruto
        PRIMARY KEY NONCLUSTERED (id)
)"""
_IDX_VOL_BRUTO = ("CREATE CLUSTERED INDEX CIX_mutualistas_colocaciones_volumen_credito_bruto "
                  "ON mutualistas_colocaciones_volumen_credito_bruto (anio, fecha_corte)")

# ---------------------------------------------------------------------------
# DDL — mutualistas_colocaciones_mensual_bruto
# ---------------------------------------------------------------------------

_TABLE_COL_BRUTO = "mutualistas_colocaciones_mensual_bruto"

_DDL_COL_BRUTO = """
CREATE TABLE mutualistas_colocaciones_mensual_bruto (
    id                   BIGINT IDENTITY(1,1) NOT NULL,
    anio                 SMALLINT       NOT NULL,
    fecha_corte          DATE           NOT NULL,
    segmento             NVARCHAR(50)   NULL,
    tipo_operacion       NVARCHAR(200)  NULL,
    tipo_credito_general NVARCHAR(200)  NULL,
    tipo_credito_espec   NVARCHAR(200)  NULL,
    cod_destino_fin      NVARCHAR(20)   NULL,
    destino_financiero   NVARCHAR(200)  NULL,
    cod_actividad_eco    NVARCHAR(20)   NULL,
    acteco_seccion       NVARCHAR(500)  NULL,
    acteco_division      NVARCHAR(500)  NULL,
    acteco_grupo         NVARCHAR(500)  NULL,
    acteco_clase         NVARCHAR(500)  NULL,
    acteco_subclase      NVARCHAR(500)  NULL,
    actividad_economica  NVARCHAR(2000) NULL,
    provincia            NVARCHAR(100)  NULL,
    canton               NVARCHAR(200)  NULL,
    parroquia            NVARCHAR(200)  NULL,
    cod_dpa              NVARCHAR(20)   NULL,
    tipo_persona         NVARCHAR(50)   NULL,
    sexo                 NVARCHAR(20)   NULL,
    rango_edad           NVARCHAR(50)   NULL,
    nivel_instruccion    NVARCHAR(100)  NULL,
    rango_saldo          NVARCHAR(100)  NULL,
    nro_sujetos          INT            NULL,
    nro_operaciones      INT            NULL,
    valor_por_vencer_usd FLOAT          NULL,
    valor_no_devenga_usd FLOAT          NULL,
    valor_vencido_usd    FLOAT          NULL,
    valor_saldo_total_usd FLOAT         NULL,
    hash_registro        NVARCHAR(64)   NOT NULL,
    fecha_carga          DATETIME2      DEFAULT GETDATE(),
    CONSTRAINT PK_mutualistas_colocaciones_mensual_bruto PRIMARY KEY NONCLUSTERED (id)
)"""
_IDX_COL_BRUTO = ("CREATE CLUSTERED INDEX CIX_mutualistas_colocaciones_mensual_bruto "
                  "ON mutualistas_colocaciones_mensual_bruto (anio, fecha_corte)")

# ---------------------------------------------------------------------------
# DDL — mutualistas_colocaciones_tarjetas_con_forma_pago
# ---------------------------------------------------------------------------

_TABLE_TAR_CON = "mutualistas_colocaciones_tarjetas_con_forma_pago"

_DDL_TAR_CON = """
CREATE TABLE mutualistas_colocaciones_tarjetas_con_forma_pago (
    id                BIGINT IDENTITY(1,1) NOT NULL,
    anio              SMALLINT      NOT NULL,
    fecha             DATE          NOT NULL,
    forma_pago        NVARCHAR(100) NULL,
    sexo              NVARCHAR(20)  NULL,
    instruccion       NVARCHAR(100) NULL,
    rango_edad        NVARCHAR(50)  NULL,
    capital_por_vencer FLOAT        NULL,
    capital_vencido   FLOAT         NULL,
    capital_no_devenga FLOAT        NULL,
    capital_consumo   FLOAT         NULL,
    hash_registro     NVARCHAR(64)  NOT NULL,
    fecha_carga       DATETIME2     DEFAULT GETDATE(),
    CONSTRAINT PK_mutualistas_colocaciones_tarjetas_con_forma_pago
        PRIMARY KEY NONCLUSTERED (id)
)"""
_IDX_TAR_CON = ("CREATE CLUSTERED INDEX CIX_mutualistas_colocaciones_tarjetas_con_forma_pago "
                "ON mutualistas_colocaciones_tarjetas_con_forma_pago (anio, fecha)")

# ---------------------------------------------------------------------------
# DDL — mutualistas_colocaciones_tarjetas_sin_forma_pago
# ---------------------------------------------------------------------------

_TABLE_TAR_SIN = "mutualistas_colocaciones_tarjetas_sin_forma_pago"

_DDL_TAR_SIN = """
CREATE TABLE mutualistas_colocaciones_tarjetas_sin_forma_pago (
    id                   BIGINT IDENTITY(1,1) NOT NULL,
    anio                 SMALLINT      NOT NULL,
    fecha                DATE          NOT NULL,
    sexo                 NVARCHAR(20)  NULL,
    instruccion          NVARCHAR(100) NULL,
    rango_edad           NVARCHAR(50)  NULL,
    num_tarjetahabientes INT           NULL,
    num_tarjetas         INT           NULL,
    capital_por_vencer   FLOAT         NULL,
    capital_vencido      FLOAT         NULL,
    capital_no_devenga   FLOAT         NULL,
    capital_consumo      FLOAT         NULL,
    hash_registro        NVARCHAR(64)  NOT NULL,
    fecha_carga          DATETIME2     DEFAULT GETDATE(),
    CONSTRAINT PK_mutualistas_colocaciones_tarjetas_sin_forma_pago
        PRIMARY KEY NONCLUSTERED (id)
)"""
_IDX_TAR_SIN = ("CREATE CLUSTERED INDEX CIX_mutualistas_colocaciones_tarjetas_sin_forma_pago "
                "ON mutualistas_colocaciones_tarjetas_sin_forma_pago (anio, fecha)")


# ---------------------------------------------------------------------------
# Punto de entrada
# ---------------------------------------------------------------------------

def load(files: dict) -> None:
    """
    files = {
      "volumen":       list[Path],
      "colocaciones":  list[Path],
      "volumen_bruto": list[Path],
      "col_bruto":     list[Path],
      "tarjetas":      list[Path],
      "min_year":      int,          (opcional)
    }
    """
    engine   = get_master_engine()
    _ensure_tables(engine)
    min_year = files.get("min_year", 2017)

    ex_vol_mut  = _get_hashes(engine, _TABLE_VOL_MUT)
    ex_vol_sec  = _get_hashes(engine, _TABLE_VOL_SEC)
    ex_col_mut  = _get_hashes(engine, _TABLE_COL_MUT)
    ex_col_sec  = _get_hashes(engine, _TABLE_COL_SEC)
    ex_vol_bru  = _get_hashes(engine, _TABLE_VOL_BRUTO)
    ex_col_bru  = _get_hashes(engine, _TABLE_COL_BRUTO)
    ex_tar_con  = _get_hashes(engine, _TABLE_TAR_CON)
    ex_tar_sin  = _get_hashes(engine, _TABLE_TAR_SIN)

    tot = {t: 0 for t in [_TABLE_VOL_MUT, _TABLE_VOL_SEC, _TABLE_COL_MUT,
                           _TABLE_COL_SEC, _TABLE_VOL_BRUTO, _TABLE_COL_BRUTO,
                           _TABLE_TAR_CON, _TABLE_TAR_SIN]}

    for p in files.get("volumen", []):
        m, s = _load_xlsm_zip(p, "Base_vcredito", _parse_volumen_row,
                               _TABLE_VOL_MUT, _TABLE_VOL_SEC,
                               ex_vol_mut, ex_vol_sec, engine, min_year)
        tot[_TABLE_VOL_MUT] += m; tot[_TABLE_VOL_SEC] += s

    for p in files.get("colocaciones", []):
        m, s = _load_xlsm_zip(p, "Base_colocaciones", _parse_col_row,
                               _TABLE_COL_MUT, _TABLE_COL_SEC,
                               ex_col_mut, ex_col_sec, engine, min_year)
        tot[_TABLE_COL_MUT] += m; tot[_TABLE_COL_SEC] += s

    for p in files.get("volumen_bruto", []):
        tot[_TABLE_VOL_BRUTO] += _load_bruto_zip(
            p, _TABLE_VOL_BRUTO, _parse_vol_bruto_row, engine, ex_vol_bru,
            use_deflate64=False, xlsm_sheets=["Base_vcredito", "Base_Vcredito"])

    for p in files.get("col_bruto", []):
        if not _DEFLATE64_OK:
            print("[colocaciones] col_bruto omitido: instalar 'inflate64' "
                  "(pip install inflate64) para soporte Deflate64.")
        else:
            tot[_TABLE_COL_BRUTO] += _load_bruto_zip(
                p, _TABLE_COL_BRUTO, _parse_col_bruto_row, engine, ex_col_bru,
                use_deflate64=True, xlsm_sheets=["Base_colocaciones", "Base_Colocaciones"])

    for p in files.get("tarjetas", []):
        c, s = _load_tarjetas_zip(p, engine, ex_tar_con, ex_tar_sin)
        tot[_TABLE_TAR_CON] += c; tot[_TABLE_TAR_SIN] += s

    print("[colocaciones] Resumen:")
    for t, n in tot.items():
        print(f"  {t:<50} {n:>10,} filas nuevas")


# ---------------------------------------------------------------------------
# DDL helpers
# ---------------------------------------------------------------------------

def _ensure_tables(engine) -> None:
    from sqlalchemy import inspect as sa_inspect
    insp = sa_inspect(engine)
    for table, ddl, idx in [
        (_TABLE_VOL_MUT,  _DDL_VOL_MUT,  _IDX_VOL_MUT),
        (_TABLE_VOL_SEC,  _DDL_VOL_SEC,  _IDX_VOL_SEC),
        (_TABLE_COL_MUT,  _DDL_COL_MUT,  _IDX_COL_MUT),
        (_TABLE_COL_SEC,  _DDL_COL_SEC,  _IDX_COL_SEC),
        (_TABLE_VOL_BRUTO, _DDL_VOL_BRUTO, _IDX_VOL_BRUTO),
        (_TABLE_COL_BRUTO, _DDL_COL_BRUTO, _IDX_COL_BRUTO),
        (_TABLE_TAR_CON,  _DDL_TAR_CON,  _IDX_TAR_CON),
        (_TABLE_TAR_SIN,  _DDL_TAR_SIN,  _IDX_TAR_SIN),
    ]:
        if not insp.has_table(table):
            with engine.begin() as conn:
                conn.execute(text(ddl))
                conn.execute(text(idx))
            print(f"[colocaciones] Tabla {table} creada.")


def _get_hashes(engine, table: str) -> set[str]:
    with engine.connect() as conn:
        rows = conn.execute(
            text(f"SELECT hash_registro FROM {table}")
        ).fetchall()
    return {r[0] for r in rows}


# ---------------------------------------------------------------------------
# Loader XLSM (volumen + colocaciones)
# ---------------------------------------------------------------------------

def _load_xlsm_zip(zip_path: Path, sheet_name: str, parser,
                   table_mut: str, table_sec: str,
                   existing_mut: set, existing_sec: set,
                   engine, min_year: int = 2017) -> tuple[int, int]:
    label = zip_path.name
    print(f"[colocaciones] XLSM {sheet_name}: {label}")
    try:
        zf = zipfile.ZipFile(zip_path)
    except Exception as ex:
        print(f"[colocaciones] Error abriendo {label}: {ex}")
        return 0, 0

    _XLSM = {".xlsm", ".xlsx"}
    _CSV  = {".txt", ".csv"}
    new_mut = new_sec = 0

    for fname in zf.namelist():
        basename = fname.split("/")[-1]
        ext = Path(basename).suffix.lower()

        if ext not in _XLSM and ext not in _CSV:
            continue

        dir_year = _year_from_zippath(fname)
        if dir_year is not None and dir_year < min_year:
            continue

        sector = _detect_sector(basename)
        is_mut = _is_mutualista(basename)
        if not is_mut and sector is None:
            continue

        tbl      = table_mut if is_mut else table_sec
        existing = existing_mut if is_mut else existing_sec

        if ext in _XLSM:
            try:
                wb = openpyxl.load_workbook(
                    io.BytesIO(zf.read(fname)), read_only=True, data_only=True
                )
                # Buscar hoja: exacto → keyword → cualquier "base" → primera
                _kw = sheet_name.split("_")[-1].lower()
                _sh = (sheet_name if sheet_name in wb.sheetnames else None)
                if _sh is None:
                    _cands = [s for s in wb.sheetnames
                              if _kw in s.lower() or "base" in s.lower()]
                    _sh = _cands[0] if _cands else (wb.sheetnames[0] if wb.sheetnames else None)
                if _sh is None:
                    wb.close()
                    continue
                ws   = wb[_sh]
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
            except Exception as ex:
                print(f"[colocaciones] Error leyendo {fname}: {ex}")
                continue

            if not rows or len(rows) < 2:
                continue

            header  = [_norm_col(c) for c in rows[0]]
            records = []
            for row in rows[1:]:
                rec = parser(header, row, sector, is_mut)
                if rec is None:
                    continue
                h = _hash_rec(rec)
                rec["hash_registro"] = h
                if h not in existing:
                    records.append(rec)
                    existing.add(h)

            if records:
                _insert(records, tbl, engine)
                if is_mut:
                    new_mut += len(records)
                else:
                    new_sec += len(records)
                print(f"[colocaciones] {fname}: {len(records):,} -> {tbl}")
            else:
                print(f"[colocaciones] {fname}: sin filas nuevas.")

        else:  # TXT / CSV
            try:
                with zf.open(fname) as f:
                    header_bytes = f.read(4096)
                enc = _detect_enc(header_bytes)
                sep = _detect_sep_bytes(header_bytes, enc)
                inserted = 0
                with zf.open(fname) as f:
                    reader = pd.read_csv(
                        f, sep=sep, encoding=enc, dtype=str,
                        chunksize=_CHUNKSIZE, on_bad_lines="skip",
                        low_memory=False,
                    )
                    for chunk in reader:
                        header = [_norm_col(c) for c in chunk.columns]
                        records = []
                        for _, df_row in chunk.iterrows():
                            rec = parser(header, tuple(df_row.values), sector, is_mut)
                            if rec is None:
                                continue
                            h = _hash_rec(rec)
                            rec["hash_registro"] = h
                            if h not in existing:
                                records.append(rec)
                                existing.add(h)
                        if records:
                            _insert(records, tbl, engine)
                            inserted += len(records)
                            if is_mut:
                                new_mut += len(records)
                            else:
                                new_sec += len(records)
            except Exception as ex:
                print(f"[colocaciones] Error procesando {fname}: {ex}")
                continue

            if inserted:
                print(f"[colocaciones] {fname}: {inserted:,} -> {tbl}")
            else:
                print(f"[colocaciones] {fname}: sin filas nuevas.")

    zf.close()
    return new_mut, new_sec


# ---------------------------------------------------------------------------
# Loader bruto (TXT / Deflate64)
# ---------------------------------------------------------------------------

def _load_bruto_zip(zip_path: Path, table: str, row_parser,
                    engine, existing: set,
                    use_deflate64: bool = False,
                    xlsm_sheets: list[str] | None = None) -> int:
    label = zip_path.name
    print(f"[colocaciones] Bruto {table}: {label}")
    try:
        zf = zipfile.ZipFile(zip_path)
    except Exception as ex:
        print(f"[colocaciones] Error abriendo {label}: {ex}")
        return 0

    _XLSM_EXTS = {".xlsm", ".xlsx"}
    _CSV_EXTS  = {".txt", ".csv", ".dat"}   # .dat = Deflate64 col_bruto 2017-2021
    total_new  = 0

    for fname in zf.namelist():
        ext = Path(fname.split("/")[-1]).suffix.lower()

        if ext in _XLSM_EXTS:
            total_new += _process_bruto_xlsm(
                zf, fname, table, row_parser, engine, existing, xlsm_sheets)
            continue

        if ext == ".rar":
            print(f"[colocaciones] {fname}: formato RAR no soportado — omitido. "
                  f"Descomprimir manualmente y volver a ejecutar con --etl-only.")
            continue

        if ext not in _CSV_EXTS:
            continue

        try:
            with zf.open(fname) as f:
                header_bytes = f.read(4096)
        except Exception as ex:
            print(f"[colocaciones] Error abriendo {fname}: {ex}")
            continue

        enc = _detect_enc(header_bytes)
        sep = _detect_sep_bytes(header_bytes, enc)
        inserted = 0

        try:
            with zf.open(fname) as f:
                reader = pd.read_csv(
                    f, sep=sep, encoding=enc, dtype=str,
                    chunksize=_CHUNKSIZE, on_bad_lines="skip",
                    low_memory=False,
                )
                for chunk in reader:
                    chunk.columns = [_norm_col(c) for c in chunk.columns]
                    records = []
                    for _, row in chunk.iterrows():
                        rec = row_parser(row)
                        if rec is None:
                            continue
                        h = _hash_rec(rec)
                        rec["hash_registro"] = h
                        if h not in existing:
                            records.append(rec)
                            existing.add(h)
                    if records:
                        _insert(records, table, engine)
                        inserted += len(records)
        except Exception as ex:
            print(f"[colocaciones] Error procesando {fname}: {ex}")

        if inserted:
            print(f"[colocaciones] {fname}: {inserted:,} -> {table}")
        else:
            print(f"[colocaciones] {fname}: sin filas nuevas.")
        total_new += inserted

    zf.close()
    return total_new


def _process_bruto_xlsm(zf: zipfile.ZipFile, fname: str,
                         table: str, row_parser,
                         engine, existing: set,
                         sheet_names: list[str] | None = None) -> int:
    """Lee un XLSM/XLSX de bases brutas y lo carga con el mismo parser TXT."""
    basename = fname.split("/")[-1]
    try:
        data = zf.read(fname)
        wb   = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as ex:
        print(f"[colocaciones] Error abriendo XLSM {basename}: {ex}")
        return 0

    ws_name = None
    if sheet_names:
        for sn in sheet_names:
            if sn in wb.sheetnames:
                ws_name = sn
                break
    if ws_name is None:
        candidates = [s for s in wb.sheetnames
                      if any(k in s.lower() for k in ("base", "datos", "data"))]
        ws_name = candidates[0] if candidates else (wb.sheetnames[0] if wb.sheetnames else None)
    if ws_name is None:
        wb.close()
        return 0

    ws   = wb[ws_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows or len(rows) < 2:
        return 0

    header   = [_norm_col(c) for c in rows[0]]
    batch    = []
    inserted = 0
    for row in rows[1:]:
        series = pd.Series(dict(zip(header, row)))
        rec = row_parser(series)
        if rec is None:
            continue
        h = _hash_rec(rec)
        rec["hash_registro"] = h
        if h not in existing:
            batch.append(rec)
            existing.add(h)
        if len(batch) >= _CHUNKSIZE:
            _insert(batch, table, engine)
            inserted += len(batch)
            batch = []
    if batch:
        _insert(batch, table, engine)
        inserted += len(batch)

    if inserted:
        print(f"[colocaciones] {basename} (XLSM): {inserted:,} -> {table}")
    else:
        print(f"[colocaciones] {basename} (XLSM): sin filas nuevas.")
    return inserted


# ---------------------------------------------------------------------------
# Loader tarjetas (ZIP con TXT / CSV / XLSM / XLSX)
# ---------------------------------------------------------------------------

def _tarjeta_is_con(name: str) -> bool:
    """True = 'con forma de pago'. Detecta por presencia de 'sin' en el nombre."""
    return "sin" not in name.lower()


def _load_tarjetas_zip(zip_path: Path, engine,
                       existing_con: set, existing_sin: set) -> tuple[int, int]:
    label = zip_path.name
    print(f"[colocaciones] Tarjetas: {label}")
    try:
        zf = zipfile.ZipFile(zip_path)
    except Exception as ex:
        print(f"[colocaciones] Error abriendo {label}: {ex}")
        return 0, 0

    _XLSM = {".xlsm", ".xlsx"}
    _CSV  = {".txt", ".csv"}
    new_con = new_sin = 0

    for fname in zf.namelist():
        basename = fname.split("/")[-1]
        ext = Path(basename).suffix.lower()

        if ext in _XLSM:
            c, s = _process_tarjetas_xlsm(zf, fname, basename, engine,
                                           existing_con, existing_sin)
            new_con += c
            new_sin += s
            continue

        if ext not in _CSV:
            continue

        is_con   = _tarjeta_is_con(basename)
        table    = _TABLE_TAR_CON if is_con else _TABLE_TAR_SIN
        existing = existing_con   if is_con else existing_sin
        parser   = _parse_tarjeta_con_row if is_con else _parse_tarjeta_sin_row

        try:
            with zf.open(fname) as f:
                header_bytes = f.read(4096)
        except Exception as ex:
            print(f"[colocaciones] Error leyendo {fname}: {ex}")
            continue

        enc = _detect_enc(header_bytes)
        sep = _detect_sep_bytes(header_bytes, enc)
        inserted = 0

        try:
            with zf.open(fname) as f:
                reader = pd.read_csv(
                    f, sep=sep, encoding=enc, dtype=str,
                    chunksize=_CHUNKSIZE, on_bad_lines="skip",
                    low_memory=False,
                )
                for chunk in reader:
                    chunk.columns = [_norm_col(c) for c in chunk.columns]
                    records = []
                    for _, row in chunk.iterrows():
                        rec = parser(row)
                        if rec is None:
                            continue
                        h = _hash_rec(rec)
                        rec["hash_registro"] = h
                        if h not in existing:
                            records.append(rec)
                            existing.add(h)
                    if records:
                        _insert(records, table, engine)
                        inserted += len(records)
        except Exception as ex:
            print(f"[colocaciones] Error procesando {fname}: {ex}")

        if inserted:
            print(f"[colocaciones] {basename}: {inserted:,} -> {table}")
        else:
            print(f"[colocaciones] {basename}: sin filas nuevas.")

        if is_con:
            new_con += inserted
        else:
            new_sin += inserted

    zf.close()
    return new_con, new_sin


def _process_tarjetas_xlsm(zf: zipfile.ZipFile, fname: str, basename: str,
                             engine, existing_con: set,
                             existing_sin: set) -> tuple[int, int]:
    """Lee XLSM/XLSX de tarjetas. Detecta 'con'/'sin' por hoja o por nombre de archivo."""
    try:
        data = zf.read(fname)
        wb   = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as ex:
        print(f"[colocaciones] Error abriendo XLSM {basename}: {ex}")
        return 0, 0

    file_is_con = _tarjeta_is_con(basename)
    new_con = new_sin = 0

    for sheet_name in wb.sheetnames:
        sn = sheet_name.lower()
        # Prefer sheet name to detect type; fall back to filename
        if "sin" in sn:
            is_con = False
        elif "con" in sn:
            is_con = True
        else:
            is_con = file_is_con

        table    = _TABLE_TAR_CON if is_con else _TABLE_TAR_SIN
        existing = existing_con   if is_con else existing_sin
        parser   = _parse_tarjeta_con_row if is_con else _parse_tarjeta_sin_row

        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            continue

        header   = [_norm_col(c) for c in rows[0]]
        batch    = []
        inserted = 0
        for row in rows[1:]:
            series = pd.Series(dict(zip(header, row)))
            rec = parser(series)
            if rec is None:
                continue
            h = _hash_rec(rec)
            rec["hash_registro"] = h
            if h not in existing:
                batch.append(rec)
                existing.add(h)
            if len(batch) >= _CHUNKSIZE:
                _insert(batch, table, engine)
                inserted += len(batch)
                batch = []
        if batch:
            _insert(batch, table, engine)
            inserted += len(batch)

        if inserted:
            print(f"[colocaciones] {basename}/{sheet_name}: {inserted:,} -> {table}")
        if is_con:
            new_con += inserted
        else:
            new_sin += inserted

    wb.close()
    return new_con, new_sin


# ---------------------------------------------------------------------------
# Parsers de fila XLSM
# ---------------------------------------------------------------------------

def _parse_volumen_row(header: list, row: tuple,
                       sector: int | None, is_mut: bool) -> dict | None:
    d = dict(zip(header, row))
    fecha_corte = _to_date(d.get("FECHACORTE") or d.get("FECHA CORTE")
                           or d.get("FECHA_CORTE"))
    if fecha_corte is None:
        return None
    rec = {
        "anio":              int(fecha_corte[:4]),
        "fecha_corte":       fecha_corte,
        "num_ruc":           _clean(d.get("NUMRUC") or d.get("NUM RUC")
                                    or d.get("NUM_RUC")),
        "razon_social":      _clean(d.get("RAZONSOCIAL") or d.get("RAZON SOCIAL")
                                    or d.get("RAZON_SOCIAL")),
        "segmento":          _clean(d.get("SEGMENTO")),
        "region":            _clean(d.get("REGION")),
        "provincia":         _clean(d.get("PROVINCIA")),
        "canton":            _clean(d.get("CANTON")),
        "tipo_credito_gral": _clean(d.get("TIPO DE CREDITO GENERAL")
                                    or d.get("TIPO DE CRDITO GENERAL")
                                    or d.get("TIPODECREDITOGENERAL")),
        "estado_operacion":  _clean(d.get("ESTADOOPERACION")
                                    or d.get("ESTADO OPERACION")
                                    or d.get("ESTADO_OPERACION")),
        "actividad_economica": _clean(d.get("ACTIVIDADECONOMICA")
                                      or d.get("ACTIVIDAD ECONOMICA")
                                      or d.get("ACTIVIDAD_ECONOMICA")),
        "destino_financiero":  _clean(d.get("DESTINOFINANCIERO")
                                      or d.get("DESTINO FINANCIERO")
                                      or d.get("DESTINO_FINANCIERO")),
        "tipo_operacion":    _clean(d.get("TIPOOPERACION") or d.get("TIPO OPERACION")
                                    or d.get("TIPO_OPERACION") or d.get("TIPO OPERACIN")),
        "val_operacion":     _to_float(d.get("VALOPERACION") or d.get("VAL OPERACION")
                                       or d.get("VAL_OPERACION")),
        "sujetos_credito":   _to_int(d.get("SUJETOS DE CREDITO")
                                     or d.get("SUJETOSDECREDITO")),
        "operaciones":       _to_int(d.get("OPERACIONES")),
    }
    if not is_mut:
        rec["sector"] = sector
    return rec


def _parse_col_row(header: list, row: tuple,
                   sector: int | None, is_mut: bool) -> dict | None:
    d = dict(zip(header, row))
    fecha_corte = _to_date(d.get("FECHA DE CORTE") or d.get("FECHADECORTE")
                           or d.get("FECHA_CORTE"))
    if fecha_corte is None:
        return None
    rec = {
        "anio":               int(fecha_corte[:4]),
        "cartera_por_vencer": _to_float(d.get("CARTERA POR VENCER")
                                        or d.get("CARTERAPORVENCER")),
        "cartera_no_devenga": _to_float(d.get("CARTERA QUE NO DEVENGA INTERESES")
                                        or d.get("CARTERAQUENODEVENGANINTERESES")),
        "cartera_vencida":    _to_float(d.get("CARTERA VENCIDA")
                                        or d.get("CARTERAVENCIDA")),
        "cartera_total":      _to_float(d.get("CARTERA TOTAL")
                                        or d.get("CARTERATOTAL")),
        "num_operaciones":    _to_int(d.get("NUMERO OPERACIONES")
                                      or d.get("NUMEROOPERACIONES")),
        "num_sujetos_credito": _to_int(d.get("NUMERO SUJETOS CREDITO")
                                       or d.get("NUMEROSUJETOSCREDITO")),
        "fecha_corte":        fecha_corte,
        "region":             _clean(d.get("REGION")),
        "provincia":          _clean(d.get("PROVINCIA")),
        "canton":             _clean(d.get("CANTON")),
        "subtipo_credito":    _clean(d.get("SUBTIPO DE CREDITO")
                                     or d.get("SUBTIPODECREDITO")),
        "origen_operacion":   _clean(d.get("ORIGEN OPERACION")
                                     or d.get("ORIGENOPERACION")),
        "linea_credito":      _clean(d.get("LINEA CREDITO") or d.get("LINEACREDITO")),
        "estado_operacion":   _clean(d.get("ESTADO OPERACION")
                                     or d.get("ESTADOOPERACION")),
        "clase_credito":      _clean(d.get("CLASE DE CREDITO")
                                     or d.get("CLASEDECREDITO")),
        "actividad_economica": _clean(d.get("ACTIVIDAD ECONOMICA")
                                      or d.get("ACTIVIDADECONOMICA")),
        "ruc":                _clean(d.get("RUC")),
        "razon_social":       _clean(d.get("RAZON SOCIAL") or d.get("RAZONSOCIAL")
                                     or d.get("RAZON SOCIAL2")),
    }
    if not is_mut:
        rec["sector"] = sector
    return rec


# ---------------------------------------------------------------------------
# Parsers de fila bruto TXT
# ---------------------------------------------------------------------------

def _parse_vol_bruto_row(row: pd.Series) -> dict | None:
    # Dos formatos:
    #   Formato A (2018-2021, anterior): separador ';', columnas sin espacios
    #     FECHADECORTE, TIPOOPERACION, CODIGOACTIVIDADECONOMICA ...
    #   Formato B (2022+): separador '\t', columnas con espacios
    #     FECHA DE CORTE, TIPO OPERACION, CODIGO ACTIVIDAD ECONOMICA ...
    fecha_corte = _to_date(_sq(row.get("FECHA DE CORTE") or row.get("FECHADECORTE")))
    if fecha_corte is None:
        return None
    return {
        "anio":  int(fecha_corte[:4]),
        "tipo_operacion": _sq(
            row.get("TIPO OPERACION") or row.get("TIPOOPERACION")),
        "estado_operacion": _sq(
            row.get("ESTADO OPERACION") or row.get("ESTADOOPERACION")),
        "segmento": _sq(row.get("SEGMENTO")),
        "cod_actividad_eco": _sq(
            row.get("CODIGO ACTIVIDAD ECONOMICA") or row.get("CODIGOACTIVIDADECONOMICA")),
        "acteco_seccion": _sq(
            row.get("ACTIVIDAD ECONOMICA SECCION") or row.get("ACTIVIDADECONOMICASECCION")),
        "acteco_division": _sq(
            row.get("ACTIVIDAD ECONOMICA DIVISION") or row.get("ACTIVIDADECONOMICADIVISION")),
        "acteco_grupo": _sq(
            row.get("ACTIVIDAD ECONOMICA GRUPO") or row.get("ACTIVIDADECONOMICAGRUPO")),
        "acteco_clase": _sq(
            row.get("ACTIVIDAD ECONOMICA CLASE") or row.get("ACTIVIDADECONOMICACLASE")),
        "acteco_subclase": _sq(
            row.get("ACTIVIDAD ECONOMICA SUBCLASE") or row.get("ACTIVIDADECONOMICASUBCLASE")),
        "actividad_economica": _sq(
            row.get("ACTIVIDAD ECONOMICA") or row.get("ACTIVIDADECONOMICA")),
        "canton": _sq(row.get("CANTON")),
        "cod_dpa": _sq(
            row.get("CODIG DPA") or row.get("CODIGO DPA") or row.get("CODIGODPA")),
        "parroquia": _sq(row.get("PARROQUIA")),
        "provincia": _sq(row.get("PROVINCIA")),
        "fecha_corte": fecha_corte,
        "tipo_credito_general": _sq(
            row.get("TIPO CREDITO GENERAL") or row.get("TIPOCREDITOGENERAL")),
        "instruccion": _sq(row.get("INSTRUCCION")),
        "sexo": _sq(row.get("SEXO")),
        "destino_financiero": _sq(
            row.get("DESTINO FINANCIERO") or row.get("DESTINOFINANCIERO")),
        "rango_edad": _sq(row.get("RANGO EDAD") or row.get("RANGOEDAD")),
        "rango_monto_credito": _sq(
            row.get("RANGO MONTO CREDITO CONCEDIDO") or row.get("RANGOMONTOCREDITOCONCEDIDO")),
        "rango_plazo_original": _sq(
            row.get("RANGO PLAZO ORIGINAL CONCESION")
            or row.get("RANGOPLAZOORIGINALCONCESION")
            or row.get("RANGOPLAZOORIGINALCONC")),
        "tipo_persona": _sq(
            row.get("TIPO PERSONA") or row.get("TIPOPERSONA")),
        "tipo_credito_espec": _sq(
            row.get("TIPO DE CREDITO") or row.get("TIPO DE CRDITO")
            or row.get("TIPODECREDITO") or row.get("TIPODECR DITO")),
        # NRO.SUJETOSCREDITO (fmt A) → norm → "NRO SUJETOSCREDITO"
        "nro_sujetos_credito": _to_int(
            row.get("NRO SUJETOS CREDITO") or row.get("NRO SUJETOSCREDITO")
            or row.get("NRO. SUJETOS CREDITO")),
        "nro_operaciones": _to_int(
            row.get("NRO OPERACIONES") or row.get("NROOPERACIONES")
            or row.get("NRO. OPERACIONES")),
        "monto_concedido_usd": _to_float(
            row.get("MONTO CONCEDIDO USD") or row.get("MONTOCONCEDIDOUSD")),
    }


def _parse_col_bruto_row(row: pd.Series) -> dict | None:
    # Dos formatos:
    #   Formato A (dat 2017-2021): separador '\t' o ';', columnas sin espacios
    #     FECHADECORTE, SEGMENTO, TIPOOPERACION ...
    #   Formato B (txt 2022+): separador '\t', columnas con espacios
    #     FECHA DE CORTE, SEGMENTO, TIPO OPERACION (C02) ...
    # Nota Formato B: las columnas de valor NO llevan "USD" al final
    #   VALOR POR VENCER (no VALOR POR VENCER USD)
    #   VALOR VENCIDO (USD) -> norm -> VALOR VENCIDO USD
    #   VALOR SALDO TOTAL (USD) -> norm -> VALOR SALDO TOTAL USD
    fecha_corte = _to_date(_sq(row.get("FECHA DE CORTE") or row.get("FECHADECORTE")))
    if fecha_corte is None:
        return None
    return {
        "anio":  int(fecha_corte[:4]),
        "fecha_corte": fecha_corte,
        "segmento": _sq(row.get("SEGMENTO")),
        "tipo_operacion": _sq(
            row.get("TIPO OPERACION C02") or row.get("TIPO OPERACION")
            or row.get("TIPOOPERACION")),
        "tipo_credito_general": _sq(
            row.get("TIPO CREDITO GENERAL") or row.get("TIPOCREDITOGENERAL")),
        "tipo_credito_espec": _sq(
            row.get("TIPO CREDITO ESPECIFICO") or row.get("TIPO CREDITO ESPCIFICO")
            or row.get("TIPOCREDITOESPECIFICO") or row.get("TIPO CREDITO ESPECFICO")),
        "cod_destino_fin": _sq(
            row.get("CODIGO DESTINO FINANCIERO") or row.get("CODIGODESTINO FINANCIERO")),
        "destino_financiero": _sq(
            row.get("DESTINO FINANCIERO") or row.get("DESTINOFINANCIERO")),
        "cod_actividad_eco": _sq(
            row.get("CODIGO ACTIVIDAD ECONOMICA") or row.get("CODIGOACTIVIDADECONOMICA")),
        "acteco_seccion": _sq(
            row.get("ACTIVIDAD ECONOMICA SECCION") or row.get("ACTIVIDADECONOMICASECCION")),
        "acteco_division": _sq(
            row.get("ACTIVIDAD ECONOMICA DIVISION") or row.get("ACTIVIDADECONOMICADIVISION")),
        "acteco_grupo": _sq(
            row.get("ACTIVIDAD ECONOMICA GRUPO") or row.get("ACTIVIDADECONOMICAGRUPO")),
        "acteco_clase": _sq(
            row.get("ACTIVIDAD ECONOMICA CLASE") or row.get("ACTIVIDADECONOMICACLASE")),
        "acteco_subclase": _sq(
            row.get("ACTIVIDAD ECONOMICA SUBCLASE") or row.get("ACTIVIDADECONOMICASUBCLASE")),
        "actividad_economica": _sq(
            row.get("ACTIVIDAD ECONOMICA") or row.get("ACTIVIDADECONOMICA")),
        "provincia": _sq(row.get("PROVINCIA")),
        "canton": _sq(row.get("CANTON")),
        "parroquia": _sq(row.get("PARROQUIA")),
        "cod_dpa": _sq(
            row.get("CODIGO DPA") or row.get("CODIG DPA") or row.get("CODIGODPA")),
        "tipo_persona": _sq(
            row.get("TIPO PERSONA") or row.get("TIPOPERSONA")),
        "sexo": _sq(row.get("SEXO")),
        "rango_edad": _sq(row.get("RANGO EDAD") or row.get("RANGOEDAD")),
        "nivel_instruccion": _sq(
            row.get("NIVEL INSTRUCCION") or row.get("INSTRUCCION")
            or row.get("NIVELINSTRUCCION")),
        "rango_saldo": _sq(row.get("RANGO SALDO") or row.get("RANGOSALDO")),
        "nro_sujetos": _to_int(
            row.get("NRO SUJETOS") or row.get("NROSUJETOS")
            or row.get("NRO. SUJETOS")),
        "nro_operaciones": _to_int(
            row.get("NRO OPERACIONES") or row.get("NROOPERACIONES")
            or row.get("NRO. OPERACIONES")),
        # Formato B: "VALOR POR VENCER" (sin USD); formato A puede traer "VALOR POR VENCER USD"
        "valor_por_vencer_usd": _to_float(
            row.get("VALOR POR VENCER USD") or row.get("VALOR POR VENCER")
            or row.get("VALORPORVENCER")),
        # VALOR NO DEVENGA INTERESES (USD) -> norm -> VALOR NO DEVENGA INTERESES USD
        "valor_no_devenga_usd": _to_float(
            row.get("VALOR NO DEVENGA INTERESES USD")
            or row.get("VALOR NO DEVENGA INTERESES")),
        # VALOR VENCIDO (USD) -> norm -> VALOR VENCIDO USD
        "valor_vencido_usd": _to_float(
            row.get("VALOR VENCIDO USD") or row.get("VALOR VENCIDO")),
        # VALOR SALDO TOTAL (USD) -> norm -> VALOR SALDO TOTAL USD
        "valor_saldo_total_usd": _to_float(
            row.get("VALOR SALDO TOTAL USD") or row.get("VALOR SALDO TOTAL")),
    }


# ---------------------------------------------------------------------------
# Parsers de fila tarjetas
# ---------------------------------------------------------------------------

def _parse_tarjeta_con_row(row: pd.Series) -> dict | None:
    fecha = _to_date(_sq(row.get("FECHA")))
    if fecha is None:
        return None
    return {
        "anio":              int(fecha[:4]),
        "fecha":             fecha,
        "forma_pago":        _sq(row.get("FORMA PAGO")),
        "sexo":              _sq(row.get("SEXO")),
        "instruccion":       _sq(row.get("INSTRUCCION")),
        "rango_edad":        _sq(row.get("RANGO EDAD")),
        "capital_por_vencer": _to_float_eu(row.get("CAPITAL POR VENCER")),
        "capital_vencido":    _to_float_eu(row.get("CAPITAL VENCIDO")),
        "capital_no_devenga": _to_float_eu(row.get("CAPITAL NO DEVENGA INTERESES")),
        "capital_consumo":    _to_float_eu(row.get("CAPITAL CONSUMO")),
    }


def _parse_tarjeta_sin_row(row: pd.Series) -> dict | None:
    fecha = _to_date(_sq(row.get("FECHA")))
    if fecha is None:
        return None
    return {
        "anio":                int(fecha[:4]),
        "fecha":               fecha,
        "sexo":                _sq(row.get("SEXO")),
        "instruccion":         _sq(row.get("INSTRUCCION")),
        "rango_edad":          _sq(row.get("RANGO EDAD")),
        "num_tarjetahabientes": _to_int(row.get("NUMERO TARJETAHABIENTES")),
        "num_tarjetas":         _to_int(row.get("NUMERO TARJETAS")),
        "capital_por_vencer":  _to_float_eu(row.get("CAPITAL POR VENCER")),
        "capital_vencido":     _to_float_eu(row.get("CAPITAL VENCIDO")),
        "capital_no_devenga":  _to_float_eu(row.get("CAPITAL NO DEVENGA INTERESES")),
        "capital_consumo":     _to_float_eu(row.get("CAPITAL CONSUMO")),
    }


# ---------------------------------------------------------------------------
# Inserción
# ---------------------------------------------------------------------------

def _insert(records: list[dict], table: str, engine) -> None:
    if not records:
        return
    cols         = list(records[0].keys())
    col_list     = ", ".join(cols)
    placeholders = ", ".join(f":{c}" for c in cols)
    with engine.begin() as conn:
        conn.execute(
            text(f"INSERT INTO {table} ({col_list}) VALUES ({placeholders})"),
            records,
        )


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _norm_col(c: str) -> str:
    """Normaliza nombre de columna eliminando acentos, BOM y chars especiales."""
    s = str(c).strip()
    s = s.replace("�", "")           # reemplazo openpyxl por char no-decodificable
    s = re.sub(r"^[^\x20-\x7E]+", "", s)  # strip BOM / no-ASCII al inicio
    nfkd = unicodedata.normalize("NFKD", s)
    ascii_s = "".join(ch for ch in nfkd if unicodedata.category(ch) != "Mn")
    ascii_s = re.sub(r"[^A-Za-z0-9 _]", " ", ascii_s)
    return re.sub(r"\s+", " ", ascii_s).strip().upper()


def _sq(v) -> str | None:
    """Strip quotes y limpia valor de celda CSV."""
    if v is None:
        return None
    s = str(v).strip().strip('"').strip()
    return s if s and s.lower() not in ("nan", "none") else None


def _clean(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s.lower() not in ("nan", "none") else None


def _to_float(v) -> float | None:
    if v is None:
        return None
    s = str(v).strip().strip('"').replace(",", "")
    try:
        f = float(s)
        return None if (np.isnan(f) or np.isinf(f)) else f
    except (ValueError, TypeError):
        return None


def _to_float_eu(v) -> float | None:
    """Convierte formato europeo '19.250.868,38' o '19250868,38' a float."""
    if v is None:
        return None
    s = str(v).strip().strip('"')
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        f = float(s)
        return None if (np.isnan(f) or np.isinf(f)) else f
    except (ValueError, TypeError):
        return None


def _to_int(v) -> int | None:
    if v is None:
        return None
    try:
        return int(float(str(v).strip().strip('"')))
    except (ValueError, TypeError):
        return None


def _to_date(v) -> str | None:
    """Convierte múltiples formatos a 'YYYY-MM-DD'."""
    if v is None:
        return None
    from datetime import datetime as _d, date as _da
    if isinstance(v, (_d, _da)):
        return (v.date() if isinstance(v, _d) else v).strftime("%Y-%m-%d")
    # Excel serial
    if isinstance(v, (int, float)) and 40000 < v < 60000:
        try:
            return (_dt(1899, 12, 30) + _td(days=int(v))).strftime("%Y-%m-%d")
        except Exception:
            pass
    s = str(v).strip().rstrip(",").strip()
    if not s or s.lower() in ("nan", "none"):
        return None
    # ISO: 2026-01-31 o 2026-01-31 00:00:00,000
    m = re.match(r"(\d{4}-\d{2}-\d{2})", s)
    if m:
        return m.group(1)
    # DD/MM/YYYY
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        return f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
    return None


def _hash_rec(rec: dict) -> str:
    key = "|".join(str(v) if v is not None else "" for v in rec.values())
    return hashlib.sha256(key.encode()).hexdigest()


def _detect_sector(fname: str) -> int | None:
    m = re.search(r"[_\-\s]S([123])[_\-\.\s]", fname, re.IGNORECASE)
    if m:
        return int(m.group(1))
    m = re.search(r"S([123])$", fname.rsplit(".", 1)[0], re.IGNORECASE)
    return int(m.group(1)) if m else None


def _is_mutualista(fname: str) -> bool:
    return bool(re.search(r"Mut", fname, re.IGNORECASE))


def _year_from_zippath(fname: str) -> int | None:
    m = re.match(r"^(\d{4})[/\-]", fname)
    return int(m.group(1)) if m else None


def _detect_enc(raw: bytes) -> str:
    try:
        raw.decode("utf-8")
        return "utf-8-sig"
    except UnicodeDecodeError:
        return "latin-1"


def _detect_sep_bytes(raw: bytes, enc: str) -> str:
    line = raw.decode(enc, errors="replace").split("\n")[0]
    counts = {"\t": line.count("\t"), "|": line.count("|"), ";": line.count(";")}
    return max(counts, key=counts.get)
