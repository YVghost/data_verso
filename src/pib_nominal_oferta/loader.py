"""
ETL loader: PIB Nominal — Oferta y Utilización de Bienes y Servicios (BCE)

Fuente: Excel trimestrales BCE CNT (tou_{cod}_{YYYYQQ}.xlsx)

Tablas destino:
  pib_nominal_oferta_bruto     — hojas _bru  (Dem_Corr_bru, Dem_Cad_bru,
                                               Dem_Ivol_bru)
  pib_nominal_oferta_ajustado  — hojas _ajus (Dem_Corr_ajus, Dem_Cad_ajus,
                                               Dem_Ivol_ajus, _vY, _vQ,
                                               Dem_Prec_ajus, _vY,
                                               Dem_contr_ajus_vY, _vQ)

Estructura Excel (todas las hojas de datos):
  Fila 9  : encabezados de variable (col A = "Variables", B en adelante)
  Fila 10 : "Trimestres"
  Fila 11 : vacía
  Filas 12+: datos — col A = "YYYY.I/II/III/IV", cols B+ = valores float

Columna tipo_indice (columna diferenciadora dentro de cada tabla):
  Bruto:    nombre del índice de medición (sin "Datos Brutos" — ya está
            implícito en la tabla bruto)
  Ajustado: índice + tipo de ajuste + variación si aplica

Deduplicación: SHA-256 sobre (anio, trimestre, variable, tipo_indice).
"""

import hashlib
import re
import sys
from pathlib import Path

import openpyxl
from sqlalchemy import text

sys.path.append(str(Path(__file__).resolve().parents[2]))
from utils.base_engine import get_master_engine

_TABLE_BRUTO = "pib_nominal_oferta_bruto"
_TABLE_AJUS  = "pib_nominal_oferta_ajustado"

_DDL_BRUTO = """
CREATE TABLE pib_nominal_oferta_bruto (
    id            BIGINT IDENTITY(1,1) NOT NULL,
    anio          INT            NOT NULL,
    trimestre     NVARCHAR(5)    NOT NULL,
    variable      NVARCHAR(300)  NOT NULL,
    valor         FLOAT          NULL,
    tipo_indice   NVARCHAR(300)  NOT NULL,
    nombre_hoja   NVARCHAR(50)   NOT NULL,
    hash_registro NVARCHAR(64)   NOT NULL,
    fecha_carga   DATETIME2      DEFAULT GETDATE(),
    CONSTRAINT PK_pib_nom_ofe_bruto PRIMARY KEY NONCLUSTERED (id)
)"""

_IDX_BRUTO = (
    "CREATE CLUSTERED INDEX CIX_pib_nom_ofe_bruto "
    "ON pib_nominal_oferta_bruto (anio, trimestre)"
)

_DDL_AJUS = """
CREATE TABLE pib_nominal_oferta_ajustado (
    id            BIGINT IDENTITY(1,1) NOT NULL,
    anio          INT            NOT NULL,
    trimestre     NVARCHAR(5)    NOT NULL,
    variable      NVARCHAR(300)  NOT NULL,
    valor         FLOAT          NULL,
    tipo_indice   NVARCHAR(300)  NOT NULL,
    nombre_hoja   NVARCHAR(50)   NOT NULL,
    hash_registro NVARCHAR(64)   NOT NULL,
    fecha_carga   DATETIME2      DEFAULT GETDATE(),
    CONSTRAINT PK_pib_nom_ofe_ajus PRIMARY KEY NONCLUSTERED (id)
)"""

_IDX_AJUS = (
    "CREATE CLUSTERED INDEX CIX_pib_nom_ofe_ajus "
    "ON pib_nominal_oferta_ajustado (anio, trimestre)"
)

# ---------------------------------------------------------------------------
# Mapping hoja → (tabla, tipo_indice)
# tipo_indice para brutas NO incluye "Datos Brutos" (redundante con la tabla)
# tipo_indice para ajustadas conserva "Datos Ajustados de Estacionalidad"
# ---------------------------------------------------------------------------
_SHEET_MAP: dict[str, tuple[str, str]] = {
    # ── Datos Brutos ────────────────────────────────────────────────────────
    "Dem_Corr_bru":      (_TABLE_BRUTO, "Precios Corrientes"),
    "Dem_Cad_bru":       (_TABLE_BRUTO, "Precios Encadenados"),
    "Dem_Ivol_bru":      (_TABLE_BRUTO, "Índices de Volumen Encadenados, 2018=100"),
    # ── Datos Ajustados de Estacionalidad ────────────────────────────────────
    "Dem_Corr_ajus":     (_TABLE_AJUS,  "Precios Corrientes, Datos Ajustados de Estacionalidad"),
    "Dem_Cad_ajus":      (_TABLE_AJUS,  "Precios Encadenados, Datos Ajustados de Estacionalidad"),
    "Dem_Ivol_ajus":     (_TABLE_AJUS,  "Índices de Volumen Encadenados, 2018=100, Datos Ajustados de Estacionalidad"),
    "Dem_Ivol_ajus_vY":  (_TABLE_AJUS,  "Índices de Volumen Encadenados, 2018=100, Datos Ajustados de Estacionalidad, Variación Interanual"),
    "Dem_Ivol_ajus_vQ":  (_TABLE_AJUS,  "Índices de Volumen Encadenados, 2018=100, Datos Ajustados de Estacionalidad, Variación Intertrimestral"),
    "Dem_Prec_ajus":     (_TABLE_AJUS,  "Deflactores de Precios Encadenados, Datos Ajustados de Estacionalidad"),
    "Dem_Prec_ajus_vY":  (_TABLE_AJUS,  "Deflactores de Precios Encadenados, Datos Ajustados de Estacionalidad, Variación Interanual"),
    "Dem_contr_ajus_vY": (_TABLE_AJUS,  "Contribución al Crecimiento, Datos Ajustados de Estacionalidad, Variación Interanual"),
    "Dem_contr_ajus_vQ": (_TABLE_AJUS,  "Contribución al Crecimiento, Datos Ajustados de Estacionalidad, Variación Intertrimestral"),
}

_TRIM_RE        = re.compile(r"^(\d{4})\.(I{1,3}V?|IV)$", re.IGNORECASE)
_HDR_ROW        = 9    # fila con nombres de variable (1-indexed en openpyxl)
_DATA_ROW_START = 12   # primera fila de datos


# ---------------------------------------------------------------------------
# Punto de entrada
# ---------------------------------------------------------------------------

def load(files: list[Path]) -> None:
    if not files:
        print("[pib_oferta] Sin archivos para cargar.")
        return

    engine = get_master_engine()
    _ensure_tables(engine)

    ex_bruto = _get_existing_hashes(engine, _TABLE_BRUTO)
    ex_ajus  = _get_existing_hashes(engine, _TABLE_AJUS)

    tot_bruto = tot_ajus = 0

    for path in sorted(files):
        nb, na = _load_file(path, engine, ex_bruto, ex_ajus)
        tot_bruto += nb
        tot_ajus  += na

    print(
        f"[pib_oferta] Total: {tot_bruto} en {_TABLE_BRUTO}, "
        f"{tot_ajus} en {_TABLE_AJUS}."
    )


# ---------------------------------------------------------------------------
# DDL
# ---------------------------------------------------------------------------

def _ensure_tables(engine) -> None:
    from sqlalchemy import inspect as sa_inspect
    insp = sa_inspect(engine)
    if not insp.has_table(_TABLE_BRUTO):
        with engine.begin() as conn:
            conn.execute(text(_DDL_BRUTO))
            conn.execute(text(_IDX_BRUTO))
        print(f"[pib_oferta] Tabla {_TABLE_BRUTO} creada.")
    if not insp.has_table(_TABLE_AJUS):
        with engine.begin() as conn:
            conn.execute(text(_DDL_AJUS))
            conn.execute(text(_IDX_AJUS))
        print(f"[pib_oferta] Tabla {_TABLE_AJUS} creada.")


def _get_existing_hashes(engine, table: str) -> set[str]:
    with engine.connect() as conn:
        rows = conn.execute(
            text(f"SELECT hash_registro FROM {table}")
        ).fetchall()
    return {r[0] for r in rows}


# ---------------------------------------------------------------------------
# Por archivo
# ---------------------------------------------------------------------------

def _load_file(
    path: Path,
    engine,
    ex_bruto: set[str],
    ex_ajus:  set[str],
) -> tuple[int, int]:
    print(f"[pib_oferta] Procesando {path.name}...")
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as ex:
        print(f"[pib_oferta] Error abriendo {path.name}: {ex}")
        return 0, 0

    n_bruto = n_ajus = 0

    for sh_name in wb.sheetnames:
        if sh_name not in _SHEET_MAP:
            continue
        table, tipo_indice = _SHEET_MAP[sh_name]
        ws = wb[sh_name]

        records = _parse_sheet(ws, tipo_indice, sh_name)
        if not records:
            continue

        existing = ex_bruto if table == _TABLE_BRUTO else ex_ajus
        new = [r for r in records if r["hash_registro"] not in existing]
        if not new:
            continue

        _insert(new, table, engine)
        existing.update(r["hash_registro"] for r in new)

        if table == _TABLE_BRUTO:
            n_bruto += len(new)
        else:
            n_ajus += len(new)
        print(f"[pib_oferta]   {sh_name}: {len(new)} filas nuevas.")

    wb.close()
    return n_bruto, n_ajus


# ---------------------------------------------------------------------------
# Parsing de hoja
# ---------------------------------------------------------------------------

def _parse_sheet(ws, tipo_indice: str, sh_name: str) -> list[dict]:
    # Encabezados de variable — fila 9, columna B en adelante
    hdr = next(ws.iter_rows(min_row=_HDR_ROW, max_row=_HDR_ROW, values_only=True))
    variables = [str(v).strip() for v in hdr[1:] if v is not None]

    records: list[dict] = []

    for row in ws.iter_rows(min_row=_DATA_ROW_START, values_only=True):
        period_raw = row[0]
        if period_raw is None:
            continue

        m = _TRIM_RE.match(str(period_raw).strip())
        if not m:
            continue

        anio      = int(m.group(1))
        trimestre = m.group(2).upper()
        values    = row[1:]

        for i, var in enumerate(variables):
            if i >= len(values):
                break
            v = values[i]
            if v is None:
                continue
            try:
                valor = float(v)
            except (TypeError, ValueError):
                continue

            records.append({
                "anio":          anio,
                "trimestre":     trimestre,
                "variable":      var,
                "valor":         valor,
                "tipo_indice":   tipo_indice,
                "nombre_hoja":   sh_name,
                "hash_registro": _hash(anio, trimestre, var, tipo_indice),
            })

    return records


# ---------------------------------------------------------------------------
# Inserción e helpers
# ---------------------------------------------------------------------------

def _insert(records: list[dict], table: str, engine) -> None:
    cols         = list(records[0].keys())
    col_list     = ", ".join(cols)
    placeholders = ", ".join(f":{c}" for c in cols)
    BATCH = 2000
    with engine.begin() as conn:
        for i in range(0, len(records), BATCH):
            conn.execute(
                text(f"INSERT INTO {table} ({col_list}) VALUES ({placeholders})"),
                records[i : i + BATCH],
            )


def _hash(anio: int, trimestre: str, variable: str, tipo_indice: str) -> str:
    key = f"{anio}|{trimestre}|{variable}|{tipo_indice}"
    return hashlib.sha256(key.encode()).hexdigest()
