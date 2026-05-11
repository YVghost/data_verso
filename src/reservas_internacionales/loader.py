"""
ETL: Reservas Internacionales — BCE Ecuador

Hojas procesadas:
  "Anual 2000-2025"      -> tabla reservas_internacionales_anual
  "Mensual 2000 abr- 2026" -> tabla reservas_internacionales_mensual

Formato: largo (una fila por indicador x periodo).

Esquema ambas tablas:
  id, anio, [mes], indicador, valor_millones, hash_registro, fecha_carga

Indicadores (~10):
  Posicion neta en divisas, Caja en divisas, Depositos netos exterior,
  Inversiones depositos plazo y titulos, Oro, DEG,
  Posicion de reserva en FMI, Posicion con ALADI, Posicion SUCRE,
  RI (total)
"""

import hashlib
import re
import sys
from pathlib import Path

import openpyxl
from sqlalchemy import text

sys.path.append(str(Path(__file__).resolve().parents[2]))
from utils.base_engine import get_master_engine

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------

_TABLE_ANUAL   = "reservas_internacionales_anual"
_TABLE_MENSUAL = "reservas_internacionales_mensual"

_SHEET_ANUAL   = "Anual 2000-2025"
_SHEET_MENSUAL = "Mensual 2000 abr- 2026"

_MES_MAP = {
    "Ene": "Enero",   "Feb": "Febrero",  "Mar": "Marzo",
    "Abr": "Abril",   "May": "Mayo",     "Jun": "Junio",
    "Jul": "Julio",   "Ago": "Agosto",   "Sep": "Septiembre",
    "Oct": "Octubre", "Nov": "Noviembre","Dic": "Diciembre",
}

_DDL_ANUAL = """
CREATE TABLE reservas_internacionales_anual (
    id             BIGINT IDENTITY(1,1) NOT NULL,
    anio           INT            NOT NULL,
    indicador      NVARCHAR(300)  NOT NULL,
    valor_millones FLOAT          NULL,
    hash_registro  NVARCHAR(64)   NOT NULL,
    fecha_carga    DATETIME2      DEFAULT GETDATE(),
    CONSTRAINT PK_reservas_internacionales_anual PRIMARY KEY NONCLUSTERED (id)
)"""

_IDX_ANUAL = """
CREATE CLUSTERED INDEX CIX_reservas_internacionales_anual
ON reservas_internacionales_anual (anio, indicador)"""

_DDL_MENSUAL = """
CREATE TABLE reservas_internacionales_mensual (
    id             BIGINT IDENTITY(1,1) NOT NULL,
    anio           INT            NOT NULL,
    mes            NVARCHAR(20)   NOT NULL,
    indicador      NVARCHAR(300)  NOT NULL,
    valor_millones FLOAT          NULL,
    hash_registro  NVARCHAR(64)   NOT NULL,
    fecha_carga    DATETIME2      DEFAULT GETDATE(),
    CONSTRAINT PK_reservas_internacionales_mensual PRIMARY KEY NONCLUSTERED (id)
)"""

_IDX_MENSUAL = """
CREATE CLUSTERED INDEX CIX_reservas_internacionales_mensual
ON reservas_internacionales_mensual (anio, mes, indicador)"""

# Regex para limpiar prefijos de numeracion: "1. ", "1.1 ", "B.1.a. "
_OUTLINE_RE = re.compile(r"^(?:(?:\d+|[A-Za-z])\.)+\d*\s+")


# ---------------------------------------------------------------------------
# Punto de entrada
# ---------------------------------------------------------------------------

def load(file: Path) -> None:
    """Carga las hojas Anual y Mensual del Excel en sus respectivas tablas."""
    if file is None or not file.exists():
        print("[ri] Archivo no encontrado, omitiendo carga.")
        return

    print(f"[ri] Procesando {file.name}...")
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)

    engine = get_master_engine()
    _ensure_tables(engine)

    # Hoja anual
    if _SHEET_ANUAL in wb.sheetnames:
        rows_a = list(wb[_SHEET_ANUAL].iter_rows(values_only=True))
        records_a = _parse_anual(rows_a)
        print(f"[ri] Anual: {len(records_a)} registros parseados")
        _upsert(_TABLE_ANUAL, records_a, ["anio", "indicador"], engine)
    else:
        print(f"[ri] [warn] Hoja '{_SHEET_ANUAL}' no encontrada")

    # Hoja mensual
    if _SHEET_MENSUAL in wb.sheetnames:
        rows_m = list(wb[_SHEET_MENSUAL].iter_rows(values_only=True))
        records_m = _parse_mensual(rows_m)
        print(f"[ri] Mensual: {len(records_m)} registros parseados")
        _upsert(_TABLE_MENSUAL, records_m, ["anio", "mes", "indicador"], engine)
    else:
        print(f"[ri] [warn] Hoja '{_SHEET_MENSUAL}' no encontrada")

    wb.close()


# ---------------------------------------------------------------------------
# Parsers
# ---------------------------------------------------------------------------

def _parse_anual(rows: list) -> list[dict]:
    """
    Anual sheet layout:
      row 7 (idx): year headers at col 3+
      rows 8-17: indicators
        label = col 1 (B) if not None else col 0 (A)
        data starts at col 3
    """
    year_row = rows[7] if len(rows) > 7 else []
    years: dict[int, int] = {}  # col_idx -> year
    for ci, v in enumerate(year_row):
        if ci < 3 or v is None:
            continue
        try:
            y = int(float(str(v)))
            if 1900 <= y <= 2100:
                years[ci] = y
        except (ValueError, TypeError):
            pass

    records = []
    for ri in range(8, len(rows)):
        r = rows[ri]
        label = _get_label(r, col_a=0, col_b=1)
        if label is None:
            continue
        for ci, year in years.items():
            val = _to_float(r[ci] if ci < len(r) else None)
            if val is None:
                continue
            rec = {"anio": year, "indicador": label, "valor_millones": val}
            rec["hash_registro"] = _hash(["anio", "indicador", "valor_millones"], rec)
            records.append(rec)

    return records


def _parse_mensual(rows: list) -> list[dict]:
    """
    Mensual sheet layout:
      row 7 (idx): year headers (col 4 = 2000, then every 12 cols)
      row 8: month abbreviations
      rows 9-18: indicators
        label = col 2 (C) if not None else col 1 (B)
        data starts at col 4
    """
    year_row  = rows[7] if len(rows) > 7 else []
    month_row = rows[8] if len(rows) > 8 else []

    col_map: dict[int, tuple] = {}  # col_idx -> (year, mes_nombre)
    current_year: int | None = None
    for ci in range(4, len(year_row)):
        yv = year_row[ci] if ci < len(year_row) else None
        if yv is not None:
            try:
                current_year = int(float(str(yv)))
            except (ValueError, TypeError):
                pass
        mv = month_row[ci] if ci < len(month_row) else None
        if mv is not None and current_year is not None:
            mes = _MES_MAP.get(str(mv).strip())
            if mes:
                col_map[ci] = (current_year, mes)

    records = []
    for ri in range(9, len(rows)):
        r = rows[ri]
        label = _get_label(r, col_a=1, col_b=2)
        if label is None:
            continue
        for ci, (year, mes) in col_map.items():
            val = _to_float(r[ci] if ci < len(r) else None)
            if val is None:
                continue
            rec = {"anio": year, "mes": mes, "indicador": label, "valor_millones": val}
            rec["hash_registro"] = _hash(["anio", "mes", "indicador", "valor_millones"], rec)
            records.append(rec)

    return records


# ---------------------------------------------------------------------------
# Helpers de parseo
# ---------------------------------------------------------------------------

def _get_label(row: tuple, col_a: int, col_b: int) -> str | None:
    """col_b preferred over col_a; applies _clean_label."""
    b = row[col_b] if col_b < len(row) else None
    a = row[col_a] if col_a < len(row) else None
    raw = b if (b is not None and str(b).strip()) else a
    if raw is None:
        return None
    return _clean_label(str(raw))


def _clean_label(s: str) -> str | None:
    s = s.strip()
    if not s:
        return None
    # Skip footnote/source rows
    if re.match(r"^\([*\d]+\)", s):
        return None
    if s.lower().startswith(("fuente", "source", "nota:", "note:", "en millones", "in millions")):
        return None
    if len(s) > 295:
        return None
    # Strip footnote markers at end: "(***)  " suffix content is already filtered above
    # Strip trailing footnote references like " (***)", " (1)", etc.
    s = re.sub(r"\s*\([*\d]+\)\s*$", "", s).strip()
    # Strip leading outline numbering: "1. ", "1.1 ", "B.1.a. "
    s = _OUTLINE_RE.sub("", s).strip()
    return s or None


def _to_float(v) -> float | None:
    if v is None:
        return None
    try:
        f = float(v)
        return f
    except (ValueError, TypeError):
        return None


def _hash(keys: list, rec: dict) -> str:
    key = "|".join(str(rec.get(k, "")) for k in keys)
    return hashlib.sha256(key.encode()).hexdigest()


# ---------------------------------------------------------------------------
# DDL y carga
# ---------------------------------------------------------------------------

def _ensure_tables(engine) -> None:
    from sqlalchemy import inspect as sa_inspect
    insp = sa_inspect(engine)
    with engine.begin() as conn:
        if not insp.has_table(_TABLE_ANUAL):
            conn.execute(text(_DDL_ANUAL))
            conn.execute(text(_IDX_ANUAL))
            print(f"[ri] Tabla {_TABLE_ANUAL} creada.")
        if not insp.has_table(_TABLE_MENSUAL):
            conn.execute(text(_DDL_MENSUAL))
            conn.execute(text(_IDX_MENSUAL))
            print(f"[ri] Tabla {_TABLE_MENSUAL} creada.")


def _upsert(table: str, records: list[dict], key_cols: list[str], engine) -> None:
    """Inserta solo registros cuyo hash_registro no existe aun en la tabla."""
    if not records:
        return

    # Cargar hashes existentes
    with engine.connect() as conn:
        result = conn.execute(text(f"SELECT hash_registro FROM {table}"))
        existing = {row[0] for row in result}

    new = [r for r in records if r["hash_registro"] not in existing]
    if not new:
        print(f"[ri] [{table}] Sin registros nuevos.")
        return

    cols = list(new[0].keys())
    placeholders = ", ".join(f":{c}" for c in cols)
    col_list = ", ".join(cols)
    sql = text(f"INSERT INTO {table} ({col_list}) VALUES ({placeholders})")

    with engine.begin() as conn:
        conn.execute(sql, new)

    print(f"[ri] [{table}] {len(new)} nuevos registros insertados.")
