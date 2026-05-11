"""
ETL loader: Indice de Tipo de Cambio Real (ITCER) — BCE Ecuador

Hojas: una sola hoja "Base YYYY=100" (el anio base puede cambiar).

Estructura del Excel:
  Fila 2 : categorias (EFECTIVO REAL, BILATERAL REAL)
  Fila 3 : paises/indicadores por columna
  Filas 4-34 : bloque anual  (col A = entero: 1995, 1996, ..., 2025)
  Fila 35    : separador vacio
  Filas 36+  : bloque mensual
               col A = "YYYY Enero" al inicio de cada anio
               col A = "Febrero", "Marzo", ... para el resto de meses
               filas vacias entre bloques de 12 meses

Tablas:
  tipo_de_cambio_anual   — una fila por anio  (formato ancho, 17 indicadores)
  tipo_de_cambio_mensual — una fila por (anio, mes) (formato ancho + mes_num)

Columnas de indicadores (detectadas dinamicamente desde el Excel):
  itcer_efectivo_real, bilateral_eeuu, bilateral_china, bilateral_colombia,
  bilateral_mexico, bilateral_peru, bilateral_alemania, bilateral_panama,
  bilateral_espana, bilateral_japon, bilateral_brasil, bilateral_rusia,
  bilateral_corea_sur, bilateral_paises_bajos, bilateral_chile,
  bilateral_italia, bilateral_vietnam
"""

import hashlib
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl
from sqlalchemy import text

sys.path.append(str(Path(__file__).resolve().parents[2]))
from utils.base_engine import get_master_engine

_TABLE_ANUAL   = "tipo_de_cambio_anual"
_TABLE_MENSUAL = "tipo_de_cambio_mensual"

_MESES = {
    "Enero": 1, "Febrero": 2, "Marzo": 3, "Abril": 4,
    "Mayo": 5,  "Junio": 6,  "Julio": 7, "Agosto": 8,
    "Septiembre": 9, "Octubre": 10, "Noviembre": 11, "Diciembre": 12,
}

_COL_MAP = {
    "efectivo":       "itcer_efectivo_real",
    "estados unidos": "bilateral_eeuu",
    "china":          "bilateral_china",
    "colombia":       "bilateral_colombia",
    "mexico":         "bilateral_mexico",
    "peru":           "bilateral_peru",
    "alemania":       "bilateral_alemania",
    "panama":         "bilateral_panama",
    "espana":         "bilateral_espana",
    "japon":          "bilateral_japon",
    "brasil":         "bilateral_brasil",
    "rusia":          "bilateral_rusia",
    "corea":          "bilateral_corea_sur",
    "paises bajos":   "bilateral_paises_bajos",
    "holanda":        "bilateral_paises_bajos",
    "chile":          "bilateral_chile",
    "italia":         "bilateral_italia",
    "vietnam":        "bilateral_vietnam",
}

_INDICATOR_COLS = [
    "itcer_efectivo_real",
    "bilateral_eeuu", "bilateral_china", "bilateral_colombia",
    "bilateral_mexico", "bilateral_peru", "bilateral_alemania",
    "bilateral_panama", "bilateral_espana", "bilateral_japon",
    "bilateral_brasil", "bilateral_rusia", "bilateral_corea_sur",
    "bilateral_paises_bajos", "bilateral_chile", "bilateral_italia",
    "bilateral_vietnam",
]

_IND_COLS_DDL = "\n".join(
    f"    {col:<34} FLOAT NULL," for col in _INDICATOR_COLS
)

_DDL_ANUAL = f"""
CREATE TABLE tipo_de_cambio_anual (
    id                                 BIGINT IDENTITY(1,1) NOT NULL,
    anio                               INT    NOT NULL,
{_IND_COLS_DDL}
    hash_registro                      NVARCHAR(64) NOT NULL,
    fecha_carga                        DATETIME2 DEFAULT GETDATE(),
    CONSTRAINT PK_tipo_de_cambio_anual PRIMARY KEY NONCLUSTERED (id)
)"""

_IDX_ANUAL = "CREATE CLUSTERED INDEX CIX_tipo_de_cambio_anual ON tipo_de_cambio_anual (anio)"

_DDL_MENSUAL = f"""
CREATE TABLE tipo_de_cambio_mensual (
    id                                   BIGINT IDENTITY(1,1) NOT NULL,
    anio                                 INT     NOT NULL,
    mes_num                              TINYINT NOT NULL,
    mes                                  NVARCHAR(20) NOT NULL,
{_IND_COLS_DDL}
    hash_registro                        NVARCHAR(64) NOT NULL,
    fecha_carga                          DATETIME2 DEFAULT GETDATE(),
    CONSTRAINT PK_tipo_de_cambio_mensual PRIMARY KEY NONCLUSTERED (id)
)"""

_IDX_MENSUAL = "CREATE CLUSTERED INDEX CIX_tipo_de_cambio_mensual ON tipo_de_cambio_mensual (anio, mes_num)"


# ---------------------------------------------------------------------------
# Punto de entrada
# ---------------------------------------------------------------------------

def load(file: Path) -> None:
    if file is None or not file.exists():
        print("[tc] Archivo no encontrado, omitiendo carga.")
        return

    print(f"[tc] Procesando {file.name}...")
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)

    sheet_name = _find_sheet(wb)
    if not sheet_name:
        print("[tc] No se encontro hoja de datos.")
        wb.close()
        return

    print(f"[tc] Hoja: {sheet_name}")
    rows = list(wb[sheet_name].iter_rows(values_only=True))
    wb.close()

    col_map = _detect_columns(rows)
    print(f"[tc] Columnas detectadas: {len(col_map)} -> {list(col_map.values())}")

    engine = get_master_engine()
    _ensure_tables(engine)

    records_a = _parse_annual(rows, col_map)
    print(f"[tc] Anual: {len(records_a)} registros")
    _upsert(_TABLE_ANUAL, records_a, engine)

    records_m = _parse_monthly(rows, col_map)
    print(f"[tc] Mensual: {len(records_m)} registros")
    _upsert(_TABLE_MENSUAL, records_m, engine)


# ---------------------------------------------------------------------------
# Deteccion de hoja y columnas
# ---------------------------------------------------------------------------

def _find_sheet(wb) -> str | None:
    for name in wb.sheetnames:
        if "base" in name.lower() or "itcer" in name.lower():
            return name
    return wb.sheetnames[0] if wb.sheetnames else None


def _detect_columns(rows: list) -> dict[int, str]:
    """
    Combina filas 1 y 2 (indices) para mapear cada columna a su nombre en BD.
    Usa coincidencia por keyword para ser robusto ante cambios de nombre.
    """
    cat_row     = rows[1] if len(rows) > 1 else ()
    country_row = rows[2] if len(rows) > 2 else ()

    col_map: dict[int, str] = {}
    n_cols = max(len(cat_row), len(country_row))
    for ci in range(1, n_cols):
        cat     = _norm(cat_row[ci]     if ci < len(cat_row)     else None)
        country = _norm(country_row[ci] if ci < len(country_row) else None)
        db_col  = _match_col(cat) or _match_col(country) or _match_col(f"{cat} {country}")
        if db_col and db_col not in col_map.values():
            col_map[ci] = db_col

    return col_map


def _norm(v) -> str:
    if v is None:
        return ""
    # Quitar tildes y normalizar a ASCII antes de comparar
    s = unicodedata.normalize("NFD", str(v).lower())
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"[^\w\s]", "", s).strip()


def _match_col(text: str) -> str | None:
    for keyword, col in _COL_MAP.items():
        if keyword in text:
            return col
    return None


# ---------------------------------------------------------------------------
# Parsers
# ---------------------------------------------------------------------------

def _parse_annual(rows: list, col_map: dict) -> list[dict]:
    records = []
    for row in rows[3:]:
        year = _as_year(row[0] if row else None)
        if year is None:
            continue
        rec = {"anio": year}
        rec.update(_extract_values(row, col_map))
        rec["hash_registro"] = _hash(["anio"], rec)
        records.append(rec)
    return records


def _parse_monthly(rows: list, col_map: dict) -> list[dict]:
    records      = []
    current_year = None

    for row in rows[3:]:
        if not row or row[0] is None:
            continue
        period_str = str(row[0]).strip()

        # Ignorar filas anuales y notas al pie
        if _as_year(row[0]) is not None:
            continue
        if period_str.lower().startswith(("fuente", "nota", "(*")):
            break

        # "YYYY MesNombre" o solo "MesNombre"
        parts = period_str.split(" ", 1)
        if len(parts) == 2 and parts[0].isdigit() and parts[1] in _MESES:
            current_year = int(parts[0])
            mes_nombre   = parts[1]
        elif period_str in _MESES:
            mes_nombre = period_str
        else:
            continue

        if current_year is None:
            continue

        rec = {
            "anio":    current_year,
            "mes_num": _MESES[mes_nombre],
            "mes":     mes_nombre,
        }
        rec.update(_extract_values(row, col_map))
        rec["hash_registro"] = _hash(["anio", "mes_num"], rec)
        records.append(rec)

    return records


def _extract_values(row: tuple, col_map: dict) -> dict:
    result = {col: None for col in _INDICATOR_COLS}
    for ci, db_col in col_map.items():
        val = _to_float(row[ci] if ci < len(row) else None)
        if val is not None:
            result[db_col] = val
    return result


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _as_year(v) -> int | None:
    if v is None:
        return None
    try:
        f = float(str(v).strip())
        y = int(f)
        if 1900 <= y <= 2100 and f == y:
            return y
    except (ValueError, TypeError):
        pass
    return None


def _to_float(v) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _hash(key_cols: list, rec: dict) -> str:
    key = "|".join(str(rec.get(k, "")) for k in key_cols)
    return hashlib.sha256(key.encode()).hexdigest()


# ---------------------------------------------------------------------------
# DDL y carga
# ---------------------------------------------------------------------------

def _ensure_tables(engine) -> None:
    from sqlalchemy import inspect as sa_inspect
    insp = sa_inspect(engine)
    with engine.begin() as conn:
        if not insp.has_table(_TABLE_ANUAL):
            conn.execute(text(_DDL_ANUAL))
            conn.execute(text(_IDX_ANUAL))
            print(f"[tc] Tabla {_TABLE_ANUAL} creada.")
        if not insp.has_table(_TABLE_MENSUAL):
            conn.execute(text(_DDL_MENSUAL))
            conn.execute(text(_IDX_MENSUAL))
            print(f"[tc] Tabla {_TABLE_MENSUAL} creada.")


def _upsert(table: str, records: list[dict], engine) -> None:
    if not records:
        return

    with engine.connect() as conn:
        existing = {row[0] for row in conn.execute(
            text(f"SELECT hash_registro FROM {table}")
        )}

    new = [r for r in records if r["hash_registro"] not in existing]
    if not new:
        print(f"[tc] [{table}] Sin registros nuevos.")
        return

    cols         = list(new[0].keys())
    col_list     = ", ".join(cols)
    placeholders = ", ".join(f":{c}" for c in cols)
    with engine.begin() as conn:
        conn.execute(text(f"INSERT INTO {table} ({col_list}) VALUES ({placeholders})"), new)

    print(f"[tc] [{table}] {len(new)} nuevos registros insertados.")
